#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
except ImportError as exc:
    raise SystemExit("缺少 reportlab，请先安装：python -m pip install reportlab") from exc

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:160]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> list[str]:
        req = models.DescribeTableRequest()
        req.DatabaseName = db
        req.TableName = table
        resp = self.client.DescribeTable(req)
        data = json.loads(resp.to_json_string())
        table_info = data.get("Table", data)
        columns = table_info.get("Columns") or []
        names: list[str] = []
        for item in columns:
            if isinstance(item, dict) and item.get("Name"):
                names.append(item["Name"])
        return names


def fetch_single_value(runner: DlcRunner, sql: str, db: str, default: Any = 0) -> Any:
    rows = runner.exec_sql(sql, db)
    if not rows or not rows[0]:
        return default
    return rows[0][0]


def resolve_column_name(columns: set[str], candidates: list[str]) -> str | None:
    for name in candidates:
        if name in columns:
            return name
    return None


def non_empty_expr(column_name: str) -> str:
    return f"CASE WHEN `{column_name}` IS NULL OR TRIM(CAST(`{column_name}` AS STRING)) = '' THEN '无' ELSE TRIM(CAST(`{column_name}` AS STRING)) END"


def build_distribution_count_sql(table_fqn: str, field_expr: str) -> str:
    return f"""
    WITH agg AS (
        SELECT
            {field_expr} AS field_value,
            COUNT(*) AS row_count
        FROM {table_fqn}
        GROUP BY {field_expr}
    )
    SELECT COUNT(*) FROM agg
    """


def build_top_distribution_sql(table_fqn: str, field_expr: str, limit_n: int | None) -> str:
    limit_clause = f"\n    LIMIT {limit_n}" if limit_n else ""
    return f"""
    WITH agg AS (
        SELECT
            {field_expr} AS field_value,
            COUNT(*) AS row_count
        FROM {table_fqn}
        GROUP BY {field_expr}
    )
    SELECT
        field_value,
        row_count,
        ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS probability,
        ROUND(SUM(row_count) OVER (ORDER BY row_count DESC, field_value ASC) * 1.0 / SUM(row_count) OVER (), 6) AS cumulative_probability
    FROM agg
    ORDER BY row_count DESC, field_value ASC
    {limit_clause}
    """


def build_distribution_page_sql(table_fqn: str, field_expr: str, offset: int, page_size: int) -> str:
    return f"""
    WITH agg AS (
        SELECT
            {field_expr} AS field_value,
            COUNT(*) AS row_count
        FROM {table_fqn}
        GROUP BY {field_expr}
    ),
    ordered AS (
        SELECT
            field_value,
            row_count,
            ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS probability,
            ROUND(SUM(row_count) OVER (ORDER BY row_count DESC, field_value ASC) * 1.0 / SUM(row_count) OVER (), 6) AS cumulative_probability,
            ROW_NUMBER() OVER (ORDER BY row_count DESC, field_value ASC) AS rn
        FROM agg
    )
    SELECT
        field_value,
        row_count,
        probability,
        cumulative_probability
    FROM ordered
    WHERE rn > {offset} AND rn <= {offset + page_size}
    ORDER BY rn
    """


def build_cluster_count_sql(table_fqn: str, spu_col: str, mid_col: str, track_col: str) -> str:
    spu_expr = non_empty_expr(spu_col)
    mid_expr = non_empty_expr(mid_col)
    track_expr = non_empty_expr(track_col)
    return f"""
    WITH base AS (
        SELECT
            {spu_expr} AS spu_value,
            {mid_expr} AS mid_cate_value,
            {track_expr} AS track_value
        FROM {table_fqn}
    ),
    agg AS (
        SELECT
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            ) AS cluster_label,
            COUNT(*) AS row_count
        FROM base
        GROUP BY
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            )
    )
    SELECT COUNT(*) FROM agg
    """


def build_cluster_sql(table_fqn: str, spu_col: str, mid_col: str, track_col: str, limit_n: int | None) -> str:
    spu_expr = non_empty_expr(spu_col)
    mid_expr = non_empty_expr(mid_col)
    track_expr = non_empty_expr(track_col)
    limit_clause = f"\n    LIMIT {limit_n}" if limit_n else ""
    return f"""
    WITH base AS (
        SELECT
            {spu_expr} AS spu_value,
            {mid_expr} AS mid_cate_value,
            {track_expr} AS track_value
        FROM {table_fqn}
    ),
    agg AS (
        SELECT
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            ) AS cluster_label,
            COUNT(*) AS row_count
        FROM base
        GROUP BY
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            )
    )
    SELECT
        cluster_label,
        row_count,
        ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS probability,
        ROUND(SUM(row_count) OVER (ORDER BY row_count DESC, cluster_label ASC) * 1.0 / SUM(row_count) OVER (), 6) AS cumulative_probability
    FROM agg
    ORDER BY row_count DESC, cluster_label ASC
    {limit_clause}
    """


def build_cluster_page_sql(table_fqn: str, spu_col: str, mid_col: str, track_col: str, offset: int, page_size: int) -> str:
    spu_expr = non_empty_expr(spu_col)
    mid_expr = non_empty_expr(mid_col)
    track_expr = non_empty_expr(track_col)
    return f"""
    WITH base AS (
        SELECT
            {spu_expr} AS spu_value,
            {mid_expr} AS mid_cate_value,
            {track_expr} AS track_value
        FROM {table_fqn}
    ),
    agg AS (
        SELECT
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            ) AS cluster_label,
            COUNT(*) AS row_count
        FROM base
        GROUP BY
            concat_ws(
                ' | ',
                concat('spu=', spu_value),
                concat('mid_cate=', mid_cate_value),
                concat('track=', track_value)
            )
    ),
    ordered AS (
        SELECT
            cluster_label,
            row_count,
            ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS probability,
            ROUND(SUM(row_count) OVER (ORDER BY row_count DESC, cluster_label ASC) * 1.0 / SUM(row_count) OVER (), 6) AS cumulative_probability,
            ROW_NUMBER() OVER (ORDER BY row_count DESC, cluster_label ASC) AS rn
        FROM agg
    )
    SELECT
        cluster_label,
        row_count,
        probability,
        cumulative_probability
    FROM ordered
    WHERE rn > {offset} AND rn <= {offset + page_size}
    ORDER BY rn
    """


def dataframe_from_rows(rows: list[list[Any]], columns: list[str], numeric_cols: list[str] | None = None) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=columns)
    for col in numeric_cols or []:
        if col in df.columns and not df.empty:
            df[col] = pd.to_numeric(df[col])
    return df


def fetch_paged_distribution(
    runner: DlcRunner,
    db: str,
    table_fqn: str,
    field_expr: str,
    page_size: int = 800,
) -> pd.DataFrame:
    total_groups = int(fetch_single_value(runner, build_distribution_count_sql(table_fqn, field_expr), db, 0))
    all_rows: list[list[Any]] = []
    for offset in range(0, total_groups, page_size):
        all_rows.extend(runner.exec_sql(build_distribution_page_sql(table_fqn, field_expr, offset, page_size), db))
    return dataframe_from_rows(all_rows, ["field_value", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])


def fetch_paged_cluster_distribution(
    runner: DlcRunner,
    db: str,
    table_fqn: str,
    spu_col: str,
    mid_col: str,
    track_col: str,
    page_size: int = 800,
) -> pd.DataFrame:
    total_groups = int(fetch_single_value(runner, build_cluster_count_sql(table_fqn, spu_col, mid_col, track_col), db, 0))
    all_rows: list[list[Any]] = []
    for offset in range(0, total_groups, page_size):
        all_rows.extend(runner.exec_sql(build_cluster_page_sql(table_fqn, spu_col, mid_col, track_col, offset, page_size), db))
    return dataframe_from_rows(all_rows, ["cluster_label", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])


def analyze_cluster_table(
    runner: DlcRunner,
    db: str,
    table: str,
    label: str,
    spu_candidates: list[str],
    mid_candidates: list[str],
    track_candidates: list[str],
    id_candidates: list[str],
    distribution_limit: int | None,
    cluster_limit: int | None,
) -> dict[str, Any]:
    raw_columns = runner.describe_table(db, table)
    column_set = set(raw_columns)
    spu_col = resolve_column_name(column_set, spu_candidates)
    mid_col = resolve_column_name(column_set, mid_candidates)
    track_col = resolve_column_name(column_set, track_candidates)
    id_col = resolve_column_name(column_set, id_candidates)

    summary: dict[str, Any] = {
        "label": label,
        "table": f"{db}.{table}",
        "resolved_columns": {
            "spu": spu_col,
            "mid_cate": mid_col,
            "track": track_col,
            "id": id_col,
        },
        "raw_columns": raw_columns,
        "distribution_limit": distribution_limit,
        "cluster_limit": cluster_limit,
    }

    table_fqn = f"{db}.{table}"
    summary["total_rows"] = int(fetch_single_value(runner, f"SELECT COUNT(*) FROM {table_fqn}", db, 0))
    if spu_col:
        summary["distinct_spu"] = int(
            fetch_single_value(
                runner,
                f"SELECT COUNT(DISTINCT `{spu_col}`) FROM {table_fqn} WHERE `{spu_col}` IS NOT NULL AND TRIM(CAST(`{spu_col}` AS STRING)) <> ''",
                db,
                0,
            )
        )
    else:
        summary["distinct_spu"] = None
    if id_col:
        summary["distinct_material_id"] = int(
            fetch_single_value(
                runner,
                f"SELECT COUNT(DISTINCT `{id_col}`) FROM {table_fqn} WHERE `{id_col}` IS NOT NULL AND TRIM(CAST(`{id_col}` AS STRING)) <> ''",
                db,
                0,
            )
        )
    else:
        summary["distinct_material_id"] = None

    distributions: dict[str, pd.DataFrame] = {}
    if spu_col:
        if distribution_limit is None:
            distributions["spu"] = fetch_paged_distribution(runner, db, table_fqn, non_empty_expr(spu_col))
        else:
            rows = runner.exec_sql(build_top_distribution_sql(table_fqn, non_empty_expr(spu_col), distribution_limit), db)
            distributions["spu"] = dataframe_from_rows(rows, ["field_value", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])
    if mid_col:
        if distribution_limit is None:
            distributions["mid_cate"] = fetch_paged_distribution(runner, db, table_fqn, non_empty_expr(mid_col))
        else:
            rows = runner.exec_sql(build_top_distribution_sql(table_fqn, non_empty_expr(mid_col), distribution_limit), db)
            distributions["mid_cate"] = dataframe_from_rows(rows, ["field_value", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])
    if track_col:
        if distribution_limit is None:
            distributions["track"] = fetch_paged_distribution(runner, db, table_fqn, non_empty_expr(track_col))
        else:
            rows = runner.exec_sql(build_top_distribution_sql(table_fqn, non_empty_expr(track_col), distribution_limit), db)
            distributions["track"] = dataframe_from_rows(rows, ["field_value", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])

    if spu_col and mid_col and track_col:
        if cluster_limit is None:
            cluster_df = fetch_paged_cluster_distribution(runner, db, table_fqn, spu_col, mid_col, track_col)
        else:
            cluster_rows = runner.exec_sql(build_cluster_sql(table_fqn, spu_col, mid_col, track_col, cluster_limit), db)
            cluster_df = dataframe_from_rows(cluster_rows, ["cluster_label", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])
    else:
        cluster_df = pd.DataFrame(columns=["cluster_label", "row_count", "probability", "cumulative_probability"])
    summary["distinct_clusters"] = int(len(cluster_df.index))
    summary["cluster_probability_sum"] = float(cluster_df["probability"].sum()) if not cluster_df.empty else 0.0
    summary["distribution_cardinality"] = {key: int(len(df.index)) for key, df in distributions.items()}

    return {
        "summary": summary,
        "cluster_df": cluster_df,
        "distributions": distributions,
    }


def analyze_tb16(runner: DlcRunner, top_n: int) -> dict[str, Any]:
    db = "data_dim"
    table = "tb16_dim_product_sale_dimension"
    raw_columns = runner.describe_table(db, table)
    column_set = set(raw_columns)
    field_candidates = {
        "产品名称": ["产品名称"],
        "品牌大类": ["品牌大类"],
        "品牌中类": ["品牌中类", "中类"],
        "细分赛道": ["抖音成人商品_细分赛道", "子赛道"],
        "性别": ["性别_李宁bi", "性别内衬BI"],
        "风格": ["风格"],
        "适用场景": ["适用场景", "场景"],
    }
    summary: dict[str, Any] = {
        "table": f"{db}.{table}",
        "total_rows": int(fetch_single_value(runner, f"SELECT COUNT(*) FROM {db}.{table}", db, 0)),
        "distinct_spu": int(
            fetch_single_value(
                runner,
                f"SELECT COUNT(DISTINCT spu) FROM {db}.{table} WHERE spu IS NOT NULL AND TRIM(CAST(spu AS STRING)) <> ''",
                db,
                0,
            )
        ),
        "resolved_fields": {},
        "missing_fields": [],
        "top_n": top_n,
    }
    distributions: dict[str, pd.DataFrame] = {}
    for logical_name, candidates in field_candidates.items():
        resolved = resolve_column_name(column_set, candidates)
        if not resolved:
            summary["missing_fields"].append({"logical_name": logical_name, "candidates": candidates})
            continue
        summary["resolved_fields"][logical_name] = resolved
        rows = runner.exec_sql(build_top_distribution_sql(f"{db}.{table}", non_empty_expr(resolved), top_n), db)
        distributions[logical_name] = dataframe_from_rows(rows, ["field_value", "row_count", "probability", "cumulative_probability"], ["row_count", "probability", "cumulative_probability"])
    return {"summary": summary, "distributions": distributions}


def compute_distribution_stats(df: pd.DataFrame) -> dict[str, Any]:
    if df.empty:
        return {
            "distinct_values": 0,
            "top1_label": None,
            "top1_probability": 0.0,
            "top10_cumulative_probability": 0.0,
        }
    top1 = df.iloc[0]
    top10_prob = float(df.head(10)["probability"].sum()) if "probability" in df.columns else 0.0
    label_col = "field_value" if "field_value" in df.columns else "cluster_label"
    return {
        "distinct_values": int(len(df.index)),
        "top1_label": str(top1[label_col]),
        "top1_probability": float(top1["probability"]),
        "top10_cumulative_probability": top10_prob,
    }


def column_letter_to_index(letter: str) -> int:
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def add_excel_bar_chart(sheet, value_col: str, label_col: str, title: str) -> None:
    max_row = sheet.max_row
    if max_row <= 1:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "类别"
    chart.x_axis.title = "概率"
    chart.height = max(7, min(18, 0.42 * (max_row - 1) + 4))
    chart.width = 18
    data = Reference(sheet, min_col=column_letter_to_index(value_col), min_row=1, max_row=max_row)
    cats = Reference(sheet, min_col=column_letter_to_index(label_col), min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    sheet.add_chart(chart, "F2")


def write_excel(payload: dict[str, Any], output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_rows = [
            ["报告时间", payload["generated_at"]],
        ]
        for section_key in ["short_video", "picture_material", "tb16"]:
            section = payload[section_key]
            summary = section["summary"]
            prefix = section["summary"]["table"]
            summary_rows.append([f"{prefix} 总行数", summary.get("total_rows")])
            if "distinct_spu" in summary:
                summary_rows.append([f"{prefix} 唯一SPU", summary.get("distinct_spu")])
            if summary.get("distinct_material_id") is not None:
                summary_rows.append([f"{prefix} 唯一素材ID", summary.get("distinct_material_id")])
            if summary.get("distinct_clusters") is not None:
                summary_rows.append([f"{prefix} 聚类组合数", summary.get("distinct_clusters")])
            if summary.get("cluster_probability_sum") is not None:
                summary_rows.append([f"{prefix} 聚类分布概率和", summary.get("cluster_probability_sum")])
        pd.DataFrame(summary_rows, columns=["metric", "value"]).to_excel(writer, sheet_name="summary", index=False)

        sheet_chart_map: list[tuple[str, str, str, str]] = []

        for section_key, cluster_sheet_name in [("short_video", "short_video_cluster"), ("picture_material", "picture_cluster")]:
            section = payload[section_key]
            cluster_df = pd.DataFrame(section["cluster_distribution"])
            cluster_df.to_excel(writer, sheet_name=cluster_sheet_name, index=False)
            sheet_chart_map.append((cluster_sheet_name, "C", "A", f"{section['summary']['table']} 聚类Top概率"))
            for dist_name, records in section["distributions"].items():
                df = pd.DataFrame(records)
                sheet_name = f"{section_key}_{dist_name}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheet_chart_map.append((sheet_name, "C", "A", f"{section['summary']['table']} {dist_name} 分布"))

        tb16 = payload["tb16"]
        for logical_name, records in tb16["distributions"].items():
            df = pd.DataFrame(records)
            sheet_name = f"tb16_{logical_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet_chart_map.append((sheet_name, "C", "A", f"tb16 {logical_name} Top分布"))

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:120]]
                width = min(max(len(value) for value in values) + 2, 90)
                sheet.column_dimensions[column_cells[0].column_letter].width = width
        for sheet_name, value_col, label_col, title in sheet_chart_map:
            add_excel_bar_chart(workbook[sheet_name], value_col, label_col, title)


def write_markdown(payload: dict[str, Any], output_path: Path) -> None:
    lines: list[str] = []
    lines.append("# 短视频与图片 SPU/中类/赛道聚类报告")
    lines.append("")
    lines.append(f"生成时间：{payload['generated_at']}")
    lines.append("")
    lines.append("说明：")
    lines.append("- `dws_short_video_product_info` 与 `dim_picture_material_data_enriched` 已按全量分布统计，不再截 Top20。")
    lines.append("- 由于 `spu` 与聚类组合基数较大，PDF/Markdown 正文只展示摘要和部分样例；全量明细见 JSON / Excel。")
    lines.append("- 概率分布图已内嵌在 Excel 中。")
    lines.append("")

    for section_key, title in [("short_video", "一、dws_short_video_product_info 聚类分析"), ("picture_material", "二、dim_picture_material_data_enriched 聚类分析")]:
        section = payload[section_key]
        summary = section["summary"]
        lines.append(f"## {title}")
        lines.append("")
        lines.append(f"- 表：`{summary['table']}`")
        lines.append(f"- 总行数：{summary['total_rows']}")
        lines.append(f"- 唯一 SPU 数：{summary.get('distinct_spu')}")
        if summary.get("distinct_material_id") is not None:
            lines.append(f"- 唯一素材 ID 数：{summary.get('distinct_material_id')}")
        lines.append(f"- 聚类字段：`spu={summary['resolved_columns'].get('spu')}`，`mid_cate={summary['resolved_columns'].get('mid_cate')}`，`track={summary['resolved_columns'].get('track')}`")
        lines.append(f"- 聚类组合数：{summary.get('distinct_clusters')}")
        lines.append(f"- `spu` 分布基数：{summary.get('distribution_cardinality', {}).get('spu')}")
        lines.append(f"- `mid_cate` 分布基数：{summary.get('distribution_cardinality', {}).get('mid_cate')}")
        lines.append(f"- `track` 分布基数：{summary.get('distribution_cardinality', {}).get('track')}")
        lines.append("")
        cluster_df = pd.DataFrame(section["cluster_distribution"])
        cluster_stats = compute_distribution_stats(cluster_df)
        lines.append("### 全量聚类分布摘要")
        lines.append("")
        lines.append(f"- 聚类组合总数：{cluster_stats['distinct_values']}")
        lines.append(f"- Top1 聚类：`{cluster_stats['top1_label']}`，占比 {cluster_stats['top1_probability']:.2%}")
        lines.append(f"- Top10 聚类累计占比：{cluster_stats['top10_cumulative_probability']:.2%}")
        lines.append("")
        lines.append("### 聚类样例（前20）")
        lines.append("")
        lines.append("| cluster_label | row_count | probability | cumulative_probability |")
        lines.append("|---|---:|---:|---:|")
        for row in section["cluster_distribution"][:20]:
            lines.append(f"| {row['cluster_label']} | {int(row['row_count'])} | {float(row['probability']):.2%} | {float(row['cumulative_probability']):.2%} |")
        lines.append("")
        for dist_name, records in section["distributions"].items():
            df = pd.DataFrame(records)
            stats = compute_distribution_stats(df)
            lines.append(f"### {dist_name} 全量分布摘要")
            lines.append("")
            lines.append(f"- 不同取值数：{stats['distinct_values']}")
            lines.append(f"- Top1：`{stats['top1_label']}`，占比 {stats['top1_probability']:.2%}")
            lines.append(f"- Top10 累计占比：{stats['top10_cumulative_probability']:.2%}")
            lines.append("")
            sample_rows = records if dist_name in ("mid_cate", "track") else records[:20]
            lines.append(f"### {dist_name} 分布样例")
            lines.append("")
            lines.append("| field_value | row_count | probability | cumulative_probability |")
            lines.append("|---|---:|---:|---:|")
            for row in sample_rows:
                lines.append(f"| {row['field_value']} | {int(row['row_count'])} | {float(row['probability']):.2%} | {float(row['cumulative_probability']):.2%} |")
            lines.append("")

    tb16 = payload["tb16"]
    lines.append("## 三、tb16_dim_product_sale_dimension 重点维度 Top20")
    lines.append("")
    lines.append(f"- 表：`{tb16['summary']['table']}`")
    lines.append(f"- 总行数：{tb16['summary']['total_rows']}")
    lines.append(f"- 唯一 SPU 数：{tb16['summary']['distinct_spu']}")
    lines.append("")
    if tb16["summary"]["missing_fields"]:
        lines.append("### 未命中的候选字段")
        lines.append("")
        for item in tb16["summary"]["missing_fields"]:
            lines.append(f"- `{item['logical_name']}` 未命中候选列：`{'`, `'.join(item['candidates'])}`")
        lines.append("")
    for logical_name, records in tb16["distributions"].items():
        resolved = tb16["summary"]["resolved_fields"].get(logical_name, logical_name)
        lines.append(f"### {logical_name} Top20")
        lines.append("")
        lines.append(f"- 实际字段：`{resolved}`")
        lines.append("")
        lines.append("| field_value | row_count | probability | cumulative_probability |")
        lines.append("|---|---:|---:|---:|")
        for row in records:
            lines.append(f"| {row['field_value']} | {int(row['row_count'])} | {float(row['probability']):.2%} | {float(row['cumulative_probability']):.2%} |")
        lines.append("")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def truncate_text(text: Any, limit: int = 80) -> str:
    raw = str(text) if text is not None else ""
    return raw if len(raw) <= limit else raw[: limit - 3] + "..."


def build_pdf_table(rows: list[list[Any]], col_widths: list[float], style) -> LongTable:
    table = LongTable(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(style)
    return table


def write_pdf(payload: dict[str, Any], output_path: Path) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCN", parent=styles["Title"], fontName="STSong-Light", fontSize=18, leading=24)
    body_style = ParagraphStyle("BodyCN", parent=styles["BodyText"], fontName="STSong-Light", fontSize=10, leading=14)
    small_style = ParagraphStyle("SmallCN", parent=styles["BodyText"], fontName="STSong-Light", fontSize=8.5, leading=12)
    heading_style = ParagraphStyle("HeadingCN", parent=styles["Heading2"], fontName="STSong-Light", fontSize=13, leading=18)
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9AA5B1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F4F7FB")]),
        ]
    )

    story: list[Any] = []
    story.append(Paragraph("短视频与图片 SPU/中类/赛道全量分布报告", title_style))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"生成时间：{payload['generated_at']}", body_style))
    story.append(Paragraph("本报告基于全量分布统计生成。为控制 PDF 体积，正文展示全量统计摘要与样例，全量明细见同目录 JSON / Excel。", body_style))
    story.append(Spacer(1, 4 * mm))

    for section_key, title in [("short_video", "一、短视频表全量分布"), ("picture_material", "二、图片表全量分布")]:
        section = payload[section_key]
        summary = section["summary"]
        story.append(Paragraph(title, heading_style))
        story.append(Paragraph(f"表：{summary['table']}", body_style))
        story.append(Paragraph(f"总行数：{summary['total_rows']}；唯一 SPU：{summary.get('distinct_spu')}；聚类组合数：{summary.get('distinct_clusters')}", body_style))
        story.append(Paragraph(f"聚类字段：spu={summary['resolved_columns'].get('spu')}，mid_cate={summary['resolved_columns'].get('mid_cate')}，track={summary['resolved_columns'].get('track')}", body_style))
        story.append(Spacer(1, 2 * mm))

        summary_rows = [["指标", "不同取值数", "Top1", "Top1占比", "Top10累计占比"]]
        cluster_df = pd.DataFrame(section["cluster_distribution"])
        cluster_stats = compute_distribution_stats(cluster_df)
        summary_rows.append([
            "cluster",
            cluster_stats["distinct_values"],
            truncate_text(cluster_stats["top1_label"], 48),
            f"{cluster_stats['top1_probability']:.2%}",
            f"{cluster_stats['top10_cumulative_probability']:.2%}",
        ])
        for dist_name, records in section["distributions"].items():
            stats = compute_distribution_stats(pd.DataFrame(records))
            summary_rows.append([
                dist_name,
                stats["distinct_values"],
                truncate_text(stats["top1_label"], 48),
                f"{stats['top1_probability']:.2%}",
                f"{stats['top10_cumulative_probability']:.2%}",
            ])
        story.append(build_pdf_table(summary_rows, [26 * mm, 22 * mm, 78 * mm, 20 * mm, 24 * mm], table_style))
        story.append(Spacer(1, 3 * mm))

        cluster_sample_rows = [["cluster_label", "row_count", "probability", "cumulative_probability"]]
        for row in section["cluster_distribution"][:20]:
            cluster_sample_rows.append([
                truncate_text(row["cluster_label"], 88),
                int(row["row_count"]),
                f"{float(row['probability']):.2%}",
                f"{float(row['cumulative_probability']):.2%}",
            ])
        story.append(Paragraph("聚类样例（前20）", body_style))
        story.append(build_pdf_table(cluster_sample_rows, [110 * mm, 22 * mm, 22 * mm, 28 * mm], table_style))
        story.append(Spacer(1, 3 * mm))

        for dist_name in ["mid_cate", "track"]:
            records = section["distributions"].get(dist_name, [])
            if not records:
                continue
            dist_rows = [["field_value", "row_count", "probability", "cumulative_probability"]]
            for row in records:
                dist_rows.append([
                    truncate_text(row["field_value"], 52),
                    int(row["row_count"]),
                    f"{float(row['probability']):.2%}",
                    f"{float(row['cumulative_probability']):.2%}",
                ])
            story.append(Paragraph(f"{dist_name} 全量分布", body_style))
            story.append(build_pdf_table(dist_rows, [90 * mm, 24 * mm, 24 * mm, 28 * mm], table_style))
            story.append(Spacer(1, 3 * mm))

        spu_records = section["distributions"].get("spu", [])
        if spu_records:
            story.append(Paragraph("spu 分布样例（前50，完整清单见 Excel/JSON）", body_style))
            spu_rows = [["field_value", "row_count", "probability", "cumulative_probability"]]
            for row in spu_records[:50]:
                spu_rows.append([
                    truncate_text(row["field_value"], 28),
                    int(row["row_count"]),
                    f"{float(row['probability']):.2%}",
                    f"{float(row['cumulative_probability']):.2%}",
                ])
            story.append(build_pdf_table(spu_rows, [52 * mm, 24 * mm, 24 * mm, 28 * mm], table_style))
        story.append(PageBreak())

    tb16 = payload["tb16"]
    tb16_summary = tb16["summary"]
    story.append(Paragraph("三、tb16_dim_product_sale_dimension 重点维度 Top20", heading_style))
    story.append(Paragraph(f"表：{tb16_summary['table']}", body_style))
    story.append(Paragraph(f"总行数：{tb16_summary['total_rows']}；唯一 SPU：{tb16_summary['distinct_spu']}", body_style))
    story.append(Spacer(1, 2 * mm))
    for logical_name, records in tb16["distributions"].items():
        story.append(Paragraph(f"{logical_name} Top20", body_style))
        rows = [["field_value", "row_count", "probability", "cumulative_probability"]]
        for row in records:
            rows.append([
                truncate_text(row["field_value"], 48),
                int(row["row_count"]),
                f"{float(row['probability']):.2%}",
                f"{float(row['cumulative_probability']):.2%}",
            ])
        story.append(build_pdf_table(rows, [86 * mm, 24 * mm, 24 * mm, 28 * mm], table_style))
        story.append(Spacer(1, 3 * mm))

    doc = SimpleDocTemplate(str(output_path), pagesize=A4, leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    doc.build(story)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成短视频/图片 SPU-中类-赛道聚类报告")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--tb16-top-n", type=int, default=20, help="tb16 重点维度 TopN")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    date_tag = now_tag()

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), args.max_wait)

    print("[1/3] 统计 dws_short_video_product_info 聚类与分布...", flush=True)
    short_video = analyze_cluster_table(
        runner,
        db="data_dws",
        table="dws_short_video_product_info",
        label="short_video",
        spu_candidates=["spu"],
        mid_candidates=["mid_cate", "reference_mid_cate"],
        track_candidates=["track", "reference_track"],
        id_candidates=["short_video_id", "file_id"],
        distribution_limit=None,
        cluster_limit=20,
    )

    print("[2/3] 统计 dim_picture_material_data_enriched 聚类与分布...", flush=True)
    picture_material = analyze_cluster_table(
        runner,
        db="data_dim",
        table="dim_picture_material_data_enriched",
        label="picture_material",
        spu_candidates=["spu"],
        mid_candidates=["mid_cate", "reference_mid_cate"],
        track_candidates=["track_first_li_ning_bi", "reference_track"],
        id_candidates=["platform_image_id", "file_id"],
        distribution_limit=None,
        cluster_limit=20,
    )

    print("[3/3] 统计 tb16_dim_product_sale_dimension 重点维度分布...", flush=True)
    tb16 = analyze_tb16(runner, args.tb16_top_n)

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "short_video": {
            "summary": short_video["summary"],
            "cluster_distribution": short_video["cluster_df"].to_dict(orient="records"),
            "distributions": {key: df.to_dict(orient="records") for key, df in short_video["distributions"].items()},
        },
        "picture_material": {
            "summary": picture_material["summary"],
            "cluster_distribution": picture_material["cluster_df"].to_dict(orient="records"),
            "distributions": {key: df.to_dict(orient="records") for key, df in picture_material["distributions"].items()},
        },
        "tb16": {
            "summary": tb16["summary"],
            "distributions": {key: df.to_dict(orient="records") for key, df in tb16["distributions"].items()},
        },
    }

    md_path = output_dir / f"spu-mid-track-cluster-report-{date_tag}.md"
    json_path = output_dir / f"spu-mid-track-cluster-report-{date_tag}.json"
    xlsx_path = output_dir / f"spu-mid-track-cluster-report-{date_tag}.xlsx"
    pdf_path = output_dir / f"spu-mid-track-cluster-report-{date_tag}.pdf"
    write_markdown(payload, md_path)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(payload, xlsx_path)
    write_pdf(payload, pdf_path)

    print("报告已生成：")
    print(md_path)
    print(json_path)
    print(xlsx_path)
    print(pdf_path)


if __name__ == "__main__":
    main()
