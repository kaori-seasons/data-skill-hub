#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_JSON_NAME = "sample-002-gender-gap-report-{date_tag}.json"
DEFAULT_MD_NAME = "sample-002-gender-gap-report-{date_tag}.md"
DEFAULT_XLSX_NAME = "sample-002-gender-gap-report-{date_tag}.xlsx"


@dataclass
class ReportPayload:
    generated_at: str
    target_table: str
    sql_file: str
    target_gender_distribution: list[dict[str, Any]]
    tb16_gender_distribution: list[dict[str, Any]]
    tb16_lining_bi_distribution: list[dict[str, Any]]
    gender_wu_root_cause_summary: list[dict[str, Any]]
    gender_wu_samples: list[dict[str, Any]]
    key_findings: list[str]


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def to_int(value: Any) -> int:
    try:
        return int(str(value))
    except Exception:
        return 0


def value_at(rows: list[list[Any]] | None, row: int = 0, col: int = 0, default: Any = None) -> Any:
    try:
        return rows[row][col]
    except Exception:
        return default


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:180]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")


def apply_sheet_style(ws) -> None:
    align = Alignment(vertical="top", horizontal="left", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 80))
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, min(max_len + 2, 80))


def write_xlsx(path: Path, tables: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        for ws in writer.book.worksheets:
            apply_sheet_style(ws)


def norm_expr(column_name: str) -> str:
    return f"COALESCE(NULLIF(TRIM(CAST({column_name} AS STRING)), ''), '<<EMPTY>>')"


def fetch_target_gender_distribution(runner: DlcRunner) -> pd.DataFrame:
    sql = f"""
    SELECT
        {norm_expr('gender')} AS gender,
        COUNT(*) AS row_cnt,
        COUNT(DISTINCT UPPER(TRIM(CAST(spu AS STRING)))) AS spu_cnt
    FROM data_dwd.dwd_file_label_id_spu
    GROUP BY {norm_expr('gender')}
    ORDER BY row_cnt DESC, gender
    """
    rows = runner.exec_sql(sql, "data_dwd")
    return pd.DataFrame(rows, columns=["gender", "row_cnt", "spu_cnt"])


def fetch_tb16_distribution(runner: DlcRunner, field_expr: str, alias: str) -> pd.DataFrame:
    sql = f"""
    WITH spu_level AS (
        SELECT
            UPPER(TRIM(CAST(spu AS STRING))) AS spu,
            {norm_expr(field_expr)} AS gender_value
        FROM data_dim.tb16_dim_product_sale_dimension
        WHERE spu IS NOT NULL
          AND TRIM(CAST(spu AS STRING)) <> ''
    )
    SELECT
        gender_value AS {alias},
        COUNT(*) AS row_cnt,
        COUNT(DISTINCT spu) AS spu_cnt
    FROM spu_level
    GROUP BY gender_value
    ORDER BY row_cnt DESC, {alias}
    """
    rows = runner.exec_sql(sql, "data_dim")
    return pd.DataFrame(rows, columns=[alias, "row_cnt", "spu_cnt"])


def build_gender_lineage_cte() -> str:
    return """
    WITH target AS (
        SELECT
            TRIM(CAST(file_id AS STRING)) AS file_id,
            UPPER(TRIM(CAST(spu AS STRING))) AS spu,
            TRIM(CAST(gender AS STRING)) AS final_gender
        FROM data_dwd.dwd_file_label_id_spu
        WHERE file_id IS NOT NULL
          AND TRIM(CAST(file_id AS STRING)) <> ''
          AND spu IS NOT NULL
          AND TRIM(CAST(spu AS STRING)) <> ''
    ),
    tb16_spu AS (
        SELECT
            UPPER(TRIM(CAST(spu AS STRING))) AS spu,
            MAX(TRIM(CAST(`性别` AS STRING))) AS gender_tb16,
            MAX(TRIM(CAST(`性别_李宁bi` AS STRING))) AS gender_lining_bi
        FROM data_dim.tb16_dim_product_sale_dimension
        WHERE spu IS NOT NULL
          AND TRIM(CAST(spu AS STRING)) <> ''
        GROUP BY UPPER(TRIM(CAST(spu AS STRING)))
    ),
    joined AS (
        SELECT
            t.file_id,
            t.spu,
            COALESCE(NULLIF(TRIM(CAST(t.final_gender AS STRING)), ''), '<<EMPTY>>') AS final_gender,
            COALESCE(NULLIF(TRIM(CAST(b.gender_tb16 AS STRING)), ''), '<<EMPTY>>') AS gender_tb16,
            COALESCE(NULLIF(TRIM(CAST(b.gender_lining_bi AS STRING)), ''), '<<EMPTY>>') AS gender_lining_bi,
            CASE
                WHEN b.spu IS NULL THEN 'NO_TB16_MATCH'
                WHEN COALESCE(NULLIF(TRIM(CAST(b.gender_lining_bi AS STRING)), ''), '') = ''
                     AND COALESCE(NULLIF(TRIM(CAST(b.gender_tb16 AS STRING)), ''), '') <> ''
                    THEN 'TB16_GENDER_PRESENT_BUT_LINING_BI_EMPTY'
                WHEN COALESCE(NULLIF(TRIM(CAST(b.gender_lining_bi AS STRING)), ''), '') = ''
                     AND COALESCE(NULLIF(TRIM(CAST(b.gender_tb16 AS STRING)), ''), '') = ''
                    THEN 'TB16_BOTH_EMPTY'
                WHEN TRIM(CAST(b.gender_lining_bi AS STRING)) = '无'
                    THEN 'LINING_BI_IS_WU'
                WHEN TRIM(CAST(b.gender_lining_bi AS STRING)) NOT IN (
                    '中', '中性', '女', '女性', '男', '男性', '女童', '男童', '婴幼儿', '童'
                )
                    THEN 'UNMAPPED_LINING_BI_ENUM'
                ELSE 'OTHER'
            END AS root_cause
        FROM target t
        LEFT JOIN tb16_spu b
          ON t.spu = b.spu
        WHERE COALESCE(NULLIF(TRIM(CAST(t.final_gender AS STRING)), ''), '') = '无'
    )
    """


def fetch_gender_wu_summary(runner: DlcRunner) -> pd.DataFrame:
    sql = f"""
    {build_gender_lineage_cte()}
    SELECT
        root_cause,
        COUNT(*) AS row_cnt,
        COUNT(DISTINCT spu) AS spu_cnt
    FROM joined
    GROUP BY root_cause
    ORDER BY row_cnt DESC, root_cause
    """
    rows = runner.exec_sql(sql, "data_dwd")
    return pd.DataFrame(rows, columns=["root_cause", "row_cnt", "spu_cnt"])


def fetch_gender_wu_samples(runner: DlcRunner, sample_limit: int) -> pd.DataFrame:
    summary_df = fetch_gender_wu_summary(runner)
    sample_rows: list[list[Any]] = []
    for root_cause in summary_df["root_cause"].astype(str).tolist():
        sql = f"""
        {build_gender_lineage_cte()}
        SELECT
            root_cause,
            file_id,
            spu,
            final_gender,
            gender_tb16,
            gender_lining_bi
        FROM joined
        WHERE root_cause = '{root_cause}'
        LIMIT {sample_limit}
        """
        rows = runner.exec_sql(sql, "data_dwd")
        sample_rows.extend(rows)
    return pd.DataFrame(
        sample_rows,
        columns=["root_cause", "file_id", "spu", "final_gender", "gender_tb16", "gender_lining_bi"],
    )


def build_findings(
    target_df: pd.DataFrame,
    tb16_gender_df: pd.DataFrame,
    tb16_lining_df: pd.DataFrame,
    root_df: pd.DataFrame,
) -> list[str]:
    findings: list[str] = []
    target_wu_rows = 0
    target_total_rows = int(target_df["row_cnt"].astype(int).sum()) if not target_df.empty else 0
    target_wu_match = target_df.loc[target_df["gender"] == "无", "row_cnt"]
    if not target_wu_match.empty:
        target_wu_rows = int(target_wu_match.iloc[0])
    findings.append(
        f"目标表 data_dwd.dwd_file_label_id_spu 中 gender='无' 共 {target_wu_rows} 行，占目标表 {round(target_wu_rows / target_total_rows * 100, 4) if target_total_rows else 0}% 。"
    )

    tb16_gender_empty = 0
    tb16_gender_match = tb16_gender_df.loc[tb16_gender_df["gender_tb16"] == "<<EMPTY>>", "spu_cnt"]
    if not tb16_gender_match.empty:
        tb16_gender_empty = int(tb16_gender_match.iloc[0])
    tb16_lining_empty = 0
    tb16_lining_match = tb16_lining_df.loc[tb16_lining_df["gender_lining_bi"] == "<<EMPTY>>", "spu_cnt"]
    if not tb16_lining_match.empty:
        tb16_lining_empty = int(tb16_lining_match.iloc[0])
    findings.append(
        f"tb16 按 SPU 看，`性别` 空值 SPU 数为 {tb16_gender_empty}，`性别_李宁bi` 空值 SPU 数为 {tb16_lining_empty}。如果后者显著更高，说明 sample-002 当前取错了性别来源字段。"
    )

    if not root_df.empty:
        top = root_df.iloc[0]
        findings.append(
            f"`gender='无'` 的主因是 {top['root_cause']}，涉及 {int(top['row_cnt'])} 行、{int(top['spu_cnt'])} 个 SPU。"
        )
        causes = set(root_df["root_cause"].astype(str).tolist())
        if "TB16_GENDER_PRESENT_BUT_LINING_BI_EMPTY" in causes:
            findings.append("存在大量样本在 tb16.性别 有值，但 tb16.性别_李宁bi 为空，说明当前 SQL 取 `性别_李宁bi` 会把本可映射的 SPU 误写成 `无`。")
        if "NO_TB16_MATCH" in causes:
            findings.append("存在一批 `dwd_file_label_id_spu` 的 SPU 在 tb16 完全对不上，这部分即使改映射规则也仍会落成 `无`，需要补查上游 SPU 对齐。")
    return findings


def write_md(path: Path, payload: ReportPayload) -> None:
    lines: list[str] = []
    lines.append("# sample-002 性别字段异常检测报告")
    lines.append("")
    lines.append(f"- 生成时间：{payload.generated_at}")
    lines.append(f"- SQL 文件：`{payload.sql_file}`")
    lines.append(f"- 目标表：`{payload.target_table}`")
    lines.append("")
    lines.append("## 一、关键结论")
    lines.append("")
    for item in payload.key_findings:
        lines.append(f"- {item}")
    lines.append("")
    lines.append("## 二、目标表 gender 分布")
    lines.append("")
    lines.append("| gender | row_cnt | spu_cnt |")
    lines.append("| --- | ---: | ---: |")
    for row in payload.target_gender_distribution:
        lines.append(f"| {row['gender']} | {row['row_cnt']} | {row['spu_cnt']} |")
    lines.append("")
    lines.append("## 三、tb16 原始字段分布")
    lines.append("")
    lines.append("### 3.1 `性别`")
    lines.append("")
    lines.append("| gender_tb16 | row_cnt | spu_cnt |")
    lines.append("| --- | ---: | ---: |")
    for row in payload.tb16_gender_distribution:
        lines.append(f"| {row['gender_tb16']} | {row['row_cnt']} | {row['spu_cnt']} |")
    lines.append("")
    lines.append("### 3.2 `性别_李宁bi`")
    lines.append("")
    lines.append("| gender_lining_bi | row_cnt | spu_cnt |")
    lines.append("| --- | ---: | ---: |")
    for row in payload.tb16_lining_bi_distribution:
        lines.append(f"| {row['gender_lining_bi']} | {row['row_cnt']} | {row['spu_cnt']} |")
    lines.append("")
    lines.append("## 四、`gender=无` 根因拆解")
    lines.append("")
    lines.append("| root_cause | row_cnt | spu_cnt |")
    lines.append("| --- | ---: | ---: |")
    for row in payload.gender_wu_root_cause_summary:
        lines.append(f"| {row['root_cause']} | {row['row_cnt']} | {row['spu_cnt']} |")
    lines.append("")
    lines.append("## 五、抽样")
    lines.append("")
    lines.append("| root_cause | file_id | spu | final_gender | gender_tb16 | gender_lining_bi |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for row in payload.gender_wu_samples:
        lines.append(
            f"| {row['root_cause']} | {row['file_id']} | {row['spu']} | {row['final_gender']} | "
            f"{row['gender_tb16']} | {row['gender_lining_bi']} |"
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="检测 sample-002.sql 落表后的 gender='无' 异常根因。")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--sample-limit", type=int, default=10, help="每类根因抽样条数")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最大等待秒数")
    args = parser.parse_args()

    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    runner = DlcRunner(os.getenv("DLC_USER"), os.getenv("DLC_PASSWORD"), max_wait=args.max_wait)

    print("[1/4] 统计目标表 gender 分布...", flush=True)
    target_df = fetch_target_gender_distribution(runner)
    print("[2/4] 统计 tb16 性别字段分布...", flush=True)
    tb16_gender_df = fetch_tb16_distribution(runner, "`性别`", "gender_tb16")
    tb16_lining_df = fetch_tb16_distribution(runner, "`性别_李宁bi`", "gender_lining_bi")
    print("[3/4] 拆解 gender='无' 的根因...", flush=True)
    root_df = fetch_gender_wu_summary(runner)
    print("[4/4] 抽样核对异常样本...", flush=True)
    sample_df = fetch_gender_wu_samples(runner, args.sample_limit)

    payload = ReportPayload(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        target_table="data_dwd.dwd_file_label_id_spu",
        sql_file="/Users/windwheel/.copaw/workspaces/sample-002.sql",
        target_gender_distribution=target_df.to_dict(orient="records"),
        tb16_gender_distribution=tb16_gender_df.to_dict(orient="records"),
        tb16_lining_bi_distribution=tb16_lining_df.to_dict(orient="records"),
        gender_wu_root_cause_summary=root_df.to_dict(orient="records"),
        gender_wu_samples=sample_df.to_dict(orient="records"),
        key_findings=build_findings(target_df, tb16_gender_df, tb16_lining_df, root_df),
    )

    date_tag = now_tag()
    json_path = output_dir / DEFAULT_JSON_NAME.format(date_tag=date_tag)
    md_path = output_dir / DEFAULT_MD_NAME.format(date_tag=date_tag)
    xlsx_path = output_dir / DEFAULT_XLSX_NAME.format(date_tag=date_tag)

    json_path.write_text(json.dumps(asdict(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    write_md(md_path, payload)
    write_xlsx(
        xlsx_path,
        {
            "target_gender": target_df,
            "tb16_gender": tb16_gender_df,
            "tb16_lining_bi": tb16_lining_df,
            "root_cause": root_df,
            "samples": sample_df,
        },
    )

    print(f"[OK] JSON: {json_path}")
    print(f"[OK] MD:   {md_path}")
    print(f"[OK] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
