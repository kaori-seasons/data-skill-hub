#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    Alignment = None
    Font = None
    PatternFill = None
    HAS_OPENPYXL = False

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_DB = "data_dwd"
DEFAULT_TABLE = "dwd_platform_order_detail_di"
DEFAULT_REPORT_PREFIX = "order-live-room-spu-distribution-probe"

FIELD_CANDIDATES: dict[str, list[str]] = {
    "shop_id": ["erp_shop_id", "shop_id", "store_id", "seller_shop_id"],
    "shop_name": ["shop_name", "shop_nick", "store_name", "seller_shop_name"],
    "product_id": ["product_id", "goods_id", "item_id"],
    "spu": ["spu", "spu_id"],
    "sku_id": ["sku_id", "sku", "item_sku_id"],
    "child_order_id": ["child_order_id", "child_order_no", "sub_order_id", "order_detail_id", "order_item_id"],
    "live_room_id": ["live_room_id", "live_room", "room_id"],
    "pay_time": ["pay_time", "payment_time", "order_pay_time", "paid_time"],
}

MAPPING_PAIRS: list[tuple[str, str, str]] = [
    ("shop_id_to_shop_name", "shop_id", "shop_name"),
    ("product_id_to_spu", "product_id", "spu"),
    ("spu_to_sku_id", "spu", "sku_id"),
    ("child_order_id_to_spu", "child_order_id", "spu"),
    ("child_order_id_to_live_room_id", "child_order_id", "live_room_id"),
    ("live_room_id_to_pay_date", "live_room_id", "pay_date"),
    ("live_room_id_to_spu", "live_room_id", "spu"),
    ("spu_to_live_room_id", "spu", "live_room_id"),
]


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def to_int(value: Any) -> int:
    try:
        return int(str(value))
    except Exception:
        return 0


def to_float(value: Any) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def dataframe_from_rows(rows: list[list[Any]], columns: list[str], numeric_columns: list[str] | None = None) -> pd.DataFrame:
    frame = pd.DataFrame(rows, columns=columns)
    for column in numeric_columns or []:
        if column in frame.columns:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def normalize_expr(column_name: str, upper: bool = False) -> str:
    raw = f"TRIM(CAST(`{column_name}` AS STRING))"
    value_expr = f"UPPER({raw})" if upper else raw
    return f"CASE WHEN `{column_name}` IS NULL OR {raw} = '' THEN NULL ELSE {value_expr} END"


def build_time_prefix_expr(column_name: str, length: int) -> str:
    raw = f"TRIM(CAST(`{column_name}` AS STRING))"
    return f"CASE WHEN `{column_name}` IS NULL OR {raw} = '' THEN NULL ELSE SUBSTR({raw}, 1, {length}) END"


def bucket_case_expr(metric_name: str, chinese_day: bool = False) -> str:
    if chinese_day:
        return f"""
        CASE
            WHEN {metric_name} = 1 THEN '1天'
            WHEN {metric_name} BETWEEN 2 AND 3 THEN '2-3天'
            WHEN {metric_name} BETWEEN 4 AND 7 THEN '4-7天'
            WHEN {metric_name} BETWEEN 8 AND 15 THEN '8-15天'
            ELSE '16天以上'
        END
        """
    return f"""
    CASE
        WHEN {metric_name} = 1 THEN '1'
        WHEN {metric_name} BETWEEN 2 AND 5 THEN '2-5'
        WHEN {metric_name} BETWEEN 6 AND 10 THEN '6-10'
        WHEN {metric_name} BETWEEN 11 AND 20 THEN '11-20'
        ELSE '21+'
    END
    """


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 600):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:180]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> list[str]:
        req = models.DescribeTableRequest()
        req.DatabaseName = db
        req.TableName = table
        resp = self.client.DescribeTable(req)
        data = json.loads(resp.to_json_string())
        table_info = data.get("Table", data)
        columns = table_info.get("Columns") or []
        names: list[str] = []
        for item in columns:
            if isinstance(item, dict) and item.get("Name"):
                names.append(item["Name"])
        return names


def resolve_columns(raw_columns: list[str]) -> tuple[dict[str, str | None], list[dict[str, Any]]]:
    column_set = set(raw_columns)
    resolved: dict[str, str | None] = {}
    missing: list[dict[str, Any]] = []
    for logical_name, candidates in FIELD_CANDIDATES.items():
        hit = None
        for candidate in candidates:
            if candidate in column_set:
                hit = candidate
                break
        resolved[logical_name] = hit
        if not hit:
            missing.append({"logical_name": logical_name, "candidates": candidates})
    resolved["pay_date"] = resolved.get("pay_time")
    resolved["pay_hour"] = resolved.get("pay_time")
    return resolved, missing


def build_base_cte(table_fqn: str, resolved: dict[str, str | None], where_clause: str | None) -> str:
    select_items = {
        "shop_id": normalize_expr(resolved["shop_id"]) if resolved.get("shop_id") else "CAST(NULL AS STRING)",
        "shop_name": normalize_expr(resolved["shop_name"]) if resolved.get("shop_name") else "CAST(NULL AS STRING)",
        "product_id": normalize_expr(resolved["product_id"]) if resolved.get("product_id") else "CAST(NULL AS STRING)",
        "spu": normalize_expr(resolved["spu"], upper=True) if resolved.get("spu") else "CAST(NULL AS STRING)",
        "sku_id": normalize_expr(resolved["sku_id"], upper=True) if resolved.get("sku_id") else "CAST(NULL AS STRING)",
        "child_order_id": normalize_expr(resolved["child_order_id"]) if resolved.get("child_order_id") else "CAST(NULL AS STRING)",
        "live_room_id": normalize_expr(resolved["live_room_id"]) if resolved.get("live_room_id") else "CAST(NULL AS STRING)",
        "pay_time": normalize_expr(resolved["pay_time"]) if resolved.get("pay_time") else "CAST(NULL AS STRING)",
        "pay_date": build_time_prefix_expr(resolved["pay_time"], 10) if resolved.get("pay_time") else "CAST(NULL AS STRING)",
        "pay_hour": build_time_prefix_expr(resolved["pay_time"], 13) if resolved.get("pay_time") else "CAST(NULL AS STRING)",
    }
    where_sql = f"\n        AND ({where_clause})" if where_clause else ""
    lines = [
        "WITH base AS (",
        "    SELECT",
    ]
    aliases = list(select_items.keys())
    for index, alias in enumerate(aliases):
        suffix = "," if index < len(aliases) - 1 else ""
        lines.append(f"        {select_items[alias]} AS {alias}{suffix}")
    lines.extend(
        [
            f"    FROM {table_fqn}",
            "    WHERE 1 = 1" + where_sql,
            ")",
        ]
    )
    return "\n".join(lines)


def build_field_profile_sql(base_cte: str, resolved: dict[str, str | None]) -> str:
    fields = [
        ("shop_id", resolved.get("shop_id")),
        ("shop_name", resolved.get("shop_name")),
        ("product_id", resolved.get("product_id")),
        ("spu", resolved.get("spu")),
        ("sku_id", resolved.get("sku_id")),
        ("child_order_id", resolved.get("child_order_id")),
        ("live_room_id", resolved.get("live_room_id")),
        ("pay_time", resolved.get("pay_time")),
        ("pay_date", resolved.get("pay_time")),
    ]
    sql_parts: list[str] = [base_cte]
    union_parts: list[str] = []
    for logical_name, physical_column in fields:
        physical = physical_column or "未命中"
        union_parts.append(
            f"""
            SELECT
                '{logical_name}' AS logical_field,
                '{physical}' AS physical_column,
                COUNT(*) AS total_rows,
                SUM(CASE WHEN {logical_name} IS NULL THEN 1 ELSE 0 END) AS null_rows,
                ROUND(SUM(CASE WHEN {logical_name} IS NULL THEN 1 ELSE 0 END) * 1.0 / COUNT(*), 6) AS null_ratio,
                COUNT(DISTINCT {logical_name}) AS distinct_values
            FROM base
            """
        )
    sql_parts.append("\nUNION ALL\n".join(union_parts))
    sql_parts.append("ORDER BY logical_field")
    return "\n".join(sql_parts)


def query_field_profile(runner: DlcRunner, db: str, base_cte: str, resolved: dict[str, str | None]) -> pd.DataFrame:
    rows = runner.exec_sql(build_field_profile_sql(base_cte, resolved), db)
    return dataframe_from_rows(
        rows,
        ["logical_field", "physical_column", "total_rows", "null_rows", "null_ratio", "distinct_values"],
        ["total_rows", "null_rows", "null_ratio", "distinct_values"],
    )


def build_mapping_summary_sql(base_cte: str, source_field: str, target_field: str) -> str:
    return f"""
    {base_cte},
    pair_stats AS (
        SELECT
            {source_field} AS source_value,
            COUNT(*) AS row_count,
            COUNT(DISTINCT {target_field}) AS target_distinct_cnt
        FROM base
        WHERE {source_field} IS NOT NULL
          AND {target_field} IS NOT NULL
        GROUP BY {source_field}
    )
    SELECT
        COUNT(*) AS source_key_count,
        SUM(row_count) AS covered_rows,
        ROUND(AVG(target_distinct_cnt), 4) AS avg_target_per_source,
        percentile_approx(target_distinct_cnt, 0.5) AS p50_target_per_source,
        percentile_approx(target_distinct_cnt, 0.9) AS p90_target_per_source,
        MAX(target_distinct_cnt) AS max_target_per_source,
        SUM(CASE WHEN target_distinct_cnt > 1 THEN 1 ELSE 0 END) AS multi_mapping_source_keys,
        ROUND(SUM(CASE WHEN target_distinct_cnt > 1 THEN 1 ELSE 0 END) * 1.0 / COUNT(*), 6) AS multi_mapping_ratio
    FROM pair_stats
    """


def build_mapping_bucket_sql(base_cte: str, source_field: str, target_field: str, chinese_day: bool = False) -> str:
    bucket_expr = bucket_case_expr("target_distinct_cnt", chinese_day=chinese_day)
    return f"""
    {base_cte},
    pair_stats AS (
        SELECT
            {source_field} AS source_value,
            COUNT(DISTINCT {target_field}) AS target_distinct_cnt
        FROM base
        WHERE {source_field} IS NOT NULL
          AND {target_field} IS NOT NULL
        GROUP BY {source_field}
    ),
    bucketed AS (
        SELECT
            {bucket_expr} AS target_bucket
        FROM pair_stats
    )
    SELECT
        target_bucket,
        COUNT(*) AS source_key_count,
        ROUND(COUNT(*) * 1.0 / SUM(COUNT(*)) OVER (), 6) AS source_key_ratio
    FROM bucketed
    GROUP BY target_bucket
    ORDER BY
        CASE target_bucket
            WHEN '1' THEN 1
            WHEN '2-5' THEN 2
            WHEN '6-10' THEN 3
            WHEN '11-20' THEN 4
            WHEN '21+' THEN 5
            WHEN '1天' THEN 1
            WHEN '2-3天' THEN 2
            WHEN '4-7天' THEN 3
            WHEN '8-15天' THEN 4
            ELSE 5
        END
    """


def build_mapping_sample_sql(
    base_cte: str,
    source_field: str,
    target_field: str,
    sample_limit: int,
    target_sample_limit: int,
) -> str:
    return f"""
    {base_cte},
    pair_stats AS (
        SELECT
            {source_field} AS source_value,
            COUNT(*) AS row_count,
            COUNT(DISTINCT {target_field}) AS target_distinct_cnt,
            CONCAT_WS(' | ', SLICE(SORT_ARRAY(COLLECT_SET({target_field})), 1, {target_sample_limit})) AS target_samples
        FROM base
        WHERE {source_field} IS NOT NULL
          AND {target_field} IS NOT NULL
        GROUP BY {source_field}
    )
    SELECT
        source_value,
        row_count,
        target_distinct_cnt,
        target_samples
    FROM pair_stats
    WHERE target_distinct_cnt > 1
    ORDER BY target_distinct_cnt DESC, row_count DESC, source_value ASC
    LIMIT {sample_limit}
    """


def query_mapping_check(
    runner: DlcRunner,
    db: str,
    base_cte: str,
    pair_name: str,
    source_field: str,
    target_field: str,
    sample_limit: int,
    target_sample_limit: int,
    quick_summary: bool = False,
) -> dict[str, Any]:
    summary_rows = runner.exec_sql(build_mapping_summary_sql(base_cte, source_field, target_field), db)
    summary_df = dataframe_from_rows(
        summary_rows,
        [
            "source_key_count",
            "covered_rows",
            "avg_target_per_source",
            "p50_target_per_source",
            "p90_target_per_source",
            "max_target_per_source",
            "multi_mapping_source_keys",
            "multi_mapping_ratio",
        ],
        [
            "source_key_count",
            "covered_rows",
            "avg_target_per_source",
            "p50_target_per_source",
            "p90_target_per_source",
            "max_target_per_source",
            "multi_mapping_source_keys",
            "multi_mapping_ratio",
        ],
    )
    bucket_df = pd.DataFrame(columns=["target_bucket", "source_key_count", "source_key_ratio"])
    sample_df = pd.DataFrame(columns=["source_value", "row_count", "target_distinct_cnt", "target_samples"])
    if not quick_summary:
        bucket_rows = runner.exec_sql(
            build_mapping_bucket_sql(base_cte, source_field, target_field, chinese_day=(target_field == "pay_date")),
            db,
        )
        sample_rows = runner.exec_sql(
            build_mapping_sample_sql(base_cte, source_field, target_field, sample_limit, target_sample_limit),
            db,
        )
        bucket_df = dataframe_from_rows(
            bucket_rows,
            ["target_bucket", "source_key_count", "source_key_ratio"],
            ["source_key_count", "source_key_ratio"],
        )
        sample_df = dataframe_from_rows(
            sample_rows,
            ["source_value", "row_count", "target_distinct_cnt", "target_samples"],
            ["row_count", "target_distinct_cnt"],
        )
    summary_record = summary_df.iloc[0].to_dict() if not summary_df.empty else {}
    return {
        "pair_name": pair_name,
        "source_field": source_field,
        "target_field": target_field,
        "summary": summary_record,
        "bucket_distribution": bucket_df,
        "multi_mapping_samples": sample_df,
    }


def build_live_room_daily_preview_sql(base_cte: str, limit_n: int) -> str:
    return f"""
    {base_cte}
    SELECT
        live_room_id,
        COALESCE(pay_date, '无支付日期') AS pay_date,
        COUNT(*) AS row_count,
        COUNT(DISTINCT child_order_id) AS child_order_count,
        COUNT(DISTINCT spu) AS spu_count,
        COUNT(DISTINCT product_id) AS product_id_count,
        COUNT(DISTINCT sku_id) AS sku_id_count,
        COUNT(DISTINCT shop_id) AS shop_id_count
    FROM base
    WHERE live_room_id IS NOT NULL
    GROUP BY live_room_id, COALESCE(pay_date, '无支付日期')
    ORDER BY child_order_count DESC, row_count DESC, live_room_id ASC
    LIMIT {limit_n}
    """


def build_spu_preview_sql(base_cte: str, limit_n: int) -> str:
    return f"""
    {base_cte}
    SELECT
        spu,
        COUNT(*) AS row_count,
        COUNT(DISTINCT child_order_id) AS child_order_count,
        COUNT(DISTINCT sku_id) AS sku_id_count,
        COUNT(DISTINCT product_id) AS product_id_count,
        COUNT(DISTINCT live_room_id) AS live_room_count,
        COUNT(DISTINCT pay_date) AS pay_date_count,
        COUNT(DISTINCT shop_id) AS shop_id_count
    FROM base
    WHERE spu IS NOT NULL
    GROUP BY spu
    ORDER BY child_order_count DESC, live_room_count DESC, row_count DESC, spu ASC
    LIMIT {limit_n}
    """


def build_overview_sql(base_cte: str) -> str:
    return f"""
    {base_cte}
    SELECT
        COUNT(*) AS total_rows,
        COUNT(DISTINCT shop_id) AS distinct_shop_id,
        COUNT(DISTINCT product_id) AS distinct_product_id,
        COUNT(DISTINCT spu) AS distinct_spu,
        COUNT(DISTINCT sku_id) AS distinct_sku_id,
        COUNT(DISTINCT child_order_id) AS distinct_child_order_id,
        COUNT(DISTINCT live_room_id) AS distinct_live_room_id,
        MIN(pay_time) AS min_pay_time,
        MAX(pay_time) AS max_pay_time
    FROM base
    """


def query_overview(runner: DlcRunner, db: str, base_cte: str) -> dict[str, Any]:
    rows = runner.exec_sql(build_overview_sql(base_cte), db)
    if not rows:
        return {}
    row = rows[0]
    return {
        "total_rows": to_int(row[0]),
        "distinct_shop_id": to_int(row[1]),
        "distinct_product_id": to_int(row[2]),
        "distinct_spu": to_int(row[3]),
        "distinct_sku_id": to_int(row[4]),
        "distinct_child_order_id": to_int(row[5]),
        "distinct_live_room_id": to_int(row[6]),
        "min_pay_time": row[7],
        "max_pay_time": row[8],
    }


def query_live_room_preview(runner: DlcRunner, db: str, base_cte: str, limit_n: int) -> pd.DataFrame:
    rows = runner.exec_sql(build_live_room_daily_preview_sql(base_cte, limit_n), db)
    return dataframe_from_rows(
        rows,
        [
            "live_room_id",
            "pay_date",
            "row_count",
            "child_order_count",
            "spu_count",
            "product_id_count",
            "sku_id_count",
            "shop_id_count",
        ],
        ["row_count", "child_order_count", "spu_count", "product_id_count", "sku_id_count", "shop_id_count"],
    )


def query_spu_preview(runner: DlcRunner, db: str, base_cte: str, limit_n: int) -> pd.DataFrame:
    rows = runner.exec_sql(build_spu_preview_sql(base_cte, limit_n), db)
    return dataframe_from_rows(
        rows,
        [
            "spu",
            "row_count",
            "child_order_count",
            "sku_id_count",
            "product_id_count",
            "live_room_count",
            "pay_date_count",
            "shop_id_count",
        ],
        ["row_count", "child_order_count", "sku_id_count", "product_id_count", "live_room_count", "pay_date_count", "shop_id_count"],
    )


def ratio_from_mapping(mapping_checks: dict[str, dict[str, Any]], pair_name: str) -> float:
    value = mapping_checks.get(pair_name, {}).get("summary", {}).get("multi_mapping_ratio")
    return to_float(value)


def max_from_mapping(mapping_checks: dict[str, dict[str, Any]], pair_name: str) -> int:
    value = mapping_checks.get(pair_name, {}).get("summary", {}).get("max_target_per_source")
    return to_int(value)


def build_key_findings(
    missing_fields: list[dict[str, Any]],
    mapping_checks: dict[str, dict[str, Any]],
    overview: dict[str, Any],
) -> list[str]:
    findings: list[str] = []
    missing_names = {item["logical_name"] for item in missing_fields}
    if "live_room_id" in missing_names:
        findings.append("订单表未命中直播间字段，当前无法直接按直播间维度梳理场次。")
    if "pay_time" in missing_names:
        findings.append("订单表未命中支付时间字段，当前无法验证“直播间ID是否足以代表场次”，需要补时间字段或补场次维表。")

    live_room_pay_date_ratio = ratio_from_mapping(mapping_checks, "live_room_id_to_pay_date")
    if live_room_pay_date_ratio > 0:
        findings.append(
            f"`live_room_id -> pay_date` 存在多值映射，约 {live_room_pay_date_ratio:.2%} 的直播间跨多个支付日，说明直播间ID不能直接等价于单一场次。"
        )

    product_spu_ratio = ratio_from_mapping(mapping_checks, "product_id_to_spu")
    if product_spu_ratio > 0:
        findings.append(
            f"`product_id -> spu` 存在多值映射，约 {product_spu_ratio:.2%} 的 product_id 对应多个 SPU，需求里必须明确以哪个字段作为商品主键。"
        )

    spu_sku_ratio = ratio_from_mapping(mapping_checks, "spu_to_sku_id")
    if spu_sku_ratio > 0:
        findings.append(
            f"`spu -> sku_id` 明显是一对多，约 {spu_sku_ratio:.2%} 的 SPU 对应多个 SKU，SPU维度汇总时不能直接按订单行数代替场次数。"
        )

    child_order_room_ratio = ratio_from_mapping(mapping_checks, "child_order_id_to_live_room_id")
    if child_order_room_ratio > 0:
        findings.append(
            f"`child_order_id -> live_room_id` 存在异常多值映射，约 {child_order_room_ratio:.2%} 的子订单命中多个直播间，需要先确认订单明细去重口径。"
        )

    spu_room_ratio = ratio_from_mapping(mapping_checks, "spu_to_live_room_id")
    spu_room_max = max_from_mapping(mapping_checks, "spu_to_live_room_id")
    if spu_room_ratio > 0:
        findings.append(
            f"`spu -> live_room_id` 存在多值映射，约 {spu_room_ratio:.2%} 的 SPU 出现在多个直播间，SPU维度场次建议统计 distinct 场次键，而不是直接 count(order rows)。"
        )
    if spu_room_max > 20:
        findings.append(f"单个 SPU 最多命中 {spu_room_max} 个直播间，长尾商品跨场次复用较明显，汇总逻辑需要先定义去重规则。")

    if overview.get("distinct_child_order_id", 0) == 0:
        findings.append("订单表未探测到有效子订单号，后续无法稳定以 child_order_id 作为最细事实粒度。")
    return findings


def safe_sheet_name(name: str) -> str:
    cleaned = name.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("[", "_").replace("]", "_")
    return cleaned[:31]


def write_excel(
    output_path: Path,
    overview: dict[str, Any],
    resolved: dict[str, str | None],
    missing_fields: list[dict[str, Any]],
    field_profile: pd.DataFrame,
    mapping_checks: dict[str, dict[str, Any]],
    live_room_preview: pd.DataFrame,
    spu_preview: pd.DataFrame,
    key_findings: list[str],
) -> None:
    if not HAS_OPENPYXL:
        print("[WARN] 未安装 openpyxl，跳过 Excel 输出。", flush=True)
        return
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_rows = [
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["总行数", overview.get("total_rows")],
            ["唯一店铺ID数", overview.get("distinct_shop_id")],
            ["唯一product_id数", overview.get("distinct_product_id")],
            ["唯一SPU数", overview.get("distinct_spu")],
            ["唯一SKU数", overview.get("distinct_sku_id")],
            ["唯一子订单数", overview.get("distinct_child_order_id")],
            ["唯一直播间数", overview.get("distinct_live_room_id")],
            ["最早支付时间", overview.get("min_pay_time")],
            ["最晚支付时间", overview.get("max_pay_time")],
        ]
        if key_findings:
            for index, finding in enumerate(key_findings, start=1):
                summary_rows.append([f"关键发现{index}", finding])
        pd.DataFrame(summary_rows, columns=["metric", "value"]).to_excel(writer, sheet_name="summary", index=False)

        resolved_rows = [{"logical_field": key, "physical_column": value or "未命中"} for key, value in resolved.items()]
        pd.DataFrame(resolved_rows).to_excel(writer, sheet_name="resolved_columns", index=False)
        pd.DataFrame(missing_fields or [{"logical_name": "无", "candidates": ""}]).to_excel(writer, sheet_name="missing_fields", index=False)
        field_profile.to_excel(writer, sheet_name="field_profile", index=False)
        live_room_preview.to_excel(writer, sheet_name="live_room_preview", index=False)
        spu_preview.to_excel(writer, sheet_name="spu_preview", index=False)

        for pair_name, content in mapping_checks.items():
            bucket_sheet = safe_sheet_name(f"{pair_name}_bucket")
            sample_sheet = safe_sheet_name(f"{pair_name}_sample")
            summary_sheet = safe_sheet_name(f"{pair_name}_summary")
            pd.DataFrame([content["summary"]]).to_excel(writer, sheet_name=summary_sheet, index=False)
            content["bucket_distribution"].to_excel(writer, sheet_name=bucket_sheet, index=False)
            content["multi_mapping_samples"].to_excel(writer, sheet_name=sample_sheet, index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:100]]
                width = min(max(len(value) for value in values) + 2, 80)
                sheet.column_dimensions[column_cells[0].column_letter].width = width


def write_markdown(
    output_path: Path,
    db: str,
    table: str,
    where_clause: str | None,
    raw_columns: list[str],
    resolved: dict[str, str | None],
    missing_fields: list[dict[str, Any]],
    overview: dict[str, Any],
    field_profile: pd.DataFrame,
    mapping_checks: dict[str, dict[str, Any]],
    live_room_preview: pd.DataFrame,
    spu_preview: pd.DataFrame,
    key_findings: list[str],
) -> None:
    lines: list[str] = []
    lines.append("# 订单表直播间/SPU映射分布探查报告")
    lines.append("")
    lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## 一、探查范围")
    lines.append("")
    lines.append(f"- 表：`{db}.{table}`")
    lines.append(f"- 过滤条件：`{where_clause or '无，全表扫描'}`")
    lines.append(f"- 原始字段数：{len(raw_columns)}")
    lines.append("")
    lines.append("## 二、字段命中情况")
    lines.append("")
    for logical_name, physical_name in resolved.items():
        if logical_name in ("pay_date", "pay_hour"):
            continue
        lines.append(f"- `{logical_name}` -> `{physical_name or '未命中'}`")
    if missing_fields:
        lines.append("")
        lines.append("未命中候选字段：")
        for item in missing_fields:
            lines.append(f"- `{item['logical_name']}` 候选列：`{'`, `'.join(item['candidates'])}`")
    lines.append("")
    lines.append("## 三、概览")
    lines.append("")
    lines.append(f"- 总行数：{overview.get('total_rows')}")
    lines.append(f"- 唯一店铺ID数：{overview.get('distinct_shop_id')}")
    lines.append(f"- 唯一 product_id 数：{overview.get('distinct_product_id')}")
    lines.append(f"- 唯一 SPU 数：{overview.get('distinct_spu')}")
    lines.append(f"- 唯一 SKU 数：{overview.get('distinct_sku_id')}")
    lines.append(f"- 唯一子订单数：{overview.get('distinct_child_order_id')}")
    lines.append(f"- 唯一直播间数：{overview.get('distinct_live_room_id')}")
    lines.append(f"- 支付时间范围：`{overview.get('min_pay_time')}` ~ `{overview.get('max_pay_time')}`")
    lines.append("")
    lines.append("## 四、字段分布")
    lines.append("")
    lines.append("| logical_field | physical_column | total_rows | null_rows | null_ratio | distinct_values |")
    lines.append("|---|---|---:|---:|---:|---:|")
    for row in field_profile.itertuples(index=False):
        lines.append(
            f"| {row.logical_field} | {row.physical_column} | {int(row.total_rows)} | {int(row.null_rows)} | {float(row.null_ratio):.2%} | {int(row.distinct_values)} |"
        )
    lines.append("")
    lines.append("## 五、关键发现")
    lines.append("")
    if key_findings:
        for finding in key_findings:
            lines.append(f"- {finding}")
    else:
        lines.append("- 当前未发现明显的多值映射风险，但仍建议结合业务确认“场次键”定义。")
    lines.append("")
    lines.append("## 六、核心映射关系")
    lines.append("")
    for pair_name, content in mapping_checks.items():
        summary = content["summary"]
        lines.append(f"### {pair_name}")
        lines.append("")
        lines.append(f"- 映射：`{content['source_field']} -> {content['target_field']}`")
        lines.append(f"- 参与映射的 source key 数：{to_int(summary.get('source_key_count'))}")
        lines.append(f"- 覆盖订单行数：{to_int(summary.get('covered_rows'))}")
        lines.append(f"- 平均每个 source 对应 target 数：{to_float(summary.get('avg_target_per_source')):.4f}")
        lines.append(f"- p50 / p90 / max：{to_int(summary.get('p50_target_per_source'))} / {to_int(summary.get('p90_target_per_source'))} / {to_int(summary.get('max_target_per_source'))}")
        lines.append(f"- 多值映射 source 占比：{to_float(summary.get('multi_mapping_ratio')):.2%}")
        lines.append("")
        lines.append("| target_bucket | source_key_count | source_key_ratio |")
        lines.append("|---|---:|---:|")
        for row in content["bucket_distribution"].itertuples(index=False):
            lines.append(f"| {row.target_bucket} | {int(row.source_key_count)} | {float(row.source_key_ratio):.2%} |")
        lines.append("")
        if not content["multi_mapping_samples"].empty:
            lines.append("多值映射样例：")
            lines.append("")
            lines.append("| source_value | row_count | target_distinct_cnt | target_samples |")
            lines.append("|---|---:|---:|---|")
            for row in content["multi_mapping_samples"].head(10).itertuples(index=False):
                lines.append(f"| {row.source_value} | {int(row.row_count)} | {int(row.target_distinct_cnt)} | {row.target_samples} |")
            lines.append("")
    lines.append("## 七、直播间维度预览")
    lines.append("")
    lines.append("说明：这里按 `live_room_id + pay_date` 展示预览，用于判断直播间ID是否足以代表单场。")
    lines.append("")
    lines.append("| live_room_id | pay_date | row_count | child_order_count | spu_count | product_id_count | sku_id_count | shop_id_count |")
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|")
    for row in live_room_preview.head(20).itertuples(index=False):
        lines.append(
            f"| {row.live_room_id} | {row.pay_date} | {int(row.row_count)} | {int(row.child_order_count)} | {int(row.spu_count)} | {int(row.product_id_count)} | {int(row.sku_id_count)} | {int(row.shop_id_count)} |"
        )
    lines.append("")
    lines.append("## 八、SPU维度预览")
    lines.append("")
    lines.append("| spu | row_count | child_order_count | sku_id_count | product_id_count | live_room_count | pay_date_count | shop_id_count |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---:|")
    for row in spu_preview.head(20).itertuples(index=False):
        lines.append(
            f"| {row.spu} | {int(row.row_count)} | {int(row.child_order_count)} | {int(row.sku_id_count)} | {int(row.product_id_count)} | {int(row.live_room_count)} | {int(row.pay_date_count)} | {int(row.shop_id_count)} |"
        )
    lines.append("")
    lines.append("## 九、需求梳理建议")
    lines.append("")
    lines.append("- 若 `live_room_id -> pay_date` 或更细时间粒度存在明显多值映射，应优先补“场次维表”或定义 `直播间 + 时间窗` 为场次键。")
    lines.append("- 若 `product_id -> spu` 不是稳定 1:1，直播间维度和 SPU 维度的统计口径必须分别定义，不能互相替代。")
    lines.append("- 若 `child_order_id -> live_room_id` 不是稳定 1:1，需要先确认订单明细是否存在重复归因或回流补数。")
    lines.append("- 若 `spu -> live_room_id` 明显一对多，SPU维度的“场次数”应统计 distinct 场次键，而不是 count(distinct child_order_id)。")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="探查订单表在直播间/SPU维度下的映射关系与场次分布风险")
    parser.add_argument("--db", default=DEFAULT_DB, help="DLC 数据库名")
    parser.add_argument("--table", default=DEFAULT_TABLE, help="DLC 表名")
    parser.add_argument("--schema-only", action="store_true", help="仅输出表字段并退出")
    parser.add_argument("--skip-field-profile", action="store_true", help="跳过字段空值率与基数统计")
    parser.add_argument("--quick-summary", action="store_true", help="仅输出映射摘要，跳过分桶和多值样例")
    parser.add_argument("--where-clause", default="", help="附加过滤条件，例如 dt >= '2026-04-01'")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--sample-limit", type=int, default=20, help="多值映射样例数量")
    parser.add_argument("--target-sample-limit", type=int, default=5, help="每个 source 展示的 target 样例数量")
    parser.add_argument("--preview-limit", type=int, default=50, help="直播间/SPU预览表输出数量")
    parser.add_argument("--max-wait", type=int, default=600, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    db = args.db
    table = args.table
    where_clause = args.where_clause.strip() or None
    table_fqn = f"{db}.{table}"
    date_tag = now_tag()

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), args.max_wait)

    print("[1/7] 解析表结构并匹配字段...", flush=True)
    raw_columns = runner.describe_table(db, table)
    resolved, missing_fields = resolve_columns(raw_columns)
    if args.schema_only:
        print("[SCHEMA] columns:")
        for column in raw_columns:
            print(column)
        print("[SCHEMA] resolved:")
        print(json.dumps(resolved, ensure_ascii=False, indent=2))
        return
    base_cte = build_base_cte(table_fqn, resolved, where_clause)

    print("[2/7] 统计整体概览...", flush=True)
    overview = query_overview(runner, db, base_cte)

    field_profile = pd.DataFrame(columns=["logical_field", "physical_column", "total_rows", "null_rows", "null_ratio", "distinct_values"])
    if args.skip_field_profile:
        print("[3/7] 跳过字段空值率与基数统计...", flush=True)
    else:
        print("[3/7] 统计字段空值率与基数...", flush=True)
        field_profile = query_field_profile(runner, db, base_cte, resolved)

    print("[4/7] 统计核心映射关系...", flush=True)
    mapping_checks: dict[str, dict[str, Any]] = {}
    for pair_name, source_field, target_field in MAPPING_PAIRS:
        if source_field not in resolved and source_field not in ("pay_date", "pay_hour"):
            continue
        if target_field not in resolved and target_field not in ("pay_date", "pay_hour"):
            continue
        if source_field not in ("pay_date", "pay_hour") and not resolved.get(source_field):
            continue
        if target_field not in ("pay_date", "pay_hour") and not resolved.get(target_field):
            continue
        mapping_checks[pair_name] = query_mapping_check(
            runner,
            db,
            base_cte,
            pair_name,
            source_field,
            target_field,
            args.sample_limit,
            args.target_sample_limit,
            args.quick_summary,
        )

    print("[5/7] 生成直播间维度预览...", flush=True)
    live_room_preview = query_live_room_preview(runner, db, base_cte, args.preview_limit)

    print("[6/7] 生成SPU维度预览...", flush=True)
    spu_preview = query_spu_preview(runner, db, base_cte, args.preview_limit)

    print("[7/7] 输出 Markdown / JSON / Excel...", flush=True)
    key_findings = build_key_findings(missing_fields, mapping_checks, overview)

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "table": table_fqn,
        "where_clause": where_clause,
        "raw_columns": raw_columns,
        "resolved_columns": resolved,
        "missing_fields": missing_fields,
        "overview": overview,
        "field_profile": field_profile.to_dict(orient="records"),
        "mapping_checks": {
            pair_name: {
                "pair_name": content["pair_name"],
                "source_field": content["source_field"],
                "target_field": content["target_field"],
                "summary": content["summary"],
                "bucket_distribution": content["bucket_distribution"].to_dict(orient="records"),
                "multi_mapping_samples": content["multi_mapping_samples"].to_dict(orient="records"),
            }
            for pair_name, content in mapping_checks.items()
        },
        "live_room_preview": live_room_preview.to_dict(orient="records"),
        "spu_preview": spu_preview.to_dict(orient="records"),
        "key_findings": key_findings,
    }

    md_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{date_tag}.md"
    json_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{date_tag}.json"
    xlsx_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{date_tag}.xlsx"

    write_markdown(
        md_path,
        db,
        table,
        where_clause,
        raw_columns,
        resolved,
        missing_fields,
        overview,
        field_profile,
        mapping_checks,
        live_room_preview,
        spu_preview,
        key_findings,
    )
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(
        xlsx_path,
        overview,
        resolved,
        missing_fields,
        field_profile,
        mapping_checks,
        live_room_preview,
        spu_preview,
        key_findings,
    )

    print("[OK] 报告已生成：")
    print(md_path)
    print(json_path)
    if HAS_OPENPYXL:
        print(xlsx_path)
    else:
        print("[WARN] Excel 未生成：当前环境缺少 openpyxl")


if __name__ == "__main__":
    main()
