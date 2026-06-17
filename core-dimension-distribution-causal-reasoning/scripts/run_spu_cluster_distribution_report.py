#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl matplotlib") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError as exc:
    matplotlib = None
    plt = None
    HAS_MATPLOTLIB = False

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:160]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> list[str]:
        req = models.DescribeTableRequest()
        req.DatabaseName = db
        req.TableName = table
        resp = self.client.DescribeTable(req)
        data = json.loads(resp.to_json_string())
        table_info = data.get("Table", data)
        columns = table_info.get("Columns") or []
        names: list[str] = []
        for item in columns:
            name = item.get("Name") if isinstance(item, dict) else None
            if name:
                names.append(name)
        return names


def non_empty_expr(column_name: str) -> str:
    return f"CASE WHEN `{column_name}` IS NULL OR TRIM(CAST(`{column_name}` AS STRING)) = '' THEN '无' ELSE TRIM(CAST(`{column_name}` AS STRING)) END"


def resolve_column_name(columns: set[str], candidates: list[str]) -> str | None:
    for name in candidates:
        if name in columns:
            return name
    return None


def fetch_single_value(runner: DlcRunner, sql: str, db: str, default: Any = 0) -> Any:
    rows = runner.exec_sql(sql, db)
    if not rows or not rows[0]:
        return default
    return rows[0][0]


def fetch_dwd_cluster_distribution(runner: DlcRunner) -> tuple[pd.DataFrame, dict[str, Any]]:
    cluster_sql = """
    WITH spu_base AS (
        SELECT
            spu,
            MAX(CASE WHEN big_cate IS NOT NULL AND TRIM(CAST(big_cate AS STRING)) <> '' THEN TRIM(CAST(big_cate AS STRING)) END) AS big_cate,
            MAX(CASE WHEN mid_cate IS NOT NULL AND TRIM(CAST(mid_cate AS STRING)) <> '' THEN TRIM(CAST(mid_cate AS STRING)) END) AS mid_cate,
            MAX(CASE WHEN sub_track IS NOT NULL AND TRIM(CAST(sub_track AS STRING)) <> '' THEN TRIM(CAST(sub_track AS STRING)) END) AS sub_track,
            MAX(CASE WHEN gender IS NOT NULL AND TRIM(CAST(gender AS STRING)) <> '' THEN TRIM(CAST(gender AS STRING)) END) AS gender,
            MAX(CASE WHEN scene IS NOT NULL AND TRIM(CAST(scene AS STRING)) <> '' THEN TRIM(CAST(scene AS STRING)) END) AS scene,
            MAX(CASE WHEN style IS NOT NULL AND TRIM(CAST(style AS STRING)) <> '' THEN TRIM(CAST(style AS STRING)) END) AS style
        FROM data_dwd.dwd_file_label_id_spu
        WHERE spu IS NOT NULL AND TRIM(CAST(spu AS STRING)) <> ''
        GROUP BY spu
    ),
    clustered AS (
        SELECT
            spu,
            COALESCE(big_cate, '无') AS big_cate,
            COALESCE(mid_cate, '无') AS mid_cate,
            COALESCE(sub_track, '无') AS sub_track,
            COALESCE(gender, '无') AS gender,
            COALESCE(scene, '无') AS scene,
            COALESCE(style, '无') AS style,
            concat_ws(
                ' | ',
                concat('big_cate=', COALESCE(big_cate, '无')),
                concat('mid_cate=', COALESCE(mid_cate, '无')),
                concat('sub_track=', COALESCE(sub_track, '无')),
                concat('gender=', COALESCE(gender, '无')),
                concat('scene=', COALESCE(scene, '无')),
                concat('style=', COALESCE(style, '无'))
            ) AS cluster_label
        FROM spu_base
    ),
    agg AS (
        SELECT
            cluster_label,
            COUNT(*) AS spu_count
        FROM clustered
        GROUP BY cluster_label
    )
    SELECT
        cluster_label,
        spu_count,
        ROUND(spu_count * 1.0 / SUM(spu_count) OVER (), 6) AS probability,
        ROUND(SUM(spu_count) OVER (ORDER BY spu_count DESC, cluster_label ASC) * 1.0 / SUM(spu_count) OVER (), 6) AS cumulative_probability
    FROM agg
    ORDER BY spu_count DESC, cluster_label ASC
    LIMIT 10
    """
    cluster_rows = runner.exec_sql(cluster_sql, "data_dwd")
    cluster_df = pd.DataFrame(
        cluster_rows,
        columns=["cluster_label", "spu_count", "probability", "cumulative_probability"],
    )
    if not cluster_df.empty:
        cluster_df["spu_count"] = pd.to_numeric(cluster_df["spu_count"])
        cluster_df["probability"] = pd.to_numeric(cluster_df["probability"])
        cluster_df["cumulative_probability"] = pd.to_numeric(cluster_df["cumulative_probability"])

    total_spu = int(
        fetch_single_value(
            runner,
            """
            SELECT COUNT(DISTINCT spu)
            FROM data_dwd.dwd_file_label_id_spu
            WHERE spu IS NOT NULL AND TRIM(CAST(spu AS STRING)) <> ''
            """,
            "data_dwd",
            0,
        )
    )
    total_rows = int(fetch_single_value(runner, "SELECT COUNT(*) FROM data_dwd.dwd_file_label_id_spu", "data_dwd", 0))
    distinct_clusters = int(
        fetch_single_value(
            runner,
            """
            WITH spu_base AS (
                SELECT
                    spu,
                    MAX(CASE WHEN big_cate IS NOT NULL AND TRIM(CAST(big_cate AS STRING)) <> '' THEN TRIM(CAST(big_cate AS STRING)) END) AS big_cate,
                    MAX(CASE WHEN mid_cate IS NOT NULL AND TRIM(CAST(mid_cate AS STRING)) <> '' THEN TRIM(CAST(mid_cate AS STRING)) END) AS mid_cate,
                    MAX(CASE WHEN sub_track IS NOT NULL AND TRIM(CAST(sub_track AS STRING)) <> '' THEN TRIM(CAST(sub_track AS STRING)) END) AS sub_track,
                    MAX(CASE WHEN gender IS NOT NULL AND TRIM(CAST(gender AS STRING)) <> '' THEN TRIM(CAST(gender AS STRING)) END) AS gender,
                    MAX(CASE WHEN scene IS NOT NULL AND TRIM(CAST(scene AS STRING)) <> '' THEN TRIM(CAST(scene AS STRING)) END) AS scene,
                    MAX(CASE WHEN style IS NOT NULL AND TRIM(CAST(style AS STRING)) <> '' THEN TRIM(CAST(style AS STRING)) END) AS style
                FROM data_dwd.dwd_file_label_id_spu
                WHERE spu IS NOT NULL AND TRIM(CAST(spu AS STRING)) <> ''
                GROUP BY spu
            )
            SELECT COUNT(DISTINCT concat_ws(
                ' | ',
                concat('big_cate=', COALESCE(big_cate, '无')),
                concat('mid_cate=', COALESCE(mid_cate, '无')),
                concat('sub_track=', COALESCE(sub_track, '无')),
                concat('gender=', COALESCE(gender, '无')),
                concat('scene=', COALESCE(scene, '无')),
                concat('style=', COALESCE(style, '无'))
            ))
            FROM spu_base
            """,
            "data_dwd",
            0,
        )
    )
    summary = {
        "table": "data_dwd.dwd_file_label_id_spu",
        "total_rows": total_rows,
        "distinct_spu": total_spu,
        "distinct_clusters": distinct_clusters,
        "top10_cluster_spu_sum": int(cluster_df["spu_count"].sum()) if not cluster_df.empty else 0,
        "top10_cluster_probability_sum": float(cluster_df["probability"].sum()) if not cluster_df.empty else 0.0,
        "cluster_dimensions": ["big_cate", "mid_cate", "sub_track", "gender", "scene", "style"],
    }
    return cluster_df, summary


def fetch_tb16_distributions(runner: DlcRunner) -> tuple[dict[str, pd.DataFrame], dict[str, Any]]:
    raw_columns = runner.describe_table("data_dim", "tb16_dim_product_sale_dimension")
    column_set = set(raw_columns)
    field_candidates = {
        "产品名称": ["产品名称"],
        "品牌大类": ["品牌大类"],
        "品牌中类": ["品牌中类", "中类"],
        "细分赛道": ["抖音成人商品_细分赛道", "子赛道"],
        "性别": ["性别_李宁bi", "性别内衬BI"],
        "风格": ["风格"],
        "适用场景": ["适用场景", "场景"],
    }

    summary: dict[str, Any] = {
        "table": "data_dim.tb16_dim_product_sale_dimension",
        "total_rows": int(fetch_single_value(runner, "SELECT COUNT(*) FROM data_dim.tb16_dim_product_sale_dimension", "data_dim", 0)),
        "distinct_spu": int(
            fetch_single_value(
                runner,
                "SELECT COUNT(DISTINCT spu) FROM data_dim.tb16_dim_product_sale_dimension WHERE spu IS NOT NULL AND TRIM(CAST(spu AS STRING)) <> ''",
                "data_dim",
                0,
            )
        ),
        "resolved_fields": {},
        "missing_fields": [],
    }
    result: dict[str, pd.DataFrame] = {}
    for logical_name, candidates in field_candidates.items():
        resolved = resolve_column_name(column_set, candidates)
        if not resolved:
            summary["missing_fields"].append({"logical_name": logical_name, "candidates": candidates})
            continue
        summary["resolved_fields"][logical_name] = resolved
        expr = non_empty_expr(resolved)
        sql = f"""
        WITH agg AS (
            SELECT
                {expr} AS field_value,
                COUNT(*) AS row_count
            FROM data_dim.tb16_dim_product_sale_dimension
            GROUP BY {expr}
        )
        SELECT
            field_value,
            row_count,
            ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS probability,
            ROUND(SUM(row_count) OVER (ORDER BY row_count DESC, field_value ASC) * 1.0 / SUM(row_count) OVER (), 6) AS cumulative_probability
        FROM agg
        ORDER BY row_count DESC, field_value ASC
        LIMIT 20
        """
        rows = runner.exec_sql(sql, "data_dim")
        df = pd.DataFrame(rows, columns=["field_value", "row_count", "probability", "cumulative_probability"])
        if not df.empty:
            df["row_count"] = pd.to_numeric(df["row_count"])
            df["probability"] = pd.to_numeric(df["probability"])
            df["cumulative_probability"] = pd.to_numeric(df["cumulative_probability"])
        result[logical_name] = df
    return result, summary


def configure_matplotlib_font() -> None:
    if not HAS_MATPLOTLIB:
        return
    candidates = [
        "Arial Unicode MS",
        "PingFang SC",
        "Heiti SC",
        "Hiragino Sans GB",
        "STHeiti",
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "WenQuanYi Zen Hei",
        "DejaVu Sans",
    ]
    plt.rcParams["font.sans-serif"] = candidates
    plt.rcParams["axes.unicode_minus"] = False


def save_bar_chart(df: pd.DataFrame, label_col: str, value_col: str, title: str, output_path: Path, top_n: int) -> None:
    if not HAS_MATPLOTLIB:
        return
    configure_matplotlib_font()
    chart_df = df.head(top_n).copy()
    if chart_df.empty:
        return
    labels = [str(item) for item in chart_df[label_col]]
    values = chart_df[value_col].astype(float).tolist()
    positions = list(range(len(chart_df)))
    fig_height = max(6, 0.55 * len(chart_df) + 1.8)
    fig, ax = plt.subplots(figsize=(14, fig_height))
    bars = ax.barh(positions, values, color="#2E6F95")
    ax.set_yticks(positions)
    ax.set_yticklabels(labels, fontsize=10)
    ax.invert_yaxis()
    ax.set_title(title, fontsize=14)
    ax.set_xlabel("概率")
    ax.grid(axis="x", linestyle="--", alpha=0.35)
    for idx, bar in enumerate(bars):
        ax.text(bar.get_width() + max(values) * 0.01, bar.get_y() + bar.get_height() / 2, f"{values[idx]:.2%}", va="center", fontsize=10)
    plt.tight_layout()
    fig.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def write_excel(
    cluster_df: pd.DataFrame,
    cluster_summary: dict[str, Any],
    dim_distributions: dict[str, pd.DataFrame],
    dim_summary: dict[str, Any],
    output_path: Path,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(
            [
                ["报告时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["dwd_file_label_id_spu 总行数", cluster_summary["total_rows"]],
                ["dwd_file_label_id_spu 唯一 SPU 数", cluster_summary["distinct_spu"]],
                ["dwd_file_label_id_spu 聚类组合数", cluster_summary["distinct_clusters"]],
                ["聚类维度", ", ".join(cluster_summary["cluster_dimensions"])],
                ["tb16_dim_product_sale_dimension 总行数", dim_summary["total_rows"]],
                ["tb16_dim_product_sale_dimension 唯一 SPU 数", dim_summary["distinct_spu"]],
            ],
            columns=["metric", "value"],
        )
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        cluster_df.to_excel(writer, sheet_name="dwd_spu_cluster_top10", index=False)
        for logical_name, df in dim_distributions.items():
            sheet_name = f"tb16_{logical_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:100]]
                width = min(max(len(value) for value in values) + 2, 80)
                sheet.column_dimensions[column_cells[0].column_letter].width = width

        add_excel_bar_chart(workbook["dwd_spu_cluster_top10"], "B", "A", "DWD SPU 聚类 Top10 概率分布")
        for logical_name in dim_distributions:
            sheet_name = f"tb16_{logical_name}"[:31]
            add_excel_bar_chart(workbook[sheet_name], "C", "A", f"TB16 {logical_name} Top20 概率分布")


def add_excel_bar_chart(sheet, value_col: str, label_col: str, title: str) -> None:
    max_row = sheet.max_row
    if max_row <= 1:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "类别"
    chart.x_axis.title = "概率"
    chart.height = max(7, min(18, 0.45 * (max_row - 1) + 4))
    chart.width = 18
    data = Reference(sheet, min_col=column_letter_to_index(value_col), min_row=1, max_row=max_row)
    cats = Reference(sheet, min_col=column_letter_to_index(label_col), min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    sheet.add_chart(chart, "F2")


def column_letter_to_index(letter: str) -> int:
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def write_markdown(
    cluster_df: pd.DataFrame,
    cluster_summary: dict[str, Any],
    dim_distributions: dict[str, pd.DataFrame],
    dim_summary: dict[str, Any],
    chart_paths: dict[str, Path],
    output_path: Path,
) -> None:
    lines: list[str] = []
    lines.append("# SPU 聚类与商品维度分布报告")
    lines.append("")
    lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")
    lines.append("## 一、dwd_file_label_id_spu 按 SPU 聚类 Top10")
    lines.append("")
    lines.append(f"- 表：`{cluster_summary['table']}`")
    lines.append(f"- 总行数：{cluster_summary['total_rows']}")
    lines.append(f"- 唯一 SPU 数：{cluster_summary['distinct_spu']}")
    lines.append(f"- 聚类组合数：{cluster_summary['distinct_clusters']}")
    lines.append(f"- 聚类维度：`{'`, `'.join(cluster_summary['cluster_dimensions'])}`")
    lines.append(f"- Top10 聚类累计概率：{cluster_summary['top10_cluster_probability_sum']:.2%}")
    if not HAS_MATPLOTLIB:
        lines.append("- 图表输出：当前环境未安装 `matplotlib`，已在 Excel 文件中内嵌概率分布图。")
    lines.append("")
    if "dwd_cluster_top10" in chart_paths:
        lines.append(f"![dwd_spu_cluster_top10]({chart_paths['dwd_cluster_top10'].name})")
        lines.append("")
    lines.append("| cluster_label | spu_count | probability | cumulative_probability |")
    lines.append("|---|---:|---:|---:|")
    for row in cluster_df.itertuples(index=False):
        lines.append(
            f"| {row.cluster_label} | {int(row.spu_count)} | {float(row.probability):.2%} | {float(row.cumulative_probability):.2%} |"
        )
    lines.append("")
    lines.append("## 二、tb16_dim_product_sale_dimension 重点维度 Top20")
    lines.append("")
    lines.append(f"- 表：`{dim_summary['table']}`")
    lines.append(f"- 总行数：{dim_summary['total_rows']}")
    lines.append(f"- 唯一 SPU 数：{dim_summary['distinct_spu']}")
    lines.append("")
    if dim_summary["missing_fields"]:
        lines.append("### 未命中的候选字段")
        lines.append("")
        for item in dim_summary["missing_fields"]:
            lines.append(f"- `{item['logical_name']}` 未命中候选列：`{'`, `'.join(item['candidates'])}`")
        lines.append("")
    for logical_name, df in dim_distributions.items():
        resolved = dim_summary["resolved_fields"].get(logical_name, logical_name)
        lines.append(f"### {logical_name} Top20")
        lines.append("")
        lines.append(f"- 实际字段：`{resolved}`")
        if logical_name in chart_paths:
            lines.append(f"- 分布图：`{chart_paths[logical_name].name}`")
        elif not HAS_MATPLOTLIB:
            lines.append("- 分布图：见 Excel 内嵌图表。")
        lines.append("")
        if logical_name in chart_paths:
            lines.append(f"![{logical_name}]({chart_paths[logical_name].name})")
            lines.append("")
        lines.append("| field_value | row_count | probability | cumulative_probability |")
        lines.append("|---|---:|---:|---:|")
        for row in df.itertuples(index=False):
            lines.append(
                f"| {row.field_value} | {int(row.row_count)} | {float(row.probability):.2%} | {float(row.cumulative_probability):.2%} |"
            )
        lines.append("")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成 SPU 聚类与 tb16 商品维度分布报告")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    date_tag = now_tag()

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), args.max_wait)

    print("[1/4] 统计 dwd_file_label_id_spu 的 SPU 聚类分布...", flush=True)
    cluster_df, cluster_summary = fetch_dwd_cluster_distribution(runner)

    print("[2/4] 统计 tb16_dim_product_sale_dimension 的重点维度分布...", flush=True)
    dim_distributions, dim_summary = fetch_tb16_distributions(runner)

    print("[3/4] 生成图表...", flush=True)
    chart_paths: dict[str, Path] = {}
    if HAS_MATPLOTLIB:
        dwd_chart = output_dir / f"spu-cluster-top10-{date_tag}.png"
        save_bar_chart(cluster_df, "cluster_label", "probability", "dwd_file_label_id_spu 按 SPU 聚类 Top10 概率分布", dwd_chart, 10)
        chart_paths["dwd_cluster_top10"] = dwd_chart
        for logical_name, df in dim_distributions.items():
            chart_path = output_dir / f"tb16-{logical_name}-top20-{date_tag}.png"
            save_bar_chart(df, "field_value", "probability", f"tb16_dim_product_sale_dimension {logical_name} Top20 概率分布", chart_path, 20)
            chart_paths[logical_name] = chart_path

    print("[4/4] 输出 Markdown / JSON / Excel...", flush=True)
    md_path = output_dir / f"spu-cluster-distribution-report-{date_tag}.md"
    json_path = output_dir / f"spu-cluster-distribution-report-{date_tag}.json"
    xlsx_path = output_dir / f"spu-cluster-distribution-report-{date_tag}.xlsx"
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "cluster_summary": cluster_summary,
        "cluster_top10": cluster_df.to_dict(orient="records"),
        "tb16_summary": dim_summary,
        "tb16_distributions": {key: df.to_dict(orient="records") for key, df in dim_distributions.items()},
        "chart_files": {key: str(path) for key, path in chart_paths.items()},
    }
    write_markdown(cluster_df, cluster_summary, dim_distributions, dim_summary, chart_paths, md_path)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(cluster_df, cluster_summary, dim_distributions, dim_summary, xlsx_path)

    print("报告已生成：")
    print(md_path)
    print(json_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
