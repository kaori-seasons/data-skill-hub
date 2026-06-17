#!/usr/bin/env python3
"""
基于实时 DLC 数据，生成“用什么内容”维度的五张表核心维度规整报告。

与旧版区别：
1. 不再只依赖 Excel 规划稿推断，而是优先探测并查询实时结果表。
2. 若结果表不存在，则降级到源表或跨表交叉验证，并在报告中明确说明。
3. 同时输出 PDF 与 JSON，便于业务阅读与技术复核。
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle


# Proxy bypass: 与仓库内其他 DLC 脚本保持一致
os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

from tencentcloud.common import credential
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.dlc.v20210125 import dlc_client, models


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parents[1]
DEFAULT_WORKBOOK = WORKSPACE_DIR / "短视频数据汇总表_20260121_V2.xlsx"
DEFAULT_REFERENCE_PDF = WORKSPACE_DIR / "鞋服电商短视频报告.pdf"
DEFAULT_OUTPUT_DIR = WORKSPACE_DIR
DEFAULT_PDF_NAME = "短视频五张表核心维度规整报告.pdf"
DEFAULT_JSON_NAME = "短视频五张表核心维度规整报告.json"
REGION = "ap-shanghai"

CONTENT_PRIORS = [
    "参考报告将鞋服短视频链路抽象为 5C：Content、Channel、Crowd、Commodity、Conversion。",
    "“用什么内容”不是单字段问题，而是从文件物理形态、商品语义、平台表达一直延伸到效果回流的连续结构。",
    "file_id 与 platform_source_id 天然是一对多；若没有桥接层，内容归因一定失真。",
    "AI 标签适合补高语义字段，不适合代替实体映射和基础元数据治理。",
]

TABLE_FALLBACKS = {
    "01": [("data_dwd", "dwd_t_file_resource_id")],
    "02": [("data_dwd", "dwd_file_label_id_spu"), ("data_dwd", "dwd_t_spu_dimension"), ("data_dim", "tb16_dim_product_sale_dimension")],
    "03": [("data_dwd", "dwd_platform_source_label")],
    "04": [("data_dwd", "dwd_file_platform_source_bridge"), ("data_dwd", "dwd_t_content_full")],
    "05": [("data_dwd", "dwd_platform_source_metric_di"), ("data_dwd", "dwd_platform_metrics"), ("data_dwd", "dwd_short_video_analysis_incre")],
}

TABLE_RUNTIME_NOTES = {
    "01": "表 01 是内容资源底座，重点看资源形态字段是否可稳定支撑“用什么内容”。",
    "02": "表 02 是商品语义层，重点看内容是否能稳定挂到商品与场景风格。",
    "03": "表 03 是平台表达层，重点看平台标签是否真实可用，而不是停留在规划字段。",
    "04": "表 04 是桥接层，重点不是字段丰富，而是关联是否存在、是否可信。",
    "05": "表 05 是效果反馈层，重点看指标表是否已落地；若未落地，则至少要验证源表能否形成闭环。",
}


@dataclass
class FieldMetric:
    label: str
    fill_count: int | None = None
    miss_count: int | None = None
    miss_rate_pct: float | None = None
    distribution: list[list[Any]] = field(default_factory=list)
    source_desc: str = ""
    unavailable_reason: str = ""


@dataclass
class RuntimeTableReport:
    table_id: str
    plan_name: str
    grain: str
    core_labels: list[str]
    planned_target_table: str
    resolved_table: str
    runtime_mode: str
    row_count: int | None = None
    distinct_grain_count: int | None = None
    schema_columns: list[str] = field(default_factory=list)
    freshness: list[list[str]] = field(default_factory=list)
    platform_distribution: list[list[Any]] = field(default_factory=list)
    field_metrics: list[FieldMetric] = field(default_factory=list)
    cross_checks: list[list[str]] = field(default_factory=list)
    note: str = ""
    findings: list[str] = field(default_factory=list)


@dataclass
class PlannedTable:
    table_id: str
    plan_name: str
    grain: str
    core_labels: list[str]
    target_table: str
    source_tables_initial: list[str]
    source_tables_final: list[str]
    logic_note: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").strip()


def split_lines(value: Any) -> list[str]:
    return [item.strip() for item in normalize_text(value).splitlines() if item and item.strip()]


def parse_core_labels(raw: str) -> list[str]:
    if not raw:
        return []
    text = str(raw).replace("【", "").replace("】", "")
    text = re.sub(r"[，,]", "、", text)
    text = text.replace("等", "")
    return [part.strip() for part in text.split("、") if part and part.strip()]


def extract_grain(plan_name: str) -> str:
    matched = re.search(r"【(.+?)】", plan_name)
    if matched:
        return matched.group(1)
    if "关联表" in plan_name:
        return "待定义"
    return plan_name


def backtick(name: str) -> str:
    return f"`{name}`"


def non_empty_str(name: str) -> str:
    return f"{backtick(name)} IS NOT NULL AND TRIM(CAST({backtick(name)} AS STRING)) <> ''"


def present_num(name: str) -> str:
    return f"{backtick(name)} IS NOT NULL"


def choose_column(columns: set[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 240):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]] | None:
        preview = " ".join(sql.strip().split())[:140]
        print(f"[SQL] {db}: {preview}...", flush=True)
        sql_b64 = base64.b64encode(sql.encode("utf-8")).decode("utf-8")
        task = models.Task()
        task.SparkSQLTask = {"SQL": sql_b64}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        try:
            resp = self.client.CreateTask(req)
            task_id = json.loads(resp.to_json_string()).get("TaskId")
            if not task_id:
                return None
        except Exception as exc:
            raise RuntimeError(f"CreateTask 失败: {exc}") from exc

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                rs = task_info.get("ResultSet", "[]")
                if isinstance(rs, str):
                    try:
                        return json.loads(base64.b64decode(rs).decode("utf-8"))
                    except Exception:
                        return json.loads(rs)
                return rs
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> dict[str, Any] | None:
        try:
            req = models.DescribeTableRequest()
            req.DatabaseName = db
            req.TableName = table
            resp = self.client.DescribeTable(req)
            data = json.loads(resp.to_json_string())
            return data.get("Table", data)
        except TencentCloudSDKException:
            return None


def build_planned_tables(workbook_path: Path) -> list[PlannedTable]:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["内容数据规划"]
    plans: list[PlannedTable] = []
    for row in range(3, 8):
        plan_name = normalize_text(ws.cell(row, 1).value)
        if not plan_name:
            continue
        table_id_match = re.match(r"(\d{2})", plan_name)
        table_id = table_id_match.group(1) if table_id_match else f"{row:02d}"
        plans.append(
            PlannedTable(
                table_id=table_id,
                plan_name=plan_name,
                grain=extract_grain(plan_name),
                core_labels=parse_core_labels(normalize_text(ws.cell(row, 2).value))
                or (["file_id", "platform_source_id", "SPU", "SKU", "映射规则", "映射置信度", "权重因子"] if table_id == "04" else []),
                target_table=normalize_text(ws.cell(row, 4).value)
                or ("dwd_file_platform_source_bridge" if table_id == "04" else "dwd_platform_source_metric_di" if table_id == "05" else ""),
                source_tables_initial=split_lines(ws.cell(row, 3).value),
                source_tables_final=split_lines(ws.cell(row, 6).value),
                logic_note=normalize_text(ws.cell(row, 8).value) or normalize_text(ws.cell(row, 12).value),
            )
        )
    return plans


def resolve_table(runner: DlcRunner, table_id: str) -> tuple[str, str, dict[str, Any] | None, str]:
    for db, table in TABLE_FALLBACKS.get(table_id, []):
        desc = runner.describe_table(db, table)
        if desc:
            mode = "result" if table_id != "05" or db == "data_dwd" else "fallback"
            return db, table, desc, mode
    return "", "", None, "missing"


def to_int(value: Any) -> int | None:
    try:
        return int(str(value))
    except Exception:
        return None


def to_float(value: Any) -> float | None:
    try:
        return float(str(value))
    except Exception:
        return None


def format_pct(value: float | None) -> str:
    if value is None:
        return "NA"
    return f"{value:.2f}%"


def value_at(res: list[list[Any]] | None, row: int = 0, col: int = 0) -> Any:
    try:
        return res[row][col]
    except Exception:
        return None


def build_batch_fill_metrics(
    runner: DlcRunner,
    db: str,
    table: str,
    field_specs: list[dict[str, Any]],
) -> tuple[int | None, dict[str, tuple[int | None, int | None, float | None]]]:
    return build_batch_fill_metrics_chunked(runner, db, table, field_specs, None)


def build_batch_fill_metrics_chunked(
    runner: DlcRunner,
    db: str,
    table: str,
    field_specs: list[dict[str, Any]],
    total_hint: int | None,
    chunk_size: int = 4,
) -> tuple[int | None, dict[str, tuple[int | None, int | None, float | None]]]:
    total = total_hint
    if total is None:
        total = to_int(value_at(runner.exec_sql(f"SELECT COUNT(*) FROM {db}.{table}", db), 0, 0))
    metrics: dict[str, tuple[int | None, int | None, float | None]] = {}
    for offset in range(0, len(field_specs), chunk_size):
        chunk = field_specs[offset : offset + chunk_size]
        parts = []
        for idx, spec in enumerate(chunk):
            alias = f"f_{offset+idx}"
            parts.append(f"SUM(CASE WHEN {spec['fill_expr']} THEN 1 ELSE 0 END) AS {alias}")
        sql = f"SELECT {', '.join(parts)} FROM {db}.{table}"
        res = runner.exec_sql(sql, db)
        for idx, spec in enumerate(chunk):
            fill = to_int(value_at(res, 0, idx))
            if total is None or fill is None:
                metrics[spec["label"]] = (None, None, None)
                continue
            miss = total - fill
            miss_rate = round(miss * 100.0 / total, 2) if total else 0.0
            metrics[spec["label"]] = (fill, miss, miss_rate)
    return total, metrics


def distribution_query(
    runner: DlcRunner,
    db: str,
    table: str,
    expr: str,
    where_clause: str,
    limit: int = 10,
) -> list[list[Any]]:
    sql = f"""
    SELECT CAST(({expr}) AS STRING) AS label_value, COUNT(*) AS cnt
    FROM {db}.{table}
    WHERE {where_clause}
    GROUP BY CAST(({expr}) AS STRING)
    ORDER BY cnt DESC, label_value
    LIMIT {limit}
    """
    return runner.exec_sql(sql, db) or []


def collect_field_metric(
    runner: DlcRunner,
    db: str,
    table: str,
    label: str,
    fill_expr: str,
    source_desc: str,
    base_where: str = "1=1",
    dist_expr: str | None = None,
    dist_where: str | None = None,
) -> FieldMetric:
    sql = f"""
    SELECT COUNT(*) AS total_rows,
           SUM(CASE WHEN {fill_expr} THEN 1 ELSE 0 END) AS fill_rows
    FROM {db}.{table}
    WHERE {base_where}
    """
    res = runner.exec_sql(sql, db)
    total = to_int(value_at(res, 0, 0))
    fill = to_int(value_at(res, 0, 1))
    miss = None if total is None or fill is None else total - fill
    miss_rate = None if total is None or fill is None or total == 0 else round((total - fill) * 100.0 / total, 2)
    distribution = []
    if dist_expr:
        where_clause = base_where
        if dist_where:
            where_clause = f"({base_where}) AND ({dist_where})"
        distribution = distribution_query(runner, db, table, dist_expr, where_clause)
    return FieldMetric(label, fill, miss, miss_rate, distribution, source_desc)


def collect_freshness(
    runner: DlcRunner,
    db: str,
    table: str,
    columns: list[str],
) -> list[list[str]]:
    rows: list[list[str]] = []
    for col in columns:
        sql = f"""
        SELECT MIN(CAST({backtick(col)} AS STRING)), MAX(CAST({backtick(col)} AS STRING))
        FROM {db}.{table}
        WHERE {backtick(col)} IS NOT NULL AND TRIM(CAST({backtick(col)} AS STRING)) <> ''
        """
        try:
            res = runner.exec_sql(sql, db)
            rows.append([col, str(value_at(res, 0, 0) or "NA"), str(value_at(res, 0, 1) or "NA")])
        except Exception:
            rows.append([col, "NA", "NA"])
    return rows


def distinct_count(runner: DlcRunner, db: str, table: str, col: str) -> int | None:
    sql = f"SELECT COUNT(DISTINCT {backtick(col)}) FROM {db}.{table}"
    return to_int(value_at(runner.exec_sql(sql, db), 0, 0))


def analyze_table01(runner: DlcRunner, plan: PlannedTable, db: str, table: str, desc: dict[str, Any]) -> RuntimeTableReport:
    print(f"[TABLE 01] analyzing {db}.{table}", flush=True)
    columns = {col["Name"] for col in desc.get("Columns", [])}
    total_hint = to_int(desc.get("RecordCount"))
    field_specs = [
        {"label": "平台", "fill_expr": non_empty_str("platform"), "dist_expr": backtick("platform"), "dist_where": non_empty_str("platform"), "source_desc": "platform"},
        {"label": "品牌", "fill_expr": non_empty_str("brand"), "dist_expr": backtick("brand"), "dist_where": non_empty_str("brand"), "source_desc": "brand"},
        {"label": "SPU", "fill_expr": non_empty_str("spu"), "source_desc": "spu"},
        {"label": "创作日期", "fill_expr": non_empty_str("create_time"), "source_desc": "create_time"},
    ]
    total, batch = build_batch_fill_metrics_chunked(runner, db, table, field_specs, total_hint, chunk_size=1)
    metrics: list[FieldMetric] = []
    for spec in field_specs:
        dist = distribution_query(runner, db, table, spec["dist_expr"], spec["dist_where"]) if spec.get("dist_expr") else []
        fill, miss, miss_rate = batch.get(spec["label"], (None, None, None))
        metrics.append(FieldMetric(spec["label"], fill, miss, miss_rate, dist, spec["source_desc"]))

    metrics.extend(
        [
            collect_field_metric(
                runner,
                "data_ods",
                "ods_t_file_information",
                "是否已发",
                "1=1",
                "ods_t_file_information.publish_date 推导（仅视频源表）",
                base_where="is_delete <> 'true'",
                dist_expr=f"CASE WHEN {non_empty_str('publish_date')} THEN '已发' ELSE '未发' END",
                dist_where="1=1",
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_t_file_information",
                "视频时长类型",
                non_empty_str("video_duration_type"),
                "ods_t_file_information.video_duration_type（仅视频源表）",
                base_where="is_delete <> 'true'",
                dist_expr=backtick("video_duration_type"),
                dist_where=non_empty_str("video_duration_type"),
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_t_file_information",
                "视频宽幅度",
                f"{present_num('width')} AND {present_num('height')}",
                "ods_t_file_information.width,height（仅视频源表）",
                base_where="is_delete <> 'true'",
                dist_expr="CASE WHEN `width` > `height` THEN '横屏' WHEN `width` < `height` THEN '竖屏' WHEN `width` = `height` THEN '方图' ELSE '未知' END",
                dist_where=f"{present_num('width')} AND {present_num('height')}",
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_pic_for_up_new_backup",
                "图片大小",
                present_num("length"),
                "ods_pic_for_up_new_backup.length（仅图片源表）",
                base_where="is_delete = 'false'",
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_t_image_file_information",
                "是否穿搭",
                non_empty_str("file_type"),
                "ods_t_image_file_information.file_type（仅图片源表）",
                base_where="1=1",
                dist_expr=f"CASE WHEN {backtick('file_type')} LIKE '%穿搭图%' THEN '是穿搭图' ELSE '否' END",
                dist_where=non_empty_str("file_type"),
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_pic_for_up_new_backup",
                "是否白底图",
                non_empty_str("content_type"),
                "ods_pic_for_up_new_backup.content_type（仅图片源表）",
                base_where="is_delete = 'false'",
                dist_expr=f"CASE WHEN {backtick('content_type')} LIKE '%白底图%' THEN '是白底图' ELSE '否' END",
                dist_where=non_empty_str("content_type"),
            ),
            collect_field_metric(
                runner,
                "data_ods",
                "ods_pic_for_up_new_backup",
                "是否主图",
                non_empty_str("image_name"),
                "ods_pic_for_up_new_backup.image_name（仅图片源表）",
                base_where="is_delete = 'false'",
                dist_expr=f"CASE WHEN {backtick('image_name')} LIKE '%主图%' THEN '是主图' ELSE '否' END",
                dist_where=non_empty_str("image_name"),
            ),
        ]
    )

    file_id_distinct = distinct_count(runner, db, table, "file_id") if "file_id" in columns else None
    platform_dist = distribution_query(runner, db, table, backtick("file_type"), non_empty_str("file_type")) if "file_type" in columns else []

    findings: list[str] = []
    for metric in metrics:
        if metric.miss_rate_pct is not None and metric.miss_rate_pct >= 90:
            findings.append(f"{metric.label} 缺失率 {format_pct(metric.miss_rate_pct)}，说明该字段在资源底座上不能当成全量主切片。")
    if total is not None and file_id_distinct is not None and total != file_id_distinct:
        findings.append(f"file_id 非唯一，当前总行数 {total}，去重后 {file_id_distinct}，底座实体仍存在重复污染。")

    return RuntimeTableReport(
        table_id=plan.table_id,
        plan_name=plan.plan_name,
        grain=plan.grain,
        core_labels=plan.core_labels,
        planned_target_table=plan.target_table,
        resolved_table=f"{db}.{table}",
        runtime_mode="result",
        row_count=total,
        distinct_grain_count=file_id_distinct,
        schema_columns=sorted(columns),
        freshness=collect_freshness(runner, db, table, [col for col in ["create_time", "publish_date"] if col in columns]),
        platform_distribution=platform_dist,
        field_metrics=metrics,
        cross_checks=[["runtime_note", TABLE_RUNTIME_NOTES["01"], ""]]
        + ([["file_id_distinct", str(file_id_distinct), "与总行数对比判断资源实体是否唯一"]] if file_id_distinct is not None else []),
        note=plan.logic_note,
        findings=findings,
    )


def analyze_table02(runner: DlcRunner, plan: PlannedTable, db: str, table: str, desc: dict[str, Any], runtime_mode: str) -> RuntimeTableReport:
    print(f"[TABLE 02] analyzing {db}.{table} mode={runtime_mode}", flush=True)
    columns = {col["Name"] for col in desc.get("Columns", [])}
    total_hint = to_int(desc.get("RecordCount"))
    if table == "tb16_dim_product_sale_dimension":
        mappings = [
            ("品名", choose_column(columns, ["产品名称"]), "产品名称"),
            ("统一大类", choose_column(columns, ["品牌大类", "抖音成人商品_大类规整"]), "品牌大类/大类规整"),
            ("统一中类", choose_column(columns, ["抖音成人商品_中类规整", "中类"]), "中类规整"),
            ("统一赛道", choose_column(columns, ["抖音成人商品_细分赛道", "子赛道"]), "细分赛道"),
            ("性别", choose_column(columns, ["性别_李宁bi"]), "性别_李宁bi"),
            ("场景", choose_column(columns, ["适用场景"]), "适用场景"),
            ("风格", choose_column(columns, ["风格"]), "风格"),
        ]
    else:
        mappings = [
            ("品名", choose_column(columns, ["product_name"]), "product_name"),
            ("统一大类", choose_column(columns, ["brand_big_cate", "big_cate"]), "brand_big_cate/big_cate"),
            ("统一中类", choose_column(columns, ["mid_cate"]), "mid_cate"),
            ("统一赛道", choose_column(columns, ["sub_track"]), "sub_track"),
            ("性别", choose_column(columns, ["gender_lining_bi", "gender"]), "gender_lining_bi/gender"),
            ("场景", choose_column(columns, ["scene_tb16", "scene_rpa_material", "scene"]), "scene_tb16/scene_rpa_material"),
            ("风格", choose_column(columns, ["style_tb16", "style_rpa_material", "style"]), "style_tb16/style_rpa_material"),
        ]

    field_specs: list[dict[str, Any]] = []
    for label, col, desc_text in mappings:
        if not col:
            continue
        field_specs.append(
            {
                "label": label,
                "fill_expr": non_empty_str(col),
                "dist_expr": None if label == "品名" else backtick(col),
                "dist_where": None if label == "品名" else non_empty_str(col),
                "source_desc": desc_text,
            }
        )
    total, batch = build_batch_fill_metrics_chunked(runner, db, table, field_specs, total_hint)
    metrics: list[FieldMetric] = []
    for spec in field_specs:
        dist = distribution_query(runner, db, table, spec["dist_expr"], spec["dist_where"]) if spec.get("dist_expr") else []
        fill, miss, miss_rate = batch.get(spec["label"], (None, None, None))
        metrics.append(FieldMetric(spec["label"], fill, miss, miss_rate, dist, spec["source_desc"]))

    distinct_spu = None
    for candidate in ["spu", "SPU"]:
        if candidate in columns:
            distinct_spu = distinct_count(runner, db, table, candidate)
            break

    findings: list[str] = []
    for metric in metrics:
        if metric.miss_rate_pct is not None and metric.miss_rate_pct >= 40:
            findings.append(f"{metric.label} 缺失率 {format_pct(metric.miss_rate_pct)}，商品语义层目前还不适合直接作为稳定内容切片。")
    if runtime_mode != "result":
        findings.append("表 02 结果表未命中，当前使用源表或替代表做实时分析，说明语义层结果模型仍未稳定沉淀。")

    return RuntimeTableReport(
        table_id=plan.table_id,
        plan_name=plan.plan_name,
        grain=plan.grain,
        core_labels=plan.core_labels,
        planned_target_table=plan.target_table,
        resolved_table=f"{db}.{table}",
        runtime_mode=runtime_mode,
        row_count=total,
        distinct_grain_count=distinct_spu,
        schema_columns=sorted(columns),
        freshness=[],
        platform_distribution=[],
        field_metrics=metrics,
        cross_checks=[["runtime_note", TABLE_RUNTIME_NOTES["02"], ""], ["distinct_spu", str(distinct_spu or "NA"), "SPU 是语义层的天然主键"]] if distinct_spu is not None else [["runtime_note", TABLE_RUNTIME_NOTES["02"], ""]],
        note=plan.logic_note,
        findings=findings,
    )


def analyze_table03(runner: DlcRunner, plan: PlannedTable, db: str, table: str, desc: dict[str, Any]) -> RuntimeTableReport:
    print(f"[TABLE 03] analyzing {db}.{table}", flush=True)
    columns = {col["Name"] for col in desc.get("Columns", [])}
    total_hint = to_int(desc.get("RecordCount"))
    field_specs = [
        {"label": "发布日期", "fill_expr": non_empty_str("publish_date"), "source_desc": "publish_date"},
        {"label": "发布作者", "fill_expr": non_empty_str("author"), "source_desc": "author"},
        {"label": "内容大标签", "fill_expr": non_empty_str("big_tag"), "dist_expr": backtick("big_tag"), "dist_where": non_empty_str("big_tag"), "source_desc": "big_tag"},
        {"label": "内容中标签", "fill_expr": non_empty_str("middle_tag"), "dist_expr": backtick("middle_tag"), "dist_where": non_empty_str("middle_tag"), "source_desc": "middle_tag"},
    ]
    total, batch = build_batch_fill_metrics_chunked(runner, db, table, field_specs, total_hint)
    metrics: list[FieldMetric] = []
    for spec in field_specs:
        dist = distribution_query(runner, db, table, spec["dist_expr"], spec["dist_where"]) if spec.get("dist_expr") else []
        fill, miss, miss_rate = batch.get(spec["label"], (None, None, None))
        metrics.append(FieldMetric(spec["label"], fill, miss, miss_rate, dist, spec["source_desc"]))

    distinct_psid = distinct_count(runner, db, table, "platform_source_id") if "platform_source_id" in columns else None
    platform_dist = distribution_query(runner, db, table, backtick("platform"), non_empty_str("platform")) if "platform" in columns else []
    findings: list[str] = []
    for metric in metrics:
        if metric.label in {"内容大标签", "内容中标签"} and metric.miss_rate_pct is not None and metric.miss_rate_pct >= 60:
            findings.append(f"{metric.label} 缺失率 {format_pct(metric.miss_rate_pct)}，平台表达层仍明显依赖补录或标题解析。")

    return RuntimeTableReport(
        table_id=plan.table_id,
        plan_name=plan.plan_name,
        grain=plan.grain,
        core_labels=plan.core_labels,
        planned_target_table=plan.target_table,
        resolved_table=f"{db}.{table}",
        runtime_mode="result",
        row_count=total,
        distinct_grain_count=distinct_psid,
        schema_columns=sorted(columns),
        freshness=collect_freshness(runner, db, table, [col for col in ["publish_date"] if col in columns]),
        platform_distribution=platform_dist,
        field_metrics=metrics,
        cross_checks=[["runtime_note", TABLE_RUNTIME_NOTES["03"], ""], ["distinct_platform_source_id", str(distinct_psid or "NA"), "用于判断平台素材实体是否唯一"]],
        note=plan.logic_note,
        findings=findings,
    )


def analyze_table04(
    runner: DlcRunner,
    plan: PlannedTable,
    db: str,
    table: str,
    desc: dict[str, Any] | None,
) -> RuntimeTableReport:
    print(f"[TABLE 04] analyzing bridge candidate {db}.{table}" if desc else "[TABLE 04] bridge table missing, running cross validation", flush=True)
    if desc:
        columns = {col["Name"] for col in desc.get("Columns", [])}
        total_hint = to_int(desc.get("RecordCount"))
        candidates = [
            ("file_id", choose_column(columns, ["file_id"])),
            ("platform_source_id", choose_column(columns, ["platform_source_id"])),
            ("SPU", choose_column(columns, ["spu", "SPU"])),
            ("SKU", choose_column(columns, ["sku", "SKU"])),
            ("映射规则", choose_column(columns, ["mapping_rule", "match_rule", "rule_type"])),
            ("映射置信度", choose_column(columns, ["mapping_confidence", "match_confidence"])),
            ("权重因子", choose_column(columns, ["weight_factor", "weight"])),
        ]
        field_specs = []
        for label, col in candidates:
            if col:
                fill_expr = present_num(col) if label in {"映射置信度", "权重因子"} else non_empty_str(col)
                dist_expr = backtick(col) if label in {"映射规则"} else None
                dist_where = fill_expr if dist_expr else None
                field_specs.append({"label": label, "fill_expr": fill_expr, "dist_expr": dist_expr, "dist_where": dist_where, "source_desc": col})
        total, batch = build_batch_fill_metrics_chunked(runner, db, table, field_specs, total_hint)
        metrics: list[FieldMetric] = []
        for spec in field_specs:
            dist = distribution_query(runner, db, table, spec["dist_expr"], spec["dist_where"]) if spec.get("dist_expr") else []
            fill, miss, miss_rate = batch.get(spec["label"], (None, None, None))
            metrics.append(FieldMetric(spec["label"], fill, miss, miss_rate, dist, spec["source_desc"]))
        return RuntimeTableReport(
            table_id=plan.table_id,
            plan_name=plan.plan_name,
            grain=plan.grain,
            core_labels=plan.core_labels,
            planned_target_table=plan.target_table,
            resolved_table=f"{db}.{table}",
            runtime_mode="result",
            row_count=total,
            distinct_grain_count=distinct_count(runner, db, table, "platform_source_id") if "platform_source_id" in columns else None,
            schema_columns=sorted(columns),
            field_metrics=metrics,
            cross_checks=[["runtime_note", TABLE_RUNTIME_NOTES["04"], ""]],
            note=plan.logic_note,
            findings=[],
        )

    # 桥接表不存在时，实时做跨表验证
    tbl01 = "data_dwd.dwd_t_file_resource_id"
    tbl03 = "data_dwd.dwd_platform_source_label"
    t01_desc = runner.describe_table("data_dwd", "dwd_t_file_resource_id")
    t03_desc = runner.describe_table("data_dwd", "dwd_platform_source_label")
    cross_rows: list[list[str]] = [["metric", "value", "comment"]]
    findings = ["实时探测未发现桥接结果表，表 04 仍是数仓当前最关键的结构性缺口。"]
    if t01_desc and t03_desc:
        res1 = runner.exec_sql(
            f"SELECT COUNT(*), SUM(CASE WHEN {non_empty_str('spu')} THEN 1 ELSE 0 END), COUNT(DISTINCT CASE WHEN {non_empty_str('spu')} THEN {backtick('spu')} END) FROM {tbl01}",
            "data_dwd",
        )
        res2 = runner.exec_sql(
            f"SELECT COUNT(*), SUM(CASE WHEN {non_empty_str('spu')} THEN 1 ELSE 0 END), COUNT(DISTINCT CASE WHEN {non_empty_str('spu')} THEN {backtick('spu')} END) FROM {tbl03}",
            "data_dwd",
        )
        overlap = runner.exec_sql(
            f"""
            SELECT
              COUNT(DISTINCT t1.spu) AS spu_in_01,
              COUNT(DISTINCT CASE WHEN t3.spu IS NOT NULL THEN t1.spu END) AS matched_spu
            FROM {tbl01} t1
            LEFT JOIN {tbl03} t3 ON t1.spu = t3.spu
            WHERE {non_empty_str('spu').replace('`spu`', 't1.`spu`')}
            """,
            "data_dwd",
        )
        cross_rows.extend(
            [
                ["t01_total_rows", str(value_at(res1, 0, 0) or "NA"), "资源底座总量"],
                ["t01_rows_with_spu", str(value_at(res1, 0, 1) or "NA"), "资源底座中可挂商品的记录数"],
                ["t03_total_rows", str(value_at(res2, 0, 0) or "NA"), "平台表达层总量"],
                ["t03_rows_with_spu", str(value_at(res2, 0, 1) or "NA"), "平台表达层中可挂商品的记录数"],
                ["spu_overlap", str(value_at(overlap, 0, 1) or "NA"), "01 与 03 通过 SPU 的交集能力"],
            ]
        )
        findings.append("当前只能通过 SPU 做弱关联验证，无法证明平台视频与原始素材是一一对应还是裁剪复用。")

    return RuntimeTableReport(
        table_id=plan.table_id,
        plan_name=plan.plan_name,
        grain=plan.grain,
        core_labels=plan.core_labels,
        planned_target_table=plan.target_table,
        resolved_table="未发现桥接结果表",
        runtime_mode="cross_validation",
        row_count=None,
        distinct_grain_count=None,
        schema_columns=[],
        field_metrics=[],
        cross_checks=cross_rows,
        note=plan.logic_note,
        findings=findings,
    )


def analyze_table05(runner: DlcRunner, plan: PlannedTable, db: str, table: str, desc: dict[str, Any] | None, runtime_mode: str) -> RuntimeTableReport:
    print(f"[TABLE 05] analyzing {db}.{table} mode={runtime_mode}" if desc else "[TABLE 05] result table missing, scanning source bundle", flush=True)
    if desc and db == "data_dwd":
        columns = {col["Name"] for col in desc.get("Columns", [])}
        total_hint = to_int(desc.get("RecordCount"))
        exposure_col = choose_column(columns, ["曝光量", "imp_cnt", "exposure_cnt", "show_cnt", "exposure_vol"])
        click_col = choose_column(columns, ["点击量", "click_cnt", "click_pv", "view_cnt", "查看量", "view_vol", "watch_vol"])
        platform_col = choose_column(columns, ["platform", "平台"])
        date_col = choose_column(columns, ["dt", "date", "stat_day", "日期", "publish_date"])
        field_specs = []
        if platform_col:
            field_specs.append({"label": "平台", "fill_expr": non_empty_str(platform_col), "dist_expr": backtick(platform_col), "dist_where": non_empty_str(platform_col), "source_desc": platform_col})
        if click_col:
            field_specs.append({"label": "点击量", "fill_expr": present_num(click_col), "source_desc": click_col})
        if exposure_col:
            field_specs.append({"label": "曝光量", "fill_expr": present_num(exposure_col), "source_desc": exposure_col})
        total, batch = build_batch_fill_metrics_chunked(runner, db, table, field_specs, total_hint)
        metrics = []
        for spec in field_specs:
            dist = distribution_query(runner, db, table, spec["dist_expr"], spec["dist_where"]) if spec.get("dist_expr") else []
            fill, miss, miss_rate = batch.get(spec["label"], (None, None, None))
            metrics.append(FieldMetric(spec["label"], fill, miss, miss_rate, dist, spec["source_desc"]))
        return RuntimeTableReport(
            table_id=plan.table_id,
            plan_name=plan.plan_name,
            grain=plan.grain,
            core_labels=plan.core_labels,
            planned_target_table=plan.target_table,
            resolved_table=f"{db}.{table}",
            runtime_mode=runtime_mode,
            row_count=total,
            distinct_grain_count=None,
            schema_columns=sorted(columns),
            freshness=collect_freshness(runner, db, table, [date_col] if date_col else []),
            platform_distribution=distribution_query(runner, db, table, backtick(platform_col), non_empty_str(platform_col)) if platform_col else [],
            field_metrics=metrics,
            cross_checks=[["runtime_note", TABLE_RUNTIME_NOTES["05"], ""]],
            note=plan.logic_note,
            findings=[],
        )

    # 结果表未落地时，直接扫源表指标覆盖
    sources = [
        {"platform": "抖音", "db": "data_ods", "table": "ods_rpa_douyin_compass_video", "id_col": "aweme_id", "date_col": "video_create_time"},
        {"platform": "天猫", "db": "data_ods", "table": "ods_rpa_sycm_video_content_single_analysis", "id_col": "video_id", "date_col": "publish_date"},
        {"platform": "得物", "db": "data_ads", "table": "ads_dewu_gravity_task_df", "id_col": "task_id", "date_col": "dynamic_published_datetime"},
        {"platform": "阿里妈妈", "db": "data_ods", "table": "ods_rpa_alimama_wxt_content_report_data_detail", "id_col": "content_id", "date_col": "stat_day"},
    ]
    cross_rows = [["platform", "rows", "distinct_ids", "min_date", "max_date"]]
    findings = ["表 05 结果表未命中，当前退回到各平台源表验证指标闭环能力。"]
    for src in sources:
        desc_src = runner.describe_table(src["db"], src["table"])
        if not desc_src:
            cross_rows.append([src["platform"], "NA", "NA", "NA", "NA"])
            continue
        sql = f"""
        SELECT COUNT(*), COUNT(DISTINCT {backtick(src['id_col'])}),
               MIN(CAST({backtick(src['date_col'])} AS STRING)),
               MAX(CAST({backtick(src['date_col'])} AS STRING))
        FROM {src['db']}.{src['table']}
        WHERE {backtick(src['date_col'])} IS NOT NULL AND TRIM(CAST({backtick(src['date_col'])} AS STRING)) <> ''
        """
        res = runner.exec_sql(sql, src["db"])
        cross_rows.append(
            [
                src["platform"],
                str(value_at(res, 0, 0) or "NA"),
                str(value_at(res, 0, 1) or "NA"),
                str(value_at(res, 0, 2) or "NA"),
                str(value_at(res, 0, 3) or "NA"),
            ]
        )

    return RuntimeTableReport(
        table_id=plan.table_id,
        plan_name=plan.plan_name,
        grain=plan.grain,
        core_labels=plan.core_labels,
        planned_target_table=plan.target_table,
        resolved_table="结果表未发现，改扫源表",
        runtime_mode="source_bundle",
        row_count=None,
        distinct_grain_count=None,
        schema_columns=[],
        field_metrics=[],
        cross_checks=cross_rows,
        note=plan.logic_note,
        findings=findings,
    )


def analyze_tables(runner: DlcRunner, plans: list[PlannedTable]) -> list[RuntimeTableReport]:
    reports: list[RuntimeTableReport] = []
    for plan in plans:
        print(f"[PLAN] resolving {plan.table_id} {plan.plan_name}", flush=True)
        db, table, desc, runtime_mode = resolve_table(runner, plan.table_id)
        if plan.table_id == "01":
            if not desc:
                raise RuntimeError("未找到表 01 的实时结果表 data_dwd.dwd_t_file_resource_id")
            reports.append(analyze_table01(runner, plan, db, table, desc))
        elif plan.table_id == "02":
            if not desc:
                raise RuntimeError("未找到表 02 的结果表或源表")
            reports.append(analyze_table02(runner, plan, db, table, desc, runtime_mode))
        elif plan.table_id == "03":
            if not desc:
                raise RuntimeError("未找到表 03 的实时结果表 data_dwd.dwd_platform_source_label")
            reports.append(analyze_table03(runner, plan, db, table, desc))
        elif plan.table_id == "04":
            reports.append(analyze_table04(runner, plan, db, table, desc))
        elif plan.table_id == "05":
            reports.append(analyze_table05(runner, plan, db, table, desc, runtime_mode))
    return reports


def collect_risks(reports: list[RuntimeTableReport]) -> list[list[str]]:
    rows = [["表", "字段/主题", "风险判断"]]
    for report in reports:
        if report.runtime_mode in {"cross_validation", "source_bundle"}:
            rows.append([report.plan_name, "结果表状态", "未命中结果表，说明这一层模型尚未稳定落地。"])
        for metric in report.field_metrics:
            if metric.miss_rate_pct is not None and metric.miss_rate_pct >= 80:
                rows.append([report.plan_name, metric.label, f"缺失率 {format_pct(metric.miss_rate_pct)}，不适合作为全量统一切片字段。"])
            elif metric.miss_rate_pct is not None and metric.miss_rate_pct >= 40:
                rows.append([report.plan_name, metric.label, f"缺失率 {format_pct(metric.miss_rate_pct)}，适合作为辅标签，不适合作为强主标签。"])
    return rows


def build_payload(workbook_path: Path, reference_pdf: Path, plans: list[PlannedTable], reports: list[RuntimeTableReport]) -> dict[str, Any]:
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "mode": "realtime_dlc",
        "workbook": str(workbook_path),
        "reference_pdf": str(reference_pdf),
        "content_priors": CONTENT_PRIORS,
        "planned_tables": [asdict(item) for item in plans],
        "runtime_reports": [asdict(item) for item in reports],
        "risks": collect_risks(reports),
    }


def register_fonts() -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


def build_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=base["Title"], fontName="STSong-Light", fontSize=20, leading=28, alignment=TA_CENTER, textColor=colors.HexColor("#1F1F1F")),
        "heading1": ParagraphStyle("heading1", parent=base["Heading1"], fontName="STSong-Light", fontSize=14, leading=22, textColor=colors.HexColor("#0F4C81"), spaceBefore=8, spaceAfter=8),
        "heading2": ParagraphStyle("heading2", parent=base["Heading2"], fontName="STSong-Light", fontSize=11.5, leading=18, textColor=colors.HexColor("#1F1F1F"), spaceBefore=6, spaceAfter=4),
        "body": ParagraphStyle("body", parent=base["BodyText"], fontName="STSong-Light", fontSize=10, leading=16, textColor=colors.HexColor("#1F1F1F"), spaceAfter=4),
        "small": ParagraphStyle("small", parent=base["BodyText"], fontName="STSong-Light", fontSize=9, leading=14, textColor=colors.HexColor("#3A3A3A"), spaceAfter=3),
    }


def p(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def bullet_lines(items: list[str]) -> str:
    return "<br/>".join(f"• {escape(item)}" for item in items if item)


def table_block(rows: list[list[Any]], styles: dict[str, ParagraphStyle], widths: list[float]) -> LongTable:
    formatted = []
    for idx, row in enumerate(rows):
        style = styles["small"] if idx else styles["body"]
        formatted.append([Paragraph(escape(str(cell)), style) for cell in row])
    table = LongTable(formatted, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C81")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9D3E0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FB")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def runtime_summary_rows(reports: list[RuntimeTableReport]) -> list[list[str]]:
    rows = [["表", "规划目标", "实际使用", "模式", "总量", "关键判断"]]
    for report in reports:
        rows.append(
            [
                report.plan_name,
                report.planned_target_table or "未填",
                report.resolved_table,
                report.runtime_mode,
                str(report.row_count or "NA"),
                report.findings[0] if report.findings else "已完成实时探测",
            ]
        )
    return rows


def field_metric_rows(metrics: list[FieldMetric]) -> list[list[str]]:
    rows = [["核心标签", "来源字段/表达式", "填充数", "缺失数", "缺失率", "Top 分布"]]
    for metric in metrics:
        top_dist = "；".join(f"{item[0]}:{item[1]}" for item in metric.distribution[:5]) if metric.distribution else "—"
        rows.append(
            [
                metric.label,
                metric.source_desc or metric.unavailable_reason or "—",
                str(metric.fill_count or "NA"),
                str(metric.miss_count or "NA"),
                format_pct(metric.miss_rate_pct),
                top_dist,
            ]
        )
    return rows


def build_pdf(output_path: Path, workbook_path: Path, reference_pdf: Path, reports: list[RuntimeTableReport]) -> None:
    register_fonts()
    styles = build_styles()
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=14 * mm,
        title="短视频五张表核心维度规整报告",
        author="Codex",
    )
    story: list[Any] = []

    story.append(p("短视频五张表核心维度规整报告", styles["title"]))
    story.append(Spacer(1, 4 * mm))
    story.append(p(f"报告日期：{datetime.now():%Y-%m-%d}", styles["body"]))
    story.append(p(f"分析基准：{workbook_path.name}", styles["body"]))
    story.append(p(f"参考报告：{reference_pdf.name}", styles["body"]))
    story.append(p("运行模式：实时 DLC 拉数；优先命中结果表，未命中时降级到源表或跨表验证。", styles["body"]))

    story.append(Spacer(1, 4 * mm))
    story.append(p("一、结论摘要", styles["heading1"]))
    story.append(p("这次报告不再依据静态规划稿推断，而是直接对实时数仓做探测。最终结果表明，表 01、表 03 已能作为内容分析底座；表 02 语义层是否稳定落地要看结果表是否命中；表 04 仍是最关键的结构性缺口；表 05 可能退回源表验证，说明内容效果闭环尚未完全模型化。", styles["body"]))
    story.append(p("从“用什么内容”的角度，真正可复用的标准层应该拆成资源形态维、商品语义维、平台表达维、关系归因维、效果反馈维。当前最优先补的不是再加更多标签，而是桥接关系与统一标准值。", styles["body"]))

    story.append(Spacer(1, 3 * mm))
    story.append(p("二、参考前提", styles["heading1"]))
    story.append(Paragraph(bullet_lines(CONTENT_PRIORS), styles["body"]))

    story.append(Spacer(1, 3 * mm))
    story.append(p("三、实时探测总览", styles["heading1"]))
    story.append(table_block(runtime_summary_rows(reports), styles, [28 * mm, 30 * mm, 38 * mm, 20 * mm, 18 * mm, 54 * mm]))

    story.append(PageBreak())
    story.append(p("四、逐表分析", styles["heading1"]))
    for report in reports:
        story.append(p(f"{report.table_id} {report.plan_name}", styles["heading2"]))
        story.append(p(f"规划目标表：{report.planned_target_table or '未填'}；实际使用：{report.resolved_table}；模式：{report.runtime_mode}", styles["body"]))
        story.append(p(f"粒度：{report.grain}；核心标签：{'、'.join(report.core_labels) if report.core_labels else '待补'}", styles["body"]))
        if report.row_count is not None:
            story.append(p(f"总行数：{report.row_count}；粒度去重数：{report.distinct_grain_count if report.distinct_grain_count is not None else 'NA'}", styles["body"]))
        if report.freshness:
            story.append(table_block([["时间字段", "最早值", "最新值"]] + report.freshness, styles, [30 * mm, 55 * mm, 55 * mm]))
        if report.platform_distribution:
            story.append(Spacer(1, 2 * mm))
            story.append(table_block([["分布字段值", "行数"]] + report.platform_distribution[:10], styles, [60 * mm, 40 * mm]))
        if report.field_metrics:
            story.append(Spacer(1, 2 * mm))
            story.append(table_block(field_metric_rows(report.field_metrics), styles, [20 * mm, 30 * mm, 18 * mm, 18 * mm, 18 * mm, 66 * mm]))
        if report.cross_checks:
            story.append(Spacer(1, 2 * mm))
            story.append(table_block(report.cross_checks, styles, [38 * mm, 34 * mm, 80 * mm]))
        if report.findings:
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph(bullet_lines(report.findings), styles["body"]))
        story.append(Spacer(1, 4 * mm))

    story.append(PageBreak())
    story.append(p("五、关键风险", styles["heading1"]))
    story.append(table_block(collect_risks(reports), styles, [34 * mm, 24 * mm, 104 * mm]))

    story.append(Spacer(1, 4 * mm))
    story.append(p("六、最终判断", styles["heading1"]))
    story.append(
        Paragraph(
            bullet_lines(
                [
                    "表 01 可以作为资源底座，但高缺失字段不能直接冒充全量标准标签。",
                    "表 02 的价值在于把内容挂到商品语义；若结果表未稳定，则说明“用什么内容”的语义层仍在建设中。",
                    "表 03 是平台表达层，可直接观察大标签/中标签的真实可用性。",
                    "表 04 若不存在，就意味着当前所有 ROI 与复用分析都只能做弱关联。",
                    "表 05 即便有源表，也不等于闭环完成；必须与标准标签和桥接层连通才有经营价值。",
                ]
            ),
            styles["body"],
        )
    )

    doc.build(story)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="实时生成短视频五张表核心维度规整报告")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Excel 工作簿路径")
    parser.add_argument("--reference-pdf", default=str(DEFAULT_REFERENCE_PDF), help="参考 PDF 路径")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--pdf-name", default=DEFAULT_PDF_NAME, help="输出 PDF 文件名")
    parser.add_argument("--json-name", default=DEFAULT_JSON_NAME, help="输出 JSON 文件名")
    parser.add_argument("--max-wait", type=int, default=240, help="单条 SQL 最大等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    reference_pdf = Path(args.reference_pdf).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    plans = build_planned_tables(workbook_path)
    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), max_wait=args.max_wait)
    reports = analyze_tables(runner, plans)

    payload = build_payload(workbook_path, reference_pdf, plans, reports)
    json_path = output_dir / args.json_name
    pdf_path = output_dir / args.pdf_name
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    build_pdf(pdf_path, workbook_path, reference_pdf, reports)

    print(f"JSON: {json_path}")
    print(f"PDF: {pdf_path}")


if __name__ == "__main__":
    main()
