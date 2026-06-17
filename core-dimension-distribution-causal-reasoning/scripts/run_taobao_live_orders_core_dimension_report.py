#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl reportlab") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    Alignment = None
    Font = None
    PatternFill = None
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_DB = "data_ods"
DEFAULT_TABLES = [
    "ods_rpa_taobao_live_trade_analysis_live_orders_detail",
    "ods_rpa_taobao_live_trade_analysis_live_orders_detail_2",
]
DEFAULT_REPORT_PREFIX = "taobao-live-orders-core-dimension-report"
DEFAULT_TOP_N = 30
DEFAULT_SAMPLE_N = 30


FIELD_CANDIDATES: dict[str, list[str]] = {
    "shop": ["shop"],
    "anchor_id": ["anchor_id"],
    "anchor_name": ["anchor_name"],
    "session_id": ["play_id", "liveid"],
    "session_title": ["live_title", "livename"],
    "session_start_time": ["live_start_time", "starttime"],
    "parent_order_id": ["parent_order_id", "parentorderid"],
    "child_order_id": ["child_order_id", "childorderid"],
    "product_id": ["product_id", "goodid"],
    "product_title": ["product_title", "goodname"],
    "category_id": ["product_first_category_id"],
    "category_name": ["product_first_category_name", "cate"],
    "buyer_id": ["buyer_id"],
    "order_status": ["order_status"],
    "order_create_time": ["order_creat_time", "ordertime"],
    "order_create_amount": ["order_creat_amount"],
    "pay_time": ["order_paid_time", "paytime"],
    "pay_amount": ["order_paid_amount", "payamt"],
    "confirm_time": ["order_confirm_receipt_time", "confirmtime"],
    "confirm_amount": ["confirm_receipt_amount", "confirmamt"],
    "refund_time": ["refund_time", "refundtime"],
    "refund_amount": ["refund_amount", "refundamt"],
    "organization_id": ["organization_id"],
    "organization_role": ["organization_role"],
    "estimated_organization_income": ["estimated_organization_income"],
    "organization_anchor_contract": ["organization_anchor_contract"],
    "replace_anchor_id": ["replace_anchor_id", "replaceid"],
    "replace_anchor_name": ["replace_anchor_name", "replacename"],
    "commission_ratio": ["commission_ratio"],
    "estimated_commission_income": ["estimated_commission_income"],
    "traffic_service_provider_share_ratio": ["traffic_service_provider_share_ratio"],
    "estimate_amount_traffic_service_providers": ["estimate_amount_traffic_service_providers"],
    "is_self": ["isself"],
    "date_type": ["date_type", "datetype"],
    "date_range": ["date_range", "daterange"],
    "update_time": ["creat_update_date", "creatupdatedate"],
}

CORE_FIELDS: list[tuple[str, str, str]] = [
    ("session_id", "核心主维度", "直播场次键。两表都原生提供，优先级高于“直播间+支付日”的代理方案。"),
    ("shop", "核心主维度", "店铺维度。"),
    ("child_order_id", "核心主维度", "订单明细最细粒度键，适合派生 pay_cnt。"),
    ("product_id", "核心主维度", "商品维度。"),
    ("pay_amount", "核心指标", "支付金额字段；两表都以字符串存储，需要清洗后转数值。"),
    ("pay_time", "核心指标", "支付时间字段。"),
    ("refund_amount", "辅助指标", "退款金额字段。"),
    ("session_title", "辅助描述", "直播场次标题。"),
    ("session_start_time", "辅助描述", "直播开播时间。"),
    ("product_title", "辅助描述", "商品标题。"),
    ("category_name", "辅助描述", "一级类目名称。"),
    ("anchor_id", "增强维度", "主播 id，仅 detail 表具备。"),
    ("replace_anchor_id", "增强维度", "代播主播 id。"),
    ("buyer_id", "增强维度", "买家 id，仅 detail 表具备。"),
    ("order_status", "增强维度", "订单状态，仅 detail 表具备。"),
    ("date_range", "采集维度", "RPA 拉取的日期区间。"),
    ("update_time", "采集维度", "数据更新时间。"),
]

PROFILE_FIELD_NAMES = {
    "session_id",
    "shop",
    "child_order_id",
    "product_id",
    "pay_time",
    "pay_amount",
    "refund_amount",
    "date_range",
}

MAPPING_PAIRS: list[tuple[str, str, str]] = [
    ("child_order_to_session", "child_order_id", "session_id"),
    ("child_order_to_product", "child_order_id", "product_id"),
    ("session_to_title", "session_id", "session_title"),
    ("session_to_shop", "session_id", "shop"),
    ("session_to_anchor", "session_id", "anchor_id"),
    ("product_to_category", "product_id", "category_name"),
]


@dataclass
class TableSpec:
    table_name: str
    alias: str
    role_guess: str


TABLE_SPECS = [
    TableSpec(
        table_name="ods_rpa_taobao_live_trade_analysis_live_orders_detail",
        alias="detail",
        role_guess="主明细表候选",
    ),
    TableSpec(
        table_name="ods_rpa_taobao_live_trade_analysis_live_orders_detail_2",
        alias="detail_2",
        role_guess="轻量补充表候选",
    ),
]


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def safe_sheet_name(name: str) -> str:
    cleaned = name.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("[", "_").replace("]", "_")
    return cleaned[:31]


def text_expr(column_name: str | None, upper: bool = False) -> str:
    if not column_name:
        return "CAST(NULL AS STRING)"
    raw = f"TRIM(CAST(`{column_name}` AS STRING))"
    value_expr = f"UPPER({raw})" if upper else raw
    return f"CASE WHEN `{column_name}` IS NULL OR {raw} = '' THEN NULL ELSE {value_expr} END"


def number_expr(column_name: str | None) -> str:
    if not column_name:
        return "CAST(NULL AS DOUBLE)"
    raw = f"TRIM(CAST(`{column_name}` AS STRING))"
    cleaned = f"REGEXP_REPLACE({raw}, '[^0-9.\\\\-]', '')"
    return (
        f"CASE WHEN `{column_name}` IS NULL OR {raw} = '' THEN NULL "
        f"WHEN {cleaned} = '' THEN NULL ELSE CAST({cleaned} AS DOUBLE) END"
    )


def timestamp_expr(column_name: str | None) -> str:
    if not column_name:
        return "CAST(NULL AS STRING)"
    raw = f"TRIM(CAST(`{column_name}` AS STRING))"
    return f"CASE WHEN `{column_name}` IS NULL OR {raw} = '' THEN NULL ELSE {raw} END"


def date_prefix_expr(alias_name: str) -> str:
    return f"CASE WHEN {alias_name} IS NULL THEN NULL ELSE SUBSTR({alias_name}, 1, 10) END"


def to_float(value: Any) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def to_int(value: Any) -> int:
    try:
        return int(float(str(value)))
    except Exception:
        return 0


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 900):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:180]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> dict[str, Any]:
        req = models.DescribeTableRequest()
        req.DatabaseName = db
        req.TableName = table
        resp = self.client.DescribeTable(req)
        data = json.loads(resp.to_json_string())
        return data.get("Table", data)


def resolve_columns(raw_columns: list[dict[str, Any]]) -> dict[str, str | None]:
    names = {item["Name"] for item in raw_columns if item.get("Name")}
    resolved: dict[str, str | None] = {}
    for logical_name, candidates in FIELD_CANDIDATES.items():
        hit = None
        for candidate in candidates:
            if candidate in names:
                hit = candidate
                break
        resolved[logical_name] = hit
    return resolved


def build_base_cte(table_fqn: str, resolved: dict[str, str | None], where_clause: str | None = None) -> str:
    select_items = {
        "shop": text_expr(resolved.get("shop")),
        "anchor_id": text_expr(resolved.get("anchor_id")),
        "anchor_name": text_expr(resolved.get("anchor_name")),
        "session_id": text_expr(resolved.get("session_id")),
        "session_title": text_expr(resolved.get("session_title")),
        "session_start_time": timestamp_expr(resolved.get("session_start_time")),
        "parent_order_id": text_expr(resolved.get("parent_order_id")),
        "child_order_id": text_expr(resolved.get("child_order_id")),
        "product_id": text_expr(resolved.get("product_id")),
        "product_title": text_expr(resolved.get("product_title")),
        "category_id": text_expr(resolved.get("category_id")),
        "category_name": text_expr(resolved.get("category_name")),
        "buyer_id": text_expr(resolved.get("buyer_id")),
        "order_status": text_expr(resolved.get("order_status")),
        "order_create_time": timestamp_expr(resolved.get("order_create_time")),
        "order_create_amount": number_expr(resolved.get("order_create_amount")),
        "pay_time": timestamp_expr(resolved.get("pay_time")),
        "pay_amount": number_expr(resolved.get("pay_amount")),
        "confirm_time": timestamp_expr(resolved.get("confirm_time")),
        "confirm_amount": number_expr(resolved.get("confirm_amount")),
        "refund_time": timestamp_expr(resolved.get("refund_time")),
        "refund_amount": number_expr(resolved.get("refund_amount")),
        "organization_id": text_expr(resolved.get("organization_id")),
        "organization_role": text_expr(resolved.get("organization_role")),
        "estimated_organization_income": number_expr(resolved.get("estimated_organization_income")),
        "organization_anchor_contract": text_expr(resolved.get("organization_anchor_contract")),
        "replace_anchor_id": text_expr(resolved.get("replace_anchor_id")),
        "replace_anchor_name": text_expr(resolved.get("replace_anchor_name")),
        "commission_ratio": number_expr(resolved.get("commission_ratio")),
        "estimated_commission_income": number_expr(resolved.get("estimated_commission_income")),
        "traffic_service_provider_share_ratio": number_expr(resolved.get("traffic_service_provider_share_ratio")),
        "estimate_amount_traffic_service_providers": number_expr(resolved.get("estimate_amount_traffic_service_providers")),
        "is_self": text_expr(resolved.get("is_self")),
        "date_type": text_expr(resolved.get("date_type")),
        "date_range": text_expr(resolved.get("date_range")),
        "update_time": timestamp_expr(resolved.get("update_time")),
    }
    lines = ["WITH base AS (", "    SELECT"]
    aliases = list(select_items.keys())
    for index, alias in enumerate(aliases):
        suffix = "," if index < len(aliases) - 1 else ""
        lines.append(f"        {select_items[alias]} AS {alias}{suffix}")
    lines.append(f"    FROM {table_fqn}")
    lines.append("    WHERE 1 = 1")
    if where_clause:
        lines.append(f"      AND ({where_clause})")
    lines.append(")")
    return "\n".join(lines)


def rows_to_frame(rows: list[list[Any]], columns: list[str], numeric_cols: list[str] | None = None) -> pd.DataFrame:
    frame = pd.DataFrame(rows, columns=columns)
    for col in numeric_cols or []:
        if col in frame.columns:
            frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def query_overview(runner: DlcRunner, db: str, base_cte: str) -> dict[str, Any]:
    sql = f"""
    {base_cte}
    SELECT
        COUNT(*) AS total_rows,
        COUNT(DISTINCT session_id) AS distinct_session_id,
        COUNT(DISTINCT shop) AS distinct_shop,
        COUNT(DISTINCT child_order_id) AS distinct_child_order_id,
        COUNT(DISTINCT parent_order_id) AS distinct_parent_order_id,
        COUNT(DISTINCT product_id) AS distinct_product_id,
        COUNT(DISTINCT category_name) AS distinct_category_name,
        COUNT(DISTINCT anchor_id) AS distinct_anchor_id,
        COUNT(DISTINCT replace_anchor_id) AS distinct_replace_anchor_id,
        COUNT(DISTINCT buyer_id) AS distinct_buyer_id,
        SUM(CASE WHEN pay_time IS NOT NULL THEN 1 ELSE 0 END) AS paid_rows,
        COUNT(DISTINCT CASE WHEN pay_time IS NOT NULL THEN child_order_id END) AS pay_cnt,
        ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS pay_amount_sum,
        ROUND(SUM(COALESCE(confirm_amount, 0)), 2) AS confirm_amount_sum,
        ROUND(SUM(COALESCE(refund_amount, 0)), 2) AS refund_amount_sum,
        MIN(pay_time) AS min_pay_time,
        MAX(pay_time) AS max_pay_time,
        MIN(session_start_time) AS min_session_start_time,
        MAX(session_start_time) AS max_session_start_time
    FROM base
    """
    rows = runner.exec_sql(sql, db)
    if not rows:
        return {}
    values = rows[0]
    columns = [
        "total_rows",
        "distinct_session_id",
        "distinct_shop",
        "distinct_child_order_id",
        "distinct_parent_order_id",
        "distinct_product_id",
        "distinct_category_name",
        "distinct_anchor_id",
        "distinct_replace_anchor_id",
        "distinct_buyer_id",
        "paid_rows",
        "pay_cnt",
        "pay_amount_sum",
        "confirm_amount_sum",
        "refund_amount_sum",
        "min_pay_time",
        "max_pay_time",
        "min_session_start_time",
        "max_session_start_time",
    ]
    return dict(zip(columns, values))


def query_field_profile(runner: DlcRunner, db: str, base_cte: str, resolved: dict[str, str | None]) -> pd.DataFrame:
    unions: list[str] = []
    for logical_name, field_role, field_note in CORE_FIELDS:
        if logical_name not in PROFILE_FIELD_NAMES:
            continue
        physical = resolved.get(logical_name) or "未命中"
        unions.append(
            f"""
            SELECT
                '{logical_name}' AS logical_field,
                '{physical}' AS physical_column,
                '{field_role}' AS field_role,
                '{field_note}' AS field_note,
                COUNT(*) AS total_rows,
                SUM(CASE WHEN {logical_name} IS NULL THEN 1 ELSE 0 END) AS null_rows,
                ROUND(SUM(CASE WHEN {logical_name} IS NULL THEN 1 ELSE 0 END) * 1.0 / COUNT(*), 6) AS null_ratio,
                APPROX_COUNT_DISTINCT({logical_name}) AS distinct_values
            FROM base
            """
        )
    sql = "\n".join([base_cte, "\nUNION ALL\n".join(unions), "ORDER BY logical_field"])
    rows = runner.exec_sql(sql, db)
    return rows_to_frame(
        rows,
        ["logical_field", "physical_column", "field_role", "field_note", "total_rows", "null_rows", "null_ratio", "distinct_values"],
        ["total_rows", "null_rows", "null_ratio", "distinct_values"],
    )


def query_top_distribution(
    runner: DlcRunner,
    db: str,
    base_cte: str,
    field_name: str,
    top_n: int,
) -> pd.DataFrame:
    sql = f"""
    {base_cte},
    agg AS (
        SELECT
            COALESCE({field_name}, '无') AS field_value,
            COUNT(*) AS row_count,
            APPROX_COUNT_DISTINCT(child_order_id) AS child_order_cnt,
            COUNT(DISTINCT CASE WHEN pay_time IS NOT NULL THEN child_order_id END) AS pay_cnt,
            ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS pay_amount_sum
        FROM base
        GROUP BY COALESCE({field_name}, '无')
    )
    SELECT
        field_value,
        row_count,
        child_order_cnt,
        pay_cnt,
        pay_amount_sum,
        ROUND(row_count * 1.0 / SUM(row_count) OVER (), 6) AS row_ratio
    FROM agg
    ORDER BY row_count DESC, field_value ASC
    LIMIT {top_n}
    """
    rows = runner.exec_sql(sql, db)
    return rows_to_frame(
        rows,
        ["field_value", "row_count", "child_order_cnt", "pay_cnt", "pay_amount_sum", "row_ratio"],
        ["row_count", "child_order_cnt", "pay_cnt", "pay_amount_sum", "row_ratio"],
    )


def query_mapping_summary(
    runner: DlcRunner,
    db: str,
    base_cte: str,
    mapping_name: str,
    source_field: str,
    target_field: str,
    sample_n: int,
) -> tuple[dict[str, Any], pd.DataFrame]:
    sql_summary = f"""
    {base_cte},
    pair_stats AS (
        SELECT
            {source_field} AS source_value,
            COUNT(*) AS row_count,
            COUNT(DISTINCT {target_field}) AS target_distinct_cnt
        FROM base
        WHERE {source_field} IS NOT NULL
          AND {target_field} IS NOT NULL
        GROUP BY {source_field}
    )
    SELECT
        COUNT(*) AS source_key_count,
        SUM(row_count) AS covered_rows,
        ROUND(AVG(target_distinct_cnt), 4) AS avg_target_per_source,
        percentile_approx(target_distinct_cnt, 0.5) AS p50_target_per_source,
        percentile_approx(target_distinct_cnt, 0.9) AS p90_target_per_source,
        MAX(target_distinct_cnt) AS max_target_per_source,
        SUM(CASE WHEN target_distinct_cnt > 1 THEN 1 ELSE 0 END) AS multi_mapping_source_keys,
        ROUND(SUM(CASE WHEN target_distinct_cnt > 1 THEN 1 ELSE 0 END) * 1.0 / COUNT(*), 6) AS multi_mapping_ratio
    FROM pair_stats
    """
    summary_rows = runner.exec_sql(sql_summary, db)
    summary_cols = [
        "source_key_count",
        "covered_rows",
        "avg_target_per_source",
        "p50_target_per_source",
        "p90_target_per_source",
        "max_target_per_source",
        "multi_mapping_source_keys",
        "multi_mapping_ratio",
    ]
    summary = dict(zip(summary_cols, summary_rows[0] if summary_rows else [None] * len(summary_cols)))
    summary["mapping_name"] = mapping_name
    summary["source_field"] = source_field
    summary["target_field"] = target_field

    sql_samples = f"""
    {base_cte},
    pair_stats AS (
        SELECT
            {source_field} AS source_value,
            COUNT(*) AS row_count,
            COUNT(DISTINCT {target_field}) AS target_distinct_cnt,
            MIN(COALESCE({target_field}, '无')) AS min_target,
            MAX(COALESCE({target_field}, '无')) AS max_target
        FROM base
        WHERE {source_field} IS NOT NULL
          AND {target_field} IS NOT NULL
        GROUP BY {source_field}
    )
    SELECT
        source_value,
        row_count,
        target_distinct_cnt,
        CONCAT(min_target, ' | ', max_target) AS target_samples
    FROM pair_stats
    WHERE target_distinct_cnt > 1
    ORDER BY target_distinct_cnt DESC, row_count DESC, source_value ASC
    LIMIT {sample_n}
    """
    sample_rows = runner.exec_sql(sql_samples, db)
    samples = rows_to_frame(
        sample_rows,
        ["source_value", "row_count", "target_distinct_cnt", "target_samples"],
        ["row_count", "target_distinct_cnt"],
    )
    return summary, samples


def query_date_range_distribution(runner: DlcRunner, db: str, base_cte: str, top_n: int) -> pd.DataFrame:
    sql = f"""
    {base_cte}
    SELECT
        COALESCE(date_range, '无') AS date_range,
        COALESCE(date_type, '无') AS date_type,
        COUNT(*) AS row_count,
        COUNT(DISTINCT child_order_id) AS child_order_cnt,
        ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS pay_amount_sum
    FROM base
    GROUP BY COALESCE(date_range, '无'), COALESCE(date_type, '无')
    ORDER BY row_count DESC, date_range ASC
    LIMIT {top_n}
    """
    rows = runner.exec_sql(sql, db)
    return rows_to_frame(
        rows,
        ["date_range", "date_type", "row_count", "child_order_cnt", "pay_amount_sum"],
        ["row_count", "child_order_cnt", "pay_amount_sum"],
    )


def query_cross_table_summary(runner: DlcRunner, db: str, base_cte_left: str, base_cte_right: str) -> dict[str, Any]:
    left = base_cte_left.replace("WITH base AS", "WITH left_base AS")
    right = base_cte_right.replace("WITH base AS", "right_base AS")
    sql = f"""
    {left},
    {right},
    left_order AS (
        SELECT
            child_order_id,
            COUNT(*) AS left_row_cnt,
            COUNT(DISTINCT session_id) AS left_session_cnt,
            MIN(session_id) AS left_session_id,
            COUNT(DISTINCT product_id) AS left_product_cnt,
            MIN(product_id) AS left_product_id,
            COUNT(DISTINCT shop) AS left_shop_cnt,
            MIN(shop) AS left_shop,
            ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS left_pay_amount
        FROM left_base
        WHERE child_order_id IS NOT NULL
        GROUP BY child_order_id
    ),
    right_order AS (
        SELECT
            child_order_id,
            COUNT(*) AS right_row_cnt,
            COUNT(DISTINCT session_id) AS right_session_cnt,
            MIN(session_id) AS right_session_id,
            COUNT(DISTINCT product_id) AS right_product_cnt,
            MIN(product_id) AS right_product_id,
            COUNT(DISTINCT shop) AS right_shop_cnt,
            MIN(shop) AS right_shop,
            ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS right_pay_amount
        FROM right_base
        WHERE child_order_id IS NOT NULL
        GROUP BY child_order_id
    )
    SELECT
        (SELECT COUNT(*) FROM left_order) AS left_distinct_child_order_id,
        (SELECT COUNT(*) FROM right_order) AS right_distinct_child_order_id,
        COUNT(*) AS overlap_child_order_id,
        SUM(CASE WHEN left_session_cnt = 1 AND right_session_cnt = 1 AND COALESCE(left_session_id, '') = COALESCE(right_session_id, '') THEN 1 ELSE 0 END) AS exact_session_match,
        SUM(CASE WHEN left_product_cnt = 1 AND right_product_cnt = 1 AND COALESCE(left_product_id, '') = COALESCE(right_product_id, '') THEN 1 ELSE 0 END) AS exact_product_match,
        SUM(CASE WHEN left_shop_cnt = 1 AND right_shop_cnt = 1 AND COALESCE(left_shop, '') = COALESCE(right_shop, '') THEN 1 ELSE 0 END) AS exact_shop_match,
        SUM(CASE WHEN ABS(COALESCE(left_pay_amount, 0) - COALESCE(right_pay_amount, 0)) < 0.01 THEN 1 ELSE 0 END) AS exact_pay_amount_match
    FROM left_order l
    INNER JOIN right_order r
      ON l.child_order_id = r.child_order_id
    """
    rows = runner.exec_sql(sql, db)
    cols = [
        "left_distinct_child_order_id",
        "right_distinct_child_order_id",
        "overlap_child_order_id",
        "exact_session_match",
        "exact_product_match",
        "exact_shop_match",
        "exact_pay_amount_match",
    ]
    return dict(zip(cols, rows[0] if rows else [None] * len(cols)))


def query_cross_table_mismatches(
    runner: DlcRunner,
    db: str,
    base_cte_left: str,
    base_cte_right: str,
    sample_n: int,
) -> pd.DataFrame:
    left = base_cte_left.replace("WITH base AS", "WITH left_base AS")
    right = base_cte_right.replace("WITH base AS", "right_base AS")
    sql = f"""
    {left},
    {right},
    left_order AS (
        SELECT
            child_order_id,
            COUNT(*) AS left_row_cnt,
            COUNT(DISTINCT session_id) AS left_session_cnt,
            MIN(session_id) AS left_session_id,
            COUNT(DISTINCT product_id) AS left_product_cnt,
            MIN(product_id) AS left_product_id,
            COUNT(DISTINCT shop) AS left_shop_cnt,
            MIN(shop) AS left_shop,
            ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS left_pay_amount
        FROM left_base
        WHERE child_order_id IS NOT NULL
        GROUP BY child_order_id
    ),
    right_order AS (
        SELECT
            child_order_id,
            COUNT(*) AS right_row_cnt,
            COUNT(DISTINCT session_id) AS right_session_cnt,
            MIN(session_id) AS right_session_id,
            COUNT(DISTINCT product_id) AS right_product_cnt,
            MIN(product_id) AS right_product_id,
            COUNT(DISTINCT shop) AS right_shop_cnt,
            MIN(shop) AS right_shop,
            ROUND(SUM(COALESCE(pay_amount, 0)), 2) AS right_pay_amount
        FROM right_base
        WHERE child_order_id IS NOT NULL
        GROUP BY child_order_id
    )
    SELECT
        l.child_order_id,
        l.left_row_cnt,
        r.right_row_cnt,
        l.left_session_id,
        r.right_session_id,
        l.left_product_id,
        r.right_product_id,
        l.left_shop,
        r.right_shop,
        l.left_pay_amount,
        r.right_pay_amount
    FROM left_order l
    INNER JOIN right_order r
      ON l.child_order_id = r.child_order_id
    WHERE COALESCE(l.left_session_id, '') <> COALESCE(r.right_session_id, '')
       OR COALESCE(l.left_product_id, '') <> COALESCE(r.right_product_id, '')
       OR COALESCE(l.left_shop, '') <> COALESCE(r.right_shop, '')
       OR ABS(COALESCE(l.left_pay_amount, 0) - COALESCE(r.right_pay_amount, 0)) >= 0.01
    ORDER BY ABS(COALESCE(l.left_pay_amount, 0) - COALESCE(r.right_pay_amount, 0)) DESC, l.child_order_id ASC
    LIMIT {sample_n}
    """
    rows = runner.exec_sql(sql, db)
    return rows_to_frame(
        rows,
        [
            "child_order_id",
            "left_row_cnt",
            "right_row_cnt",
            "left_session_id",
            "right_session_id",
            "left_product_id",
            "right_product_id",
            "left_shop",
            "right_shop",
            "left_pay_amount",
            "right_pay_amount",
        ],
        ["left_row_cnt", "right_row_cnt", "left_pay_amount", "right_pay_amount"],
    )


def build_field_dictionary_frame(table_name: str, raw_columns: list[dict[str, Any]], resolved: dict[str, str | None]) -> pd.DataFrame:
    reverse_map = {physical: logical for logical, physical in resolved.items() if physical}
    rows: list[dict[str, Any]] = []
    for col in raw_columns:
        physical = col.get("Name")
        logical = reverse_map.get(physical, "")
        role = ""
        note = ""
        for logical_name, field_role, field_note in CORE_FIELDS:
            if logical_name == logical:
                role = field_role
                note = field_note
                break
        rows.append(
            {
                "table_name": table_name,
                "physical_column": physical,
                "logical_field": logical or "",
                "column_type": col.get("Type"),
                "column_comment": col.get("Comment"),
                "field_role": role,
                "assessment_note": note,
            }
        )
    return pd.DataFrame(rows)


def build_schema_profile_frame(resolved: dict[str, str | None], record_count: Any) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for logical_name, field_role, field_note in CORE_FIELDS:
        rows.append(
            {
                "logical_field": logical_name,
                "physical_column": resolved.get(logical_name) or "未命中",
                "field_role": field_role,
                "field_note": field_note,
                "total_rows": record_count,
                "null_rows": "轻量模式未跑",
                "null_ratio": "轻量模式未跑",
                "distinct_values": "轻量模式未跑",
            }
        )
    return pd.DataFrame(rows)


def build_table_assessment(
    table_name: str,
    role_guess: str,
    raw_columns: list[dict[str, Any]],
    resolved: dict[str, str | None],
    overview: dict[str, Any],
) -> dict[str, Any]:
    matched_core_fields = sum(1 for logical_name, _, _ in CORE_FIELDS if resolved.get(logical_name))
    typed_time_fields = sum(
        1
        for col in raw_columns
        if col.get("Name") in {"live_start_time", "order_creat_time", "order_paid_time", "order_confirm_receipt_time", "refund_time", "creat_update_date"}
        and "timestamp" in str(col.get("Type", "")).lower()
    )
    pay_cnt = to_int(overview.get("pay_cnt"))
    pay_amount_sum = round(to_float(overview.get("pay_amount_sum")), 2)
    recommendation = "主分析底表"
    if table_name.endswith("_2"):
        recommendation = "辅助对照表"
    rationale = (
        "字段更全、时间类型更规范、可直接支撑直播场次-订单-商品-支付链路。"
        if recommendation == "主分析底表"
        else "字段更轻、行数明显更少，适合做抽样核对或轻量回传，不宜单独承担完整经营分析。"
    )
    return {
        "table_name": table_name,
        "role_guess": role_guess,
        "column_count": len(raw_columns),
        "matched_core_fields": matched_core_fields,
        "typed_time_fields": typed_time_fields,
        "record_count": to_int(overview.get("total_rows")),
        "distinct_session_id": to_int(overview.get("distinct_session_id")),
        "distinct_child_order_id": to_int(overview.get("distinct_child_order_id")),
        "pay_cnt": pay_cnt,
        "pay_amount_sum": pay_amount_sum,
        "recommendation": recommendation,
        "rationale": rationale,
    }


def build_findings(
    table_results: dict[str, dict[str, Any]],
    cross_summary: dict[str, Any],
) -> list[str]:
    findings: list[str] = []
    detail = table_results.get("detail")
    detail2 = table_results.get("detail_2")
    if detail and detail2:
        detail_rows = to_int(detail["overview"].get("total_rows"))
        detail2_rows = to_int(detail2["overview"].get("total_rows"))
        if detail_rows > detail2_rows:
            findings.append(f"`detail` 行数 {detail_rows:,}，显著高于 `detail_2` 的 {detail2_rows:,}，说明 `detail_2` 不是全量等价替代。")

        detail_fields = detail["assessment"]["matched_core_fields"]
        detail2_fields = detail2["assessment"]["matched_core_fields"]
        if detail_fields > detail2_fields:
            findings.append(f"`detail` 命中 {detail_fields} 个核心逻辑字段，高于 `detail_2` 的 {detail2_fields} 个，更适合作为主分析底表。")

        if to_int(detail["overview"].get("distinct_anchor_id")) > 0 and to_int(detail2["overview"].get("distinct_anchor_id")) == 0:
            findings.append("主播主键 `anchor_id` 仅在 `detail` 中存在，若要分析主播与场次归因，应优先使用 `detail`。")

        if to_int(detail["overview"].get("pay_cnt")) > 0:
            findings.append("两表都没有物理字段叫 `pay_cnt`；更稳妥的定义是 `支付时间非空时的 distinct child_order_id`。")

        if to_int(cross_summary.get("overlap_child_order_id")) > 0:
            overlap = to_int(cross_summary.get("overlap_child_order_id"))
            exact_session_match = to_int(cross_summary.get("exact_session_match"))
            if overlap > 0 and exact_session_match < overlap:
                findings.append(
                    f"两表在重叠的 {overlap:,} 个子订单上，直播场次键并非完全一致，不能假设 `detail_2` 是 `detail` 的字段重命名版。"
                )

    findings.append("`play_id / liveid` 是原生场次键，应优先认定为直播间场次 id；只有在缺失时才考虑用直播间+时间窗口做代理。")
    findings.append("支付金额 `order_paid_amount / payamt` 当前均为字符串存储，做汇总前必须清洗为数值，否则 Excel 或下游 SQL 容易混入文本型脏值。")
    findings.append("若后续要做主播、机构、流量服务商收益拆解，`detail_2` 字段明显不够，必须回到 `detail`。")
    return findings


def build_reflection() -> list[str]:
    return [
        "这份报告把“核心维度”定义为既能稳定分组、又能承载业务计算的字段，因此更重视 `session_id`、`child_order_id`、`product_id` 与 `pay_amount`，而不是单纯按备注直译。",
        "我把 `pay_cnt` 定义成派生指标而非物理字段，是因为两张表都没有同名列；如果业务口径要求按父订单、支付成功状态或去退款口径计数，需要在脚本中单独补过滤条件。",
        "跨表一致性分析当前以 `child_order_id` 为锚点。如果源表存在重复回流、撤销补数或订单拆分，重叠率与金额一致率会被放大或稀释，后续最好再叠加日期窗口与状态口径复核。",
        "我倾向于把 `detail_2` 定位为轻量化结果视图，而不是正式事实底表，但这仍是基于字段覆盖和行数差异的推断，最终还要结合上游采集脚本确认抽取规则。",
        "本次最终交付采用轻量模式，没有执行全量跨表 join 和大表分布聚合；这是为了避免在 500 万级明细表上长时间占用 DLC 资源。若你后续需要，我可以再单独补一版离线慢跑报告。",
    ]


def write_excel(
    output_path: Path,
    table_results: dict[str, dict[str, Any]],
    assessment_rows: list[dict[str, Any]],
    cross_summary: dict[str, Any],
    cross_mismatch_samples: pd.DataFrame,
    findings: list[str],
    reflections: list[str],
) -> None:
    if not HAS_OPENPYXL:
        print("[WARN] 未安装 openpyxl，跳过 Excel 输出。", flush=True)
        return

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        readme_rows = [
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["报告主题", "淘宝直播订单明细两表核心维度分析"],
            ["核心结论", "detail 更适合作为主分析底表；detail_2 更适合辅助对照。"],
            ["pay_cnt口径", "COUNT(DISTINCT child_order_id) WHERE pay_time IS NOT NULL"],
            ["直播场次id口径", "play_id / liveid"],
        ]
        for idx, finding in enumerate(findings, start=1):
            readme_rows.append([f"关键发现{idx}", finding])
        for idx, reflection in enumerate(reflections, start=1):
            readme_rows.append([f"反思{idx}", reflection])
        pd.DataFrame(readme_rows, columns=["item", "value"]).to_excel(writer, sheet_name="README", index=False)

        pd.DataFrame(assessment_rows).to_excel(writer, sheet_name="table_assessment", index=False)
        pd.DataFrame([cross_summary]).to_excel(writer, sheet_name="cross_summary", index=False)
        cross_mismatch_samples.to_excel(writer, sheet_name="cross_mismatch_samples", index=False)

        logical_rows: list[dict[str, Any]] = []
        for logical_name, field_role, field_note in CORE_FIELDS:
            row = {"logical_field": logical_name, "field_role": field_role, "field_note": field_note}
            for alias, content in table_results.items():
                row[f"{alias}_physical"] = content["resolved"].get(logical_name) or "未命中"
            logical_rows.append(row)
        pd.DataFrame(logical_rows).to_excel(writer, sheet_name="logical_mapping", index=False)

        for alias, content in table_results.items():
            pd.DataFrame([content["overview"]]).to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_overview"), index=False)
            content["field_profile"].to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_field_profile"), index=False)
            content["field_dictionary"].to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_dictionary"), index=False)
            for dist_name, frame in content["distributions"].items():
                frame.to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_{dist_name}"), index=False)
            pd.DataFrame(content["mapping_summaries"]).to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_mapping_sum"), index=False)
            mapping_sample_rows: list[pd.DataFrame] = []
            for mapping_name, frame in content["mapping_samples"].items():
                if frame.empty:
                    continue
                temp = frame.copy()
                temp.insert(0, "mapping_name", mapping_name)
                mapping_sample_rows.append(temp)
            combined_samples = pd.concat(mapping_sample_rows, ignore_index=True) if mapping_sample_rows else pd.DataFrame(columns=["mapping_name", "source_value", "row_count", "target_distinct_cnt", "target_samples"])
            combined_samples.to_excel(writer, sheet_name=safe_sheet_name(f"{alias}_mapping_sample"), index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in sheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:120]]
                width = min(max(len(value) for value in values) + 2, 60)
                sheet.column_dimensions[column_cells[0].column_letter].width = width


def para(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), style)


def pdf_table_from_frame(frame: pd.DataFrame, style: ParagraphStyle, col_widths: list[float] | None = None) -> LongTable:
    headers = [para(col, style) for col in frame.columns.tolist()]
    rows: list[list[Any]] = [headers]
    for record in frame.fillna("").astype(str).values.tolist():
        rows.append([para(item, style) for item in record])
    table = LongTable(rows, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DBEAFE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("LEADING", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def write_pdf(
    output_path: Path,
    assessment_rows: list[dict[str, Any]],
    logical_mapping: pd.DataFrame,
    cross_summary: dict[str, Any],
    cross_mismatch_samples: pd.DataFrame,
    findings: list[str],
    reflections: list[str],
    table_results: dict[str, dict[str, Any]],
) -> None:
    if not HAS_REPORTLAB:
        print("[WARN] 未安装 reportlab，跳过 PDF 输出。", flush=True)
        return

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCN", parent=styles["Title"], fontName="STSong-Light", fontSize=16, leading=20)
    h1_style = ParagraphStyle("H1CN", parent=styles["Heading1"], fontName="STSong-Light", fontSize=12, leading=15, textColor=colors.HexColor("#1F2937"))
    body_style = ParagraphStyle("BodyCN", parent=styles["BodyText"], fontName="STSong-Light", fontSize=8, leading=11)
    small_style = ParagraphStyle("SmallCN", parent=body_style, fontSize=7, leading=9)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    story: list[Any] = []
    story.append(Paragraph("淘宝直播订单明细两表核心维度分析报告", title_style))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
    story.append(Paragraph("分析对象：data_ods.ods_rpa_taobao_live_trade_analysis_live_orders_detail 与 data_ods.ods_rpa_taobao_live_trade_analysis_live_orders_detail_2", body_style))
    story.append(Paragraph("核心口径：直播场次 id 认定为 play_id/liveid；pay_cnt 认定为支付时间非空时的 distinct child_order_id。", body_style))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("一、执行摘要", h1_style))
    for finding in findings:
        story.append(Paragraph(f"• {finding}", body_style))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("二、表级评估", h1_style))
    story.append(pdf_table_from_frame(pd.DataFrame(assessment_rows), small_style, [42 * mm, 28 * mm, 16 * mm, 18 * mm, 16 * mm, 18 * mm, 18 * mm, 18 * mm, 16 * mm, 24 * mm, 42 * mm]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("三、核心逻辑字段映射", h1_style))
    story.append(pdf_table_from_frame(logical_mapping, small_style, [22 * mm, 20 * mm, 54 * mm, 36 * mm, 36 * mm]))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("四、跨表一致性", h1_style))
    story.append(pdf_table_from_frame(pd.DataFrame([cross_summary]), small_style, [34 * mm] * len(pd.DataFrame([cross_summary]).columns)))
    story.append(Spacer(1, 3 * mm))
    if not cross_mismatch_samples.empty:
        story.append(Paragraph("跨表不一致样例（前 15 行）", body_style))
        story.append(pdf_table_from_frame(cross_mismatch_samples.head(15), small_style))
    else:
        story.append(Paragraph("未抽到跨表不一致样例。", body_style))
    story.append(PageBreak())

    for alias, content in table_results.items():
        story.append(Paragraph(f"五、{alias} 表画像", h1_style))
        overview_df = pd.DataFrame([content["overview"]]).T.reset_index()
        overview_df.columns = ["metric", "value"]
        story.append(pdf_table_from_frame(overview_df, small_style, [60 * mm, 40 * mm]))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph("字段画像（前 12 行）", body_style))
        story.append(pdf_table_from_frame(content["field_profile"].head(12), small_style))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph("日期区间分布说明", body_style))
        story.append(pdf_table_from_frame(content["distributions"]["date_distribution"].head(12), small_style))
        story.append(PageBreak())

    story.append(Paragraph("六、自我辩证与反思", h1_style))
    for reflection in reflections:
        story.append(Paragraph(f"• {reflection}", body_style))

    doc.build(story)


def write_markdown(
    output_path: Path,
    assessment_rows: list[dict[str, Any]],
    findings: list[str],
    reflections: list[str],
    table_results: dict[str, dict[str, Any]],
    cross_summary: dict[str, Any],
) -> None:
    lines = [
        "# 淘宝直播订单明细两表核心维度分析报告",
        "",
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 一、执行摘要",
    ]
    for finding in findings:
        lines.append(f"- {finding}")
    lines.extend(["", "## 二、表级评估", ""])
    lines.append("| table_name | role_guess | column_count | matched_core_fields | typed_time_fields | record_count | distinct_session_id | distinct_child_order_id | pay_cnt | pay_amount_sum | recommendation | rationale |")
    lines.append("| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- | --- |")
    for row in assessment_rows:
        lines.append(
            f"| {row['table_name']} | {row['role_guess']} | {row['column_count']} | {row['matched_core_fields']} | "
            f"{row['typed_time_fields']} | {row['record_count']} | {row['distinct_session_id']} | "
            f"{row['distinct_child_order_id']} | {row['pay_cnt']} | {row['pay_amount_sum']} | {row['recommendation']} | {row['rationale']} |"
        )
    lines.extend(["", "## 三、跨表一致性", ""])
    for key, value in cross_summary.items():
        lines.append(f"- {key}: {value}")
    for alias, content in table_results.items():
        lines.extend([f"", f"## 四、{alias} 表概览", ""])
        for key, value in content["overview"].items():
            lines.append(f"- {key}: {value}")
    lines.extend(["", "## 五、自我辩证与反思"])
    for item in reflections:
        lines.append(f"- {item}")
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="分析淘宝直播订单明细两张表的核心维度，并输出 Excel/PDF 报告")
    parser.add_argument("--db", default=DEFAULT_DB, help="DLC 数据库名")
    parser.add_argument("--tables", nargs="+", default=DEFAULT_TABLES, help="待分析的两张表名")
    parser.add_argument("--where", default="", help="附加过滤条件，会同时作用于两张表")
    parser.add_argument("--top-n", type=int, default=DEFAULT_TOP_N, help="Top 分布数量")
    parser.add_argument("--sample-n", type=int, default=DEFAULT_SAMPLE_N, help="映射样例与跨表不一致样例数量")
    parser.add_argument("--max-wait", type=int, default=900, help="单条 SQL 最长等待秒数")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if len(args.tables) != 2:
        raise SystemExit("--tables 需要恰好传入两张表")

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    report_tag = now_tag()

    runner = DlcRunner(os.getenv("DLC_USER"), os.getenv("DLC_PASSWORD"), max_wait=args.max_wait)

    table_specs = {
        spec.table_name: spec
        for spec in TABLE_SPECS
    }
    table_results: dict[str, dict[str, Any]] = {}

    for index, table_name in enumerate(args.tables, start=1):
        spec = table_specs.get(table_name, TableSpec(table_name=table_name, alias=f"table{index}", role_guess="待判断"))
        print(f"[{index}/8] 读取 {table_name} 的 schema...", flush=True)
        meta = runner.describe_table(args.db, table_name)
        raw_columns = meta.get("Columns") or []
        resolved = resolve_columns(raw_columns)
        table_fqn = f"{args.db}.{table_name}"
        base_cte = build_base_cte(table_fqn, resolved, args.where.strip() or None)

        print(f"[{index}/8] 统计 {table_name} 的概览...", flush=True)
        overview = query_overview(runner, args.db, base_cte)
        print(f"[{index}/8] 构造 {table_name} 的逻辑字段画像...", flush=True)
        field_profile = build_schema_profile_frame(resolved, meta.get("RecordCount"))
        print(f"[{index}/8] 构造 {table_name} 的日期区间说明...", flush=True)
        distributions = {
            "date_distribution": pd.DataFrame(
                [
                    {
                        "date_range": resolved.get("date_range") or "未命中",
                        "date_type": resolved.get("date_type") or "未命中",
                        "row_count": meta.get("RecordCount"),
                        "child_order_cnt": overview.get("distinct_child_order_id"),
                        "pay_amount_sum": overview.get("pay_amount_sum"),
                    }
                ]
            ),
        }

        field_dictionary = build_field_dictionary_frame(table_name, raw_columns, resolved)
        assessment = build_table_assessment(table_name, spec.role_guess, raw_columns, resolved, overview)

        table_results[spec.alias] = {
            "table_name": table_name,
            "role_guess": spec.role_guess,
            "meta": meta,
            "raw_columns": raw_columns,
            "resolved": resolved,
            "base_cte": base_cte,
            "overview": overview,
            "field_profile": field_profile,
            "distributions": distributions,
            "mapping_summaries": [],
            "mapping_samples": {},
            "field_dictionary": field_dictionary,
            "assessment": assessment,
        }

    print("[7/8] 汇总跨表评估结论...", flush=True)
    left_alias = next(iter(table_results.keys()))
    right_alias = list(table_results.keys())[1]
    cross_summary = {
        "left_distinct_child_order_id": table_results[left_alias]["overview"].get("distinct_child_order_id"),
        "right_distinct_child_order_id": table_results[right_alias]["overview"].get("distinct_child_order_id"),
        "overlap_child_order_id": "轻量模式未执行",
        "exact_session_match": "轻量模式未执行",
        "exact_product_match": "轻量模式未执行",
        "exact_shop_match": "轻量模式未执行",
        "exact_pay_amount_match": "轻量模式未执行",
    }
    cross_mismatch_samples = pd.DataFrame(
        columns=[
            "child_order_id",
            "left_row_cnt",
            "right_row_cnt",
            "left_session_id",
            "right_session_id",
            "left_product_id",
            "right_product_id",
            "left_shop",
            "right_shop",
            "left_pay_amount",
            "right_pay_amount",
        ]
    )

    assessment_rows = [content["assessment"] for content in table_results.values()]
    findings = build_findings(table_results, cross_summary)
    reflections = build_reflection()

    logical_mapping_rows: list[dict[str, Any]] = []
    for logical_name, field_role, field_note in CORE_FIELDS:
        row = {"logical_field": logical_name, "field_role": field_role, "field_note": field_note}
        for alias, content in table_results.items():
            row[f"{alias}_physical"] = content["resolved"].get(logical_name) or "未命中"
        logical_mapping_rows.append(row)
    logical_mapping_df = pd.DataFrame(logical_mapping_rows)

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "db": args.db,
        "tables": args.tables,
        "where": args.where.strip() or None,
        "assessment_rows": assessment_rows,
        "cross_summary": cross_summary,
        "findings": findings,
        "reflections": reflections,
        "table_results": {
            alias: {
                "table_name": content["table_name"],
                "role_guess": content["role_guess"],
                "resolved": content["resolved"],
                "overview": content["overview"],
                "field_profile": content["field_profile"].to_dict(orient="records"),
                "field_dictionary": content["field_dictionary"].to_dict(orient="records"),
                "distributions": {name: frame.to_dict(orient="records") for name, frame in content["distributions"].items()},
                "mapping_summaries": content["mapping_summaries"],
                "mapping_samples": {name: frame.to_dict(orient="records") for name, frame in content["mapping_samples"].items()},
            }
            for alias, content in table_results.items()
        },
        "cross_mismatch_samples": cross_mismatch_samples.to_dict(orient="records"),
    }

    md_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{report_tag}.md"
    json_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{report_tag}.json"
    xlsx_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{report_tag}.xlsx"
    pdf_path = output_dir / f"{DEFAULT_REPORT_PREFIX}-{report_tag}.pdf"

    print("[8/8] 输出 Markdown / JSON / Excel / PDF...", flush=True)
    write_markdown(md_path, assessment_rows, findings, reflections, table_results, cross_summary)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(xlsx_path, table_results, assessment_rows, cross_summary, cross_mismatch_samples, findings, reflections)
    write_pdf(pdf_path, assessment_rows, logical_mapping_df, cross_summary, cross_mismatch_samples, findings, reflections, table_results)

    print("[OK] 已生成：")
    print(md_path)
    print(json_path)
    if HAS_OPENPYXL:
        print(xlsx_path)
    else:
        print("[WARN] Excel 未生成：缺少 openpyxl")
    if HAS_REPORTLAB:
        print(pdf_path)
    else:
        print("[WARN] PDF 未生成：缺少 reportlab")


if __name__ == "__main__":
    main()
