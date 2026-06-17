#!/usr/bin/env python3
"""
前台图片异常路径回溯脚本。

目的：
1. 给定一条前台展示路径，回溯 ODS -> DWS -> DIM 各层对应记录。
2. 比较“原始数据库值”和“前台展示值”的差异，判断异常是在源表、SQL 加工层，还是前台/接口格式化层引入。
3. 重点识别：
   - 前台路径的 /// 前缀
   - 前台路径中的 %20 等 URL 编码
   - 数据库字段本身是否也含 /、\、%20
   - 前台路径与数据库路径在“去分隔符 + URL 解码”后是否同源
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote
from xml.sax.saxutils import escape


os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
except ImportError as exc:
    raise SystemExit("缺少 reportlab，请先安装：python -m pip install reportlab") from exc

try:
    from tencentcloud.common import credential
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_PDF_NAME = "frontend-image-path-lineage-trace-{date_tag}.pdf"
DEFAULT_JSON_NAME = "frontend-image-path-lineage-trace-{date_tag}.json"
DEFAULT_XLSX_NAME = "frontend-image-path-lineage-trace-{date_tag}.xlsx"


@dataclass
class LayerSample:
    layer_name: str
    matched_rows: int
    rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class PathComparison:
    layer_name: str
    compare_field: str
    raw_value: str
    decoded_value: str
    compact_signature: str
    has_separator: bool
    has_pct_encoding: bool
    exact_match_frontend: bool
    decoded_match_frontend: bool
    compact_match_frontend: bool
    same_filename_frontend: bool


@dataclass
class ReportPayload:
    generated_at: str
    frontend_path: str
    frontend_decoded_path: str
    frontend_filename: str
    lookup_keyword: str
    key_findings: list[str]
    layer_samples: list[LayerSample]
    comparisons: list[PathComparison]
    recommendations: list[str]


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:150]
        print(f"[SQL] {db}: {preview}...", flush=True)
        task = models.Task()
        task.SparkSQLTask = {"SQL": base64.b64encode(sql.encode("utf-8")).decode("utf-8")}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def value_at(rows: list[list[Any]] | None, row: int = 0, col: int = 0, default: Any = None) -> Any:
    try:
        return rows[row][col]
    except Exception:
        return default


def to_int(value: Any) -> int:
    try:
        return int(str(value))
    except Exception:
        return 0


def sql_quote(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def extract_filename(path_value: str) -> str:
    decoded = unquote(path_value or "")
    if not decoded:
        return ""
    parts = re.split(r"[\\/]", decoded)
    return parts[-1] if parts else decoded


def compact_signature(path_value: str) -> str:
    decoded = unquote(path_value or "")
    return re.sub(r"[\\/]+", "", decoded)


def has_pct_encoding(path_value: str) -> bool:
    return bool(re.search(r"%[0-9a-fA-F]{2}", path_value or ""))


def has_separator(path_value: str) -> bool:
    value = path_value or ""
    return "/" in value or "\\" in value


def row_to_dict(columns: list[str], row: list[Any]) -> dict[str, Any]:
    return {col: row[idx] if idx < len(row) else None for idx, col in enumerate(columns)}


def build_lookup_keyword(frontend_path: str, explicit_keyword: str | None) -> str:
    if explicit_keyword:
        return explicit_keyword
    filename = extract_filename(frontend_path)
    if filename:
        return filename
    decoded = unquote(frontend_path or "")
    return decoded[-64:] if decoded else ""


def build_in_condition(values: list[str]) -> str:
    cleaned = [str(x) for x in values if str(x).strip()]
    if not cleaned:
        return ""
    return ", ".join(sql_quote(x) for x in cleaned)


def build_layer_samples(runner: DlcRunner, lookup_keyword: str, sample_limit: int) -> list[LayerSample]:
    like_value = "%" + lookup_keyword.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_") + "%"
    sql_value = sql_quote(like_value)
    source_specs = [
        {
            "layer_name": "ODS: ods_pic_for_up_new_backup",
            "db": "data_ods",
            "count_sql": f"""
                SELECT COUNT(*)
                FROM data_ods.ods_pic_for_up_new_backup
                WHERE is_delete = 'false'
                  AND (
                        image_name LIKE {sql_value} ESCAPE '\\\\'
                     OR image_full_path LIKE {sql_value} ESCAPE '\\\\'
                  )
            """,
            "sample_sql": f"""
                SELECT
                    CAST(id AS STRING) AS file_id,
                    image_name,
                    image_full_path,
                    content_type,
                    brand,
                    spu,
                    sku,
                    CAST(image_modify_time AS STRING) AS image_modify_time
                FROM data_ods.ods_pic_for_up_new_backup
                WHERE is_delete = 'false'
                  AND (
                        image_name LIKE {sql_value} ESCAPE '\\\\'
                     OR image_full_path LIKE {sql_value} ESCAPE '\\\\'
                  )
                ORDER BY image_modify_time DESC, file_id DESC
                LIMIT {sample_limit}
            """,
            "columns": ["file_id", "image_name", "image_full_path", "content_type", "brand", "spu", "sku", "image_modify_time"],
        },
        {
            "layer_name": "ODS: ods_t_image_file_information",
            "db": "data_ods",
            "count_sql": f"""
                SELECT COUNT(*)
                FROM data_ods.ods_t_image_file_information
                WHERE
                        file_name LIKE {sql_value} ESCAPE '\\\\'
                     OR folder_path LIKE {sql_value} ESCAPE '\\\\'
                     OR full_path LIKE {sql_value} ESCAPE '\\\\'
            """,
            "sample_sql": f"""
                SELECT
                    CAST(file_id AS STRING) AS file_id,
                    platform,
                    store,
                    spu,
                    file_name,
                    folder_path,
                    full_path,
                    file_format,
                    CAST(last_modify_time AS STRING) AS last_modify_time,
                    CAST(data_update_time AS STRING) AS data_update_time
                FROM data_ods.ods_t_image_file_information
                WHERE
                        file_name LIKE {sql_value} ESCAPE '\\\\'
                     OR folder_path LIKE {sql_value} ESCAPE '\\\\'
                     OR full_path LIKE {sql_value} ESCAPE '\\\\'
                ORDER BY data_update_time DESC, last_modify_time DESC, file_id DESC
                LIMIT {sample_limit}
            """,
            "columns": [
                "file_id",
                "platform",
                "store",
                "spu",
                "file_name",
                "folder_path",
                "full_path",
                "file_format",
                "last_modify_time",
                "data_update_time",
            ],
        },
    ]

    layer_samples: list[LayerSample] = []
    for spec in source_specs:
        matched_rows = to_int(value_at(runner.exec_sql(spec["count_sql"], spec["db"]), 0, 0, 0))
        rows = runner.exec_sql(spec["sample_sql"], spec["db"])
        layer_samples.append(
            LayerSample(
                layer_name=spec["layer_name"],
                matched_rows=matched_rows,
                rows=[row_to_dict(spec["columns"], row) for row in rows],
            )
        )

    anchor_file_ids: list[str] = []
    anchor_file_names: list[str] = []
    for layer in layer_samples:
        if layer.layer_name == "ODS: ods_t_image_file_information":
            for row in layer.rows:
                file_id = str(row.get("file_id") or "").strip()
                file_name = str(row.get("file_name") or "").strip()
                if file_id and file_id not in anchor_file_ids:
                    anchor_file_ids.append(file_id)
                if file_name and file_name not in anchor_file_names:
                    anchor_file_names.append(file_name)

    file_id_in_sql = build_in_condition(anchor_file_ids[:50])
    file_name_in_sql = build_in_condition(anchor_file_names[:20])
    dws_where_parts = ["file_type = '图片'"]
    dim_where_parts = []
    if file_id_in_sql:
        dws_where_parts.append(f"file_id IN ({file_id_in_sql})")
        dim_where_parts.append(f"file_id IN ({file_id_in_sql})")
    elif file_name_in_sql:
        dws_where_parts.append(f"file_name IN ({file_name_in_sql})")
        dim_where_parts.append(f"image_share_file_name IN ({file_name_in_sql})")
    if len(dws_where_parts) == 1:
        dws_where_parts.append(f"file_name LIKE {sql_value} ESCAPE '\\\\'")
        dws_where_parts.append(f"full_path LIKE {sql_value} ESCAPE '\\\\'")
    if not dim_where_parts:
        dim_where_parts.append(f"image_share_file_name LIKE {sql_value} ESCAPE '\\\\'")
        dim_where_parts.append(f"image_share_folder_path LIKE {sql_value} ESCAPE '\\\\'")
        dim_where_parts.append(f"full_path LIKE {sql_value} ESCAPE '\\\\'")

    downstream_specs = [
        {
            "layer_name": "DWS: dws_platform_file_resource_label_id",
            "db": "data_dws",
            "count_sql": f"""
                SELECT COUNT(*)
                FROM data_dws.dws_platform_file_resource_label_id
                WHERE {" AND ".join([dws_where_parts[0]])}
                  AND ({' OR '.join(dws_where_parts[1:])})
            """,
            "sample_sql": f"""
                SELECT
                    CAST(file_id AS STRING) AS file_id,
                    file_name,
                    full_path,
                    CAST(create_time AS STRING) AS create_time
                FROM data_dws.dws_platform_file_resource_label_id
                WHERE {" AND ".join([dws_where_parts[0]])}
                  AND ({' OR '.join(dws_where_parts[1:])})
                LIMIT {sample_limit}
            """,
            "columns": [
                "file_id",
                "file_name",
                "full_path",
                "create_time",
            ],
        },
        {
            "layer_name": "DIM: dim_picture_material_data_enriched",
            "db": "data_dim",
            "count_sql": f"""
                SELECT COUNT(*)
                FROM data_dim.dim_picture_material_data_enriched
                WHERE {' OR '.join(dim_where_parts)}
            """,
            "sample_sql": f"""
                SELECT
                    CAST(file_id AS STRING) AS file_id,
                    image_share_folder_path,
                    image_share_file_name,
                    full_path,
                    CAST(image_modify_time AS STRING) AS image_modify_time
                FROM data_dim.dim_picture_material_data_enriched
                WHERE {' OR '.join(dim_where_parts)}
                LIMIT {sample_limit}
            """,
            "columns": [
                "file_id",
                "image_share_folder_path",
                "image_share_file_name",
                "full_path",
                "image_modify_time",
            ],
        },
    ]

    for spec in downstream_specs:
        matched_rows = to_int(value_at(runner.exec_sql(spec["count_sql"], spec["db"]), 0, 0, 0))
        rows = runner.exec_sql(spec["sample_sql"], spec["db"])
        layer_samples.append(
            LayerSample(
                layer_name=spec["layer_name"],
                matched_rows=matched_rows,
                rows=[row_to_dict(spec["columns"], row) for row in rows],
            )
        )
    return layer_samples


def build_comparisons(frontend_path: str, layer_samples: list[LayerSample]) -> list[PathComparison]:
    frontend_decoded = unquote(frontend_path or "")
    frontend_filename = extract_filename(frontend_path)
    frontend_compact = compact_signature(frontend_path)
    comparisons: list[PathComparison] = []

    field_map = {
        "ODS: ods_pic_for_up_new_backup": ["image_full_path"],
        "ODS: ods_t_image_file_information": ["folder_path", "full_path"],
        "DWS: dws_platform_file_resource_label_id": ["full_path"],
        "DIM: dim_picture_material_data_enriched": ["image_share_folder_path", "full_path"],
    }

    for layer in layer_samples:
        for row in layer.rows[:8]:
            for field_name in field_map.get(layer.layer_name, []):
                raw_value = str(row.get(field_name) or "")
                if not raw_value:
                    continue
                decoded_value = unquote(raw_value)
                comparisons.append(
                    PathComparison(
                        layer_name=layer.layer_name,
                        compare_field=field_name,
                        raw_value=raw_value,
                        decoded_value=decoded_value,
                        compact_signature=compact_signature(raw_value),
                        has_separator=has_separator(raw_value),
                        has_pct_encoding=has_pct_encoding(raw_value),
                        exact_match_frontend=raw_value == frontend_path,
                        decoded_match_frontend=decoded_value == frontend_decoded,
                        compact_match_frontend=compact_signature(raw_value) == frontend_compact,
                        same_filename_frontend=extract_filename(raw_value) == frontend_filename,
                    )
                )
    return comparisons


def build_key_findings(frontend_path: str, layer_samples: list[LayerSample], comparisons: list[PathComparison]) -> list[str]:
    findings: list[str] = []
    findings.append(f"前台路径含分隔符：{'是' if has_separator(frontend_path) else '否'}；含 URL 编码：{'是' if has_pct_encoding(frontend_path) else '否'}。")

    for layer in layer_samples:
        findings.append(f"{layer.layer_name} 命中 {layer.matched_rows} 行。")

    separator_free_layers = []
    compact_match_layers = []
    for item in comparisons:
        if item.compact_match_frontend:
            compact_match_layers.append(f"{item.layer_name}.{item.compare_field}")
        if not item.has_separator and not item.has_pct_encoding:
            separator_free_layers.append(f"{item.layer_name}.{item.compare_field}")

    if compact_match_layers:
        findings.append("前台路径在“去分隔符 + URL 解码”后，与以下数据库字段同源：" + "；".join(compact_match_layers[:8]) + "。")
    if separator_free_layers:
        findings.append("命中的数据库字段原值均未出现前台那种 `/` 或 `%20` 形态，至少说明异常不是以相同字面值直接存进数据库：" + "；".join(separator_free_layers[:8]) + "。")

    source_match = [x for x in comparisons if x.layer_name.startswith("ODS") and x.compact_match_frontend]
    dws_match = [x for x in comparisons if x.layer_name.startswith("DWS") and x.compact_match_frontend]
    dim_match = [x for x in comparisons if x.layer_name.startswith("DIM") and x.compact_match_frontend]
    if source_match and dws_match and dim_match:
        findings.append("ODS、DWS、DIM 三层都能找到与前台路径同源的压扁字符串，说明数据库链路基本保持一致，异常更像发生在数据库之外的展示或接口层。")
    return findings


def build_recommendations(comparisons: list[PathComparison]) -> list[str]:
    recs = [
        "前台若需要展示共享盘路径，不应对数据库里的压扁路径做二次“补斜杠”拼装，除非掌握真实目录层级规则。",
        "若前台必须展示可读路径，应优先消费结构化字段，如 folder_path + file_name，而不是对 full_path 字符串做猜测式格式化。",
        "接口层若执行 URL encode，仅应编码 URL 参数，不应把普通文件展示名中的空格写成 `%20` 后再直接回显给用户。",
    ]
    if any(item.compact_match_frontend and not item.has_separator and not item.has_pct_encoding for item in comparisons):
        recs.append("当前证据支持：数据库侧存的是压扁路径，前台展示值中的 `///` 与 `%20` 是二次格式化产物，应重点排查前台渲染或接口适配逻辑。")
    return recs


def build_frames(payload: ReportPayload) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {
        "summary": pd.DataFrame({"key_finding": payload.key_findings}),
        "comparisons": pd.DataFrame([asdict(x) for x in payload.comparisons]),
        "layers": pd.DataFrame(
            [{"layer_name": x.layer_name, "matched_rows": x.matched_rows, "sample_rows": len(x.rows)} for x in payload.layer_samples]
        ),
    }
    for idx, layer in enumerate(payload.layer_samples, start=1):
        if layer.rows:
            frames[f"layer_{idx}"] = pd.DataFrame(layer.rows)
    return frames


def autofit_excel(writer: pd.ExcelWriter, frames: dict[str, pd.DataFrame]) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="top", wrap_text=True)
    for sheet_name, frame in frames.items():
        ws = writer.sheets[sheet_name[:31]]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        for idx, column in enumerate(frame.columns, start=1):
            max_len = len(str(column))
            if not frame.empty:
                sample_values = frame[column].head(200).tolist()
                if sample_values:
                    max_len = max(max_len, max(len(str(v)) for v in sample_values))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 60)


def build_pdf(payload: ReportPayload, pdf_path: Path) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_cn",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=18,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17375E"),
    )
    heading_style = ParagraphStyle(
        "heading_cn",
        parent=styles["Heading2"],
        fontName="STSong-Light",
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#1F4E78"),
        spaceBefore=8,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "body_cn",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=9.2,
        leading=13,
    )
    small_style = ParagraphStyle(
        "small_cn",
        parent=body_style,
        fontSize=8.0,
        leading=11,
    )

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    story: list[Any] = []
    story.append(Paragraph("前台图片异常路径回溯报告", title_style))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"生成时间：{escape(payload.generated_at)}", body_style))
    story.append(Paragraph(f"前台原始路径：{escape(payload.frontend_path)}", small_style))
    story.append(Paragraph(f"前台解码路径：{escape(payload.frontend_decoded_path)}", small_style))
    story.append(Paragraph(f"前台文件名：{escape(payload.frontend_filename)}", body_style))

    story.append(Paragraph("一、核心结论", heading_style))
    for item in payload.key_findings:
        story.append(Paragraph(f"• {escape(item)}", body_style))

    story.append(Paragraph("二、层级命中情况", heading_style))
    layer_rows = [["layer_name", "matched_rows"]]
    for item in payload.layer_samples:
        layer_rows.append([item.layer_name, str(item.matched_rows)])
    layer_table = LongTable(layer_rows, colWidths=[120 * mm, 35 * mm], repeatRows=1)
    layer_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C9D6")),
            ]
        )
    )
    story.append(layer_table)

    story.append(Paragraph("三、路径对比", heading_style))
    compare_rows = [["layer", "field", "has_sep", "has_%", "compact_match", "same_filename"]]
    for item in payload.comparisons[:24]:
        compare_rows.append(
            [
                item.layer_name,
                item.compare_field,
                "Y" if item.has_separator else "N",
                "Y" if item.has_pct_encoding else "N",
                "Y" if item.compact_match_frontend else "N",
                "Y" if item.same_filename_frontend else "N",
            ]
        )
    compare_table = LongTable(compare_rows, colWidths=[58 * mm, 28 * mm, 18 * mm, 18 * mm, 22 * mm, 22 * mm], repeatRows=1)
    compare_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.6),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C9D6")),
            ]
        )
    )
    story.append(compare_table)

    for layer in payload.layer_samples:
        if not layer.rows:
            continue
        story.append(PageBreak())
        story.append(Paragraph(layer.layer_name, heading_style))
        rows = [list(layer.rows[0].keys())] + [[str(x.get(col, "")) for col in layer.rows[0].keys()] for x in layer.rows]
        widths = [min(max(len(col), 10), 22) * 3.2 * mm for col in layer.rows[0].keys()]
        table = LongTable(rows, colWidths=widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 6.8),
                    ("LEADING", (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C9D6")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(table)

    doc.build(story)


def write_outputs(payload: ReportPayload, output_dir: Path, date_tag: str) -> tuple[Path, Path, Path]:
    pdf_path = output_dir / DEFAULT_PDF_NAME.format(date_tag=date_tag)
    json_path = output_dir / DEFAULT_JSON_NAME.format(date_tag=date_tag)
    xlsx_path = output_dir / DEFAULT_XLSX_NAME.format(date_tag=date_tag)

    json_path.write_text(
        json.dumps(
            {
                **asdict(payload),
                "layer_samples": [asdict(x) for x in payload.layer_samples],
                "comparisons": [asdict(x) for x in payload.comparisons],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    frames = build_frames(payload)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        autofit_excel(writer, frames)

    build_pdf(payload, pdf_path)
    return pdf_path, json_path, xlsx_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="前台图片异常路径回溯脚本")
    parser.add_argument("--frontend-path", required=True, help="前台看到的异常路径")
    parser.add_argument("--lookup-keyword", default="", help="可选；指定反查关键字，默认取前台路径最后的文件名")
    parser.add_argument("--sample-limit", type=int, default=20, help="每层抽样数量")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    frontend_path = args.frontend_path
    frontend_decoded = unquote(frontend_path)
    lookup_keyword = build_lookup_keyword(frontend_path, args.lookup_keyword or None)

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), args.max_wait)
    layer_samples = build_layer_samples(runner, lookup_keyword, args.sample_limit)
    comparisons = build_comparisons(frontend_path, layer_samples)
    payload = ReportPayload(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        frontend_path=frontend_path,
        frontend_decoded_path=frontend_decoded,
        frontend_filename=extract_filename(frontend_path),
        lookup_keyword=lookup_keyword,
        key_findings=build_key_findings(frontend_path, layer_samples, comparisons),
        layer_samples=layer_samples,
        comparisons=comparisons,
        recommendations=build_recommendations(comparisons),
    )
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path, json_path, xlsx_path = write_outputs(payload, output_dir, now_tag())
    print(f"[OK] PDF:  {pdf_path}")
    print(f"[OK] JSON: {json_path}")
    print(f"[OK] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
