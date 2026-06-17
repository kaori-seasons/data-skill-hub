#!/usr/bin/env python3
"""
表1 / 表2 / 表4 数据质量检测脚本

覆盖范围：
1. 表1：data_dwd.dwd_t_file_resource_id
2. 表2：data_dwd.dwd_file_label_id_spu
3. 表4：data_dws.dws_platform_file_resource_label_id

设计原则：
- 同时读取 four_table.pdf 与 “内容数据规划” sheet，确保质量标准不只来自现网 schema。
- 基于实时 DLC 数据做检测，兼顾字段质量、抽样、上下游血缘、数据新鲜度。
- 对表4额外按 four_table.pdf 的桥接设计，校验核心维度 file_id / platform_source_id / sku / spu / weight_factor。
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape


os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    from PyPDF2 import PdfReader
except ImportError as exc:
    raise SystemExit("缺少 PyPDF2，请先安装：python -m pip install PyPDF2") from exc

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
except ImportError as exc:
    raise SystemExit("缺少 reportlab，请先安装：python -m pip install reportlab") from exc

try:
    from tencentcloud.common import credential
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parents[1]
DEFAULT_WORKBOOK = WORKSPACE_DIR / "短视频数据汇总表_20260121_V2.xlsx"
DEFAULT_FOUR_TABLE_PDF = WORKSPACE_DIR / "four-table.pdf"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_PDF_NAME = "table124-quality-report-{date_tag}.pdf"
DEFAULT_JSON_NAME = "table124-quality-report-{date_tag}.json"
DEFAULT_XLSX_NAME = "table124-quality-report-{date_tag}.xlsx"


@dataclass
class PlannedTable:
    table_id: str
    plan_name: str
    core_labels: list[str]
    target_table: str
    source_tables: list[str]
    logic_desc: str
    code_excerpt: str


@dataclass
class SourceSnapshot:
    full_name: str
    exists: bool
    record_count_hint: int | None = None
    column_count: int | None = None
    key_columns_present: list[str] = field(default_factory=list)


@dataclass
class FieldMetric:
    field_name: str
    present_in_schema: bool
    fill_count: int | None
    miss_count: int | None
    miss_rate_pct: float | None
    distribution: list[list[Any]] = field(default_factory=list)
    note: str = ""


@dataclass
class CheckResult:
    check_name: str
    severity: str
    status: str
    detail: str
    row_count: int | None = None
    row_rate_pct: float | None = None


@dataclass
class SampleSet:
    title: str
    description: str
    columns: list[str]
    rows: list[list[Any]]


@dataclass
class TableReport:
    table_id: str
    plan_name: str
    target_table: str
    logic_desc: str
    code_excerpt: str
    source_tables: list[str]
    downstream_notes: list[str]
    schema_columns: list[str]
    row_count: int
    distinct_key_count: int | None
    key_definition: str
    freshness: list[list[str]]
    source_snapshots: list[SourceSnapshot]
    field_metrics: list[FieldMetric]
    checks: list[CheckResult]
    samples: list[SampleSet]


@dataclass
class Payload:
    generated_at: str
    workbook_path: str
    four_table_pdf_path: str
    four_table_expected_dims: list[str]
    four_table_pdf_excerpt: str
    summary_findings: list[str]
    table_reports: list[TableReport]


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def to_int(value: Any) -> int:
    try:
        return int(str(value))
    except Exception:
        return 0


def to_float(value: Any) -> float | None:
    try:
        return float(str(value))
    except Exception:
        return None


def value_at(rows: list[list[Any]] | None, row: int = 0, col: int = 0, default: Any = None) -> Any:
    try:
        return rows[row][col]
    except Exception:
        return default


def backtick(name: str) -> str:
    return f"`{name}`"


def non_empty_expr(name: str) -> str:
    return f"{backtick(name)} IS NOT NULL AND TRIM(CAST({backtick(name)} AS STRING)) <> ''"


def sql_quote(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        preview = " ".join(sql.strip().split())[:160]
        print(f"[SQL] {db}: {preview}...", flush=True)
        sql_b64 = base64.b64encode(sql.encode("utf-8")).decode("utf-8")
        task = models.Task()
        task.SparkSQLTask = {"SQL": sql_b64}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> dict[str, Any] | None:
        try:
            req = models.DescribeTableRequest()
            req.DatabaseName = db
            req.TableName = table
            resp = self.client.DescribeTable(req)
            data = json.loads(resp.to_json_string())
            return data.get("Table", data)
        except TencentCloudSDKException:
            return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").strip()


def split_lines(value: Any) -> list[str]:
    return [item.strip() for item in normalize_text(value).splitlines() if item and item.strip()]


def parse_core_labels(raw: str) -> list[str]:
    if not raw:
        return []
    text = str(raw).replace("【", "").replace("】", "")
    text = text.replace("（", "").replace("）", "")
    text = text.replace("(", "").replace(")", "")
    text = re.sub(r"[，,]", "、", text)
    return [part.strip() for part in text.split("、") if part and part.strip()]


def build_planned_tables(workbook_path: Path) -> dict[str, PlannedTable]:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["内容数据规划"]
    selected_rows = {"01": 3, "02": 4, "04": 6}
    tables: dict[str, PlannedTable] = {}
    for table_id, row in selected_rows.items():
        tables[table_id] = PlannedTable(
            table_id=table_id,
            plan_name=normalize_text(ws.cell(row, 1).value),
            core_labels=parse_core_labels(normalize_text(ws.cell(row, 2).value)),
            target_table=normalize_text(ws.cell(row, 3).value),
            source_tables=split_lines(ws.cell(row, 5).value),
            logic_desc=normalize_text(ws.cell(row, 8).value),
            code_excerpt=normalize_text(ws.cell(row, 7).value or ws.cell(row, 6).value)[:1800],
        )
    if not tables["04"].source_tables:
        tables["04"].source_tables = [
            "data_dwd.dwd_t_file_resource_id",
            "data_dwd.dwd_file_label_id_spu",
            "data_dwd.dwd_platform_source_label",
        ]
    return tables


def extract_four_table_pdf(pdf_path: Path) -> tuple[str, list[str]]:
    reader = PdfReader(str(pdf_path))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    dims = [dim for dim in ["file_id", "platform_source_id", "sku", "spu", "weight_factor"] if dim in text]
    excerpt = ""
    matched = re.search(r"字段设计(.+?)核心逻辑", text, flags=re.S)
    if matched:
        excerpt = re.sub(r"\s+", " ", matched.group(1)).strip()[:700]
    else:
        excerpt = re.sub(r"\s+", " ", text).strip()[:700]
    return excerpt, dims


def snapshot_sources(runner: DlcRunner, source_tables: list[str], key_columns: list[str]) -> list[SourceSnapshot]:
    snapshots: list[SourceSnapshot] = []
    for full_name in source_tables:
        if "." not in full_name:
            snapshots.append(SourceSnapshot(full_name=full_name, exists=False))
            continue
        db, table = full_name.split(".", 1)
        desc = runner.describe_table(db, table)
        if not desc:
            snapshots.append(SourceSnapshot(full_name=full_name, exists=False))
            continue
        columns = [col["Name"] for col in desc.get("Columns", [])]
        snapshots.append(
            SourceSnapshot(
                full_name=full_name,
                exists=True,
                record_count_hint=to_int(desc.get("RecordCount")) if desc.get("RecordCount") is not None else None,
                column_count=len(columns),
                key_columns_present=[col for col in key_columns if col in columns],
            )
        )
    return snapshots


def collect_freshness(runner: DlcRunner, db: str, table: str, columns: list[str]) -> list[list[str]]:
    freshness: list[list[str]] = []
    for col in columns:
        sql = f"""
        SELECT
            MIN(CAST({backtick(col)} AS STRING)),
            MAX(CAST({backtick(col)} AS STRING))
        FROM {db}.{table}
        WHERE {non_empty_expr(col)}
        """
        rows = runner.exec_sql(sql, db)
        freshness.append([col, str(value_at(rows, 0, 0, "NA")), str(value_at(rows, 0, 1, "NA"))])
    return freshness


def collect_fill_metric(
    runner: DlcRunner,
    db: str,
    table: str,
    field_name: str,
    columns: set[str],
    sample_dist: bool = False,
    where_clause: str = "1=1",
) -> FieldMetric:
    if field_name not in columns:
        return FieldMetric(field_name, False, None, None, None, [], "schema 缺失")
    sql = f"""
    SELECT
        COUNT(*) AS total_rows,
        SUM(CASE WHEN {non_empty_expr(field_name)} THEN 1 ELSE 0 END) AS fill_rows
    FROM {db}.{table}
    WHERE {where_clause}
    """
    rows = runner.exec_sql(sql, db)
    total = to_int(value_at(rows, 0, 0, 0))
    fill = to_int(value_at(rows, 0, 1, 0))
    miss = total - fill
    miss_rate = round(miss * 100.0 / total, 4) if total else 0.0
    distribution: list[list[Any]] = []
    if sample_dist:
        dist_sql = f"""
        SELECT CAST({backtick(field_name)} AS STRING) AS label_value, COUNT(*) AS cnt
        FROM {db}.{table}
        WHERE {where_clause} AND {non_empty_expr(field_name)}
        GROUP BY CAST({backtick(field_name)} AS STRING)
        ORDER BY cnt DESC, label_value
        LIMIT 8
        """
        distribution = runner.exec_sql(dist_sql, db) or []
    return FieldMetric(field_name, True, fill, miss, miss_rate, distribution)


def scalar_query(runner: DlcRunner, db: str, sql: str) -> int:
    return to_int(value_at(runner.exec_sql(sql, db), 0, 0, 0))


def build_table1_report(runner: DlcRunner, planned: PlannedTable) -> TableReport:
    db, table = "data_dwd", planned.target_table
    desc = runner.describe_table(db, table)
    if not desc:
        raise RuntimeError(f"未找到表 {db}.{table}")
    columns = {col["Name"] for col in desc.get("Columns", [])}
    row_count = scalar_query(runner, db, f"SELECT COUNT(*) FROM {db}.{table}")
    distinct_file_id = scalar_query(runner, db, f"SELECT COUNT(DISTINCT file_id) FROM {db}.{table}")
    duplicate_rows = max(row_count - distinct_file_id, 0)
    image_video_suffix = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table}
        WHERE file_type = '图片'
          AND lower(full_path) RLIKE '\\\\.(mp4|mov|avi|mkv|webm)$'
        """,
    )

    field_metrics = [
        collect_fill_metric(runner, db, table, field, columns, sample_dist=field in {"file_type", "platform", "picture_type"})
        for field in [
            "file_type",
            "file_id",
            "platform",
            "brand",
            "spu",
            "create_time",
            "full_path",
            "concatenated_path",
            "video_duration_type",
            "video_wide_range",
            "picture_size",
            "picture_wear",
            "picture_type",
            "publish_date",
        ]
    ]

    dup_sample_sql = f"""
    WITH dup AS (
        SELECT file_id, COUNT(*) AS cnt
        FROM {db}.{table}
        WHERE {non_empty_expr('file_id')}
        GROUP BY file_id
        HAVING COUNT(*) > 1
        ORDER BY cnt DESC, file_id
        LIMIT 12
    )
    SELECT t.file_id, t.file_type, t.platform, t.spu, t.create_time, t.publish_date, t.full_path, d.cnt
    FROM {db}.{table} t
    JOIN dup d ON t.file_id = d.file_id
    ORDER BY d.cnt DESC, t.file_id, t.create_time
    LIMIT 24
    """
    suffix_sample_sql = f"""
    SELECT file_id, file_type, platform, picture_type, picture_wear, full_path, create_time
    FROM {db}.{table}
    WHERE file_type = '图片'
      AND lower(full_path) RLIKE '\\\\.(mp4|mov|avi|mkv|webm)$'
    ORDER BY create_time DESC, file_id DESC
    LIMIT 20
    """

    checks = [
        CheckResult(
            check_name="file_id 唯一性",
            severity="high" if duplicate_rows > 0 else "low",
            status="fail" if duplicate_rows > 0 else "pass",
            detail=f"总行数 {row_count}，distinct file_id {distinct_file_id}，重复扩张 {duplicate_rows} 行。",
            row_count=duplicate_rows,
            row_rate_pct=round(duplicate_rows * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="图片链路混入视频后缀",
            severity="high" if image_video_suffix > 0 else "low",
            status="fail" if image_video_suffix > 0 else "pass",
            detail="表1同时承接图片与视频资源，若图片行路径带视频后缀，会直接污染下游桥接表与图片消费层。",
            row_count=image_video_suffix,
            row_rate_pct=round(image_video_suffix * 100.0 / row_count, 4) if row_count else 0.0,
        ),
    ]

    return TableReport(
        table_id=planned.table_id,
        plan_name=planned.plan_name,
        target_table=f"{db}.{table}",
        logic_desc=planned.logic_desc,
        code_excerpt=planned.code_excerpt,
        source_tables=planned.source_tables,
        downstream_notes=["下游供表2做 file_id / spu 语义补齐，下游供表4做桥接汇总。"],
        schema_columns=sorted(columns),
        row_count=row_count,
        distinct_key_count=distinct_file_id,
        key_definition="file_id",
        freshness=collect_freshness(runner, db, table, [col for col in ["create_time", "publish_date"] if col in columns]),
        source_snapshots=snapshot_sources(runner, [f"data_ods.{name}" if "." not in name else name for name in planned.source_tables], ["file_id", "spu", "publish_date"]),
        field_metrics=field_metrics,
        checks=checks,
        samples=[
            SampleSet(
                title="重复 file_id 样本",
                description="抽样展示表1中同一 file_id 被展开成多行的记录。",
                columns=["file_id", "file_type", "platform", "spu", "create_time", "publish_date", "full_path", "dup_cnt"],
                rows=runner.exec_sql(dup_sample_sql, db),
            ),
            SampleSet(
                title="图片行视频后缀样本",
                description="抽样展示 file_type='图片' 但 full_path 呈现视频后缀的记录。",
                columns=["file_id", "file_type", "platform", "picture_type", "picture_wear", "full_path", "create_time"],
                rows=runner.exec_sql(suffix_sample_sql, db),
            ),
        ],
    )


def build_table2_report(runner: DlcRunner, planned: PlannedTable) -> TableReport:
    db, table = "data_dwd", planned.target_table
    desc = runner.describe_table(db, table)
    if not desc:
        raise RuntimeError(f"未找到表 {db}.{table}")
    columns = {col["Name"] for col in desc.get("Columns", [])}
    row_count = scalar_query(runner, db, f"SELECT COUNT(*) FROM {db}.{table}")
    distinct_key = scalar_query(
        runner,
        db,
        f"SELECT COUNT(DISTINCT concat_ws('|', coalesce(file_id,''), coalesce(spu,''))) FROM {db}.{table}",
    )
    unmatched_t1 = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table} t2
        LEFT JOIN data_dwd.dwd_t_file_resource_id t1
          ON t2.file_id = t1.file_id
        WHERE {non_empty_expr('t2.file_id'.split('.')[-1])}
        """.replace("`file_id`", "t2.file_id"),
    )
    # safer explicit SQL for anti-join
    unmatched_t1 = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table} t2
        LEFT JOIN data_dwd.dwd_t_file_resource_id t1
          ON t2.file_id = t1.file_id
        WHERE t2.file_id IS NOT NULL
          AND TRIM(CAST(t2.file_id AS STRING)) <> ''
          AND t1.file_id IS NULL
        """,
    )
    freshness_proxy = runner.exec_sql(
        f"""
        SELECT
            MIN(CAST(t1.create_time AS STRING)),
            MAX(CAST(t1.create_time AS STRING))
        FROM {db}.{table} t2
        LEFT JOIN data_dwd.dwd_t_file_resource_id t1
          ON t2.file_id = t1.file_id
        WHERE t1.create_time IS NOT NULL
          AND TRIM(CAST(t1.create_time AS STRING)) <> ''
        """,
        db,
    )
    field_metrics = [
        collect_fill_metric(runner, db, table, field, columns, sample_dist=field in {"big_cate", "mid_cate", "sub_track", "gender", "scene", "style"})
        for field in ["file_id", "spu", "product_name", "big_cate", "mid_cate", "sub_track", "gender", "scene", "style"]
    ]

    missing_semantic_sql = f"""
    SELECT
        file_id,
        spu,
        product_name,
        big_cate,
        mid_cate,
        sub_track,
        gender,
        scene,
        style,
        (
            CASE WHEN product_name IS NULL OR TRIM(CAST(product_name AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN big_cate IS NULL OR TRIM(CAST(big_cate AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN mid_cate IS NULL OR TRIM(CAST(mid_cate AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN sub_track IS NULL OR TRIM(CAST(sub_track AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN gender IS NULL OR TRIM(CAST(gender AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN scene IS NULL OR TRIM(CAST(scene AS STRING)) = '' THEN 1 ELSE 0 END +
            CASE WHEN style IS NULL OR TRIM(CAST(style AS STRING)) = '' THEN 1 ELSE 0 END
        ) AS missing_semantic_cnt
    FROM {db}.{table}
    WHERE
        product_name IS NULL OR TRIM(CAST(product_name AS STRING)) = '' OR
        big_cate IS NULL OR TRIM(CAST(big_cate AS STRING)) = '' OR
        mid_cate IS NULL OR TRIM(CAST(mid_cate AS STRING)) = '' OR
        sub_track IS NULL OR TRIM(CAST(sub_track AS STRING)) = '' OR
        gender IS NULL OR TRIM(CAST(gender AS STRING)) = '' OR
        scene IS NULL OR TRIM(CAST(scene AS STRING)) = '' OR
        style IS NULL OR TRIM(CAST(style AS STRING)) = ''
    ORDER BY missing_semantic_cnt DESC, file_id, spu
    LIMIT 24
    """
    unmatched_sample_sql = f"""
    SELECT t2.file_id, t2.spu, t2.product_name, t2.big_cate, t2.mid_cate, t2.sub_track
    FROM {db}.{table} t2
    LEFT JOIN data_dwd.dwd_t_file_resource_id t1
      ON t2.file_id = t1.file_id
    WHERE t2.file_id IS NOT NULL
      AND TRIM(CAST(t2.file_id AS STRING)) <> ''
      AND t1.file_id IS NULL
    ORDER BY t2.file_id, t2.spu
    LIMIT 20
    """

    checks = [
        CheckResult(
            check_name="(file_id, spu) 唯一性",
            severity="high" if row_count != distinct_key else "low",
            status="fail" if row_count != distinct_key else "pass",
            detail=f"总行数 {row_count}，distinct(file_id,spu) {distinct_key}。",
            row_count=max(row_count - distinct_key, 0),
            row_rate_pct=round(max(row_count - distinct_key, 0) * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="file_id 上游回链表1",
            severity="high" if unmatched_t1 > 0 else "low",
            status="fail" if unmatched_t1 > 0 else "pass",
            detail="表2按规划应建立在表1 file_id 基础上补充商品语义；若回链失败，标签层粒度不可信。",
            row_count=unmatched_t1,
            row_rate_pct=round(unmatched_t1 * 100.0 / row_count, 4) if row_count else 0.0,
        ),
    ]

    return TableReport(
        table_id=planned.table_id,
        plan_name=planned.plan_name,
        target_table=f"{db}.{table}",
        logic_desc=planned.logic_desc,
        code_excerpt=planned.code_excerpt,
        source_tables=planned.source_tables,
        downstream_notes=["下游供表4做 file_id / spu 维度桥接。"],
        schema_columns=sorted(columns),
        row_count=row_count,
        distinct_key_count=distinct_key,
        key_definition="(file_id, spu)",
        freshness=[["upstream_create_time_proxy", str(value_at(freshness_proxy, 0, 0, 'NA')), str(value_at(freshness_proxy, 0, 1, 'NA'))]],
        source_snapshots=snapshot_sources(
            runner,
            [
                "data_dwd.dwd_t_file_resource_id",
                "data_dim.tb16_dim_product_sale_dimension",
                "data_ods.ods_dy_product_top_crowd_ays",
                "data_ods.ods_rpa_material_data",
            ],
            ["file_id", "spu", "商品链接id"],
        ),
        field_metrics=field_metrics,
        checks=checks,
        samples=[
            SampleSet(
                title="语义字段缺失样本",
                description="抽样展示表2中商品语义字段缺失最重的记录。",
                columns=["file_id", "spu", "product_name", "big_cate", "mid_cate", "sub_track", "gender", "scene", "style", "missing_semantic_cnt"],
                rows=runner.exec_sql(missing_semantic_sql, db),
            ),
            SampleSet(
                title="file_id 无法回链表1样本",
                description="抽样展示表2里在表1找不到对应 file_id 的记录。",
                columns=["file_id", "spu", "product_name", "big_cate", "mid_cate", "sub_track"],
                rows=runner.exec_sql(unmatched_sample_sql, db),
            ),
        ],
    )


def build_table4_report(
    runner: DlcRunner,
    planned: PlannedTable,
    expected_dims: list[str],
) -> TableReport:
    db, table = "data_dws", planned.target_table
    desc = runner.describe_table(db, table)
    if not desc:
        raise RuntimeError(f"未找到表 {db}.{table}")
    columns = {col["Name"] for col in desc.get("Columns", [])}
    row_count = scalar_query(runner, db, f"SELECT COUNT(*) FROM {db}.{table}")
    distinct_key = scalar_query(
        runner,
        db,
        f"SELECT COUNT(DISTINCT concat_ws('|', coalesce(file_id,''), coalesce(platform_source_id,''), coalesce(spu,''))) FROM {db}.{table}",
    )
    file_id_null = scalar_query(
        runner,
        db,
        f"SELECT COUNT(*) FROM {db}.{table} WHERE file_id IS NULL OR TRIM(CAST(file_id AS STRING)) = ''",
    )
    platform_source_null = scalar_query(
        runner,
        db,
        f"SELECT COUNT(*) FROM {db}.{table} WHERE platform_source_id IS NULL OR TRIM(CAST(platform_source_id AS STRING)) = ''",
    )
    spu_null = scalar_query(
        runner,
        db,
        f"SELECT COUNT(*) FROM {db}.{table} WHERE spu IS NULL OR TRIM(CAST(spu AS STRING)) = ''",
    )
    unmatched_t1 = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table} t4
        LEFT JOIN data_dwd.dwd_t_file_resource_id t1
          ON t4.file_id = t1.file_id
        WHERE t4.file_id IS NOT NULL
          AND TRIM(CAST(t4.file_id AS STRING)) <> ''
          AND t1.file_id IS NULL
        """,
    )
    unmatched_t2 = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table} t4
        LEFT JOIN data_dwd.dwd_file_label_id_spu t2
          ON t4.file_id = t2.file_id
         AND t4.spu = t2.spu
        WHERE t4.file_id IS NOT NULL
          AND TRIM(CAST(t4.file_id AS STRING)) <> ''
          AND t4.spu IS NOT NULL
          AND TRIM(CAST(t4.spu AS STRING)) <> ''
          AND t2.file_id IS NULL
        """,
    )
    unmatched_t3 = scalar_query(
        runner,
        db,
        f"""
        SELECT COUNT(*)
        FROM {db}.{table} t4
        LEFT JOIN data_dwd.dwd_platform_source_label t3
          ON t4.platform_source_id = t3.platform_source_id
        WHERE t4.platform_source_id IS NOT NULL
          AND TRIM(CAST(t4.platform_source_id AS STRING)) <> ''
          AND t3.platform_source_id IS NULL
        """,
    )
    field_metrics = [
        collect_fill_metric(runner, db, table, field, columns, sample_dist=field in {"file_type", "platform", "big_tag", "middle_tag"})
        for field in ["file_type", "file_id", "platform_source_id", "spu", "title", "author", "big_tag", "middle_tag", "publish_date"]
    ]
    for dim in expected_dims:
        if dim not in {metric.field_name for metric in field_metrics}:
            field_metrics.append(
                FieldMetric(dim, dim in columns, None, None, None, [], "来自 four_table.pdf 的核心桥接维度期望")
            )

    missing_platform_sql = f"""
    SELECT file_type, file_id, platform_source_id, spu, title, publish_date, full_path
    FROM {db}.{table}
    WHERE platform_source_id IS NULL OR TRIM(CAST(platform_source_id AS STRING)) = ''
    ORDER BY publish_date DESC, create_time DESC, file_id DESC
    LIMIT 24
    """
    unmatched_t2_sample_sql = f"""
    SELECT t4.file_type, t4.file_id, t4.platform_source_id, t4.spu, t4.title, t4.full_path
    FROM {db}.{table} t4
    LEFT JOIN data_dwd.dwd_file_label_id_spu t2
      ON t4.file_id = t2.file_id
     AND t4.spu = t2.spu
    WHERE t4.file_id IS NOT NULL
      AND TRIM(CAST(t4.file_id AS STRING)) <> ''
      AND t4.spu IS NOT NULL
      AND TRIM(CAST(t4.spu AS STRING)) <> ''
      AND t2.file_id IS NULL
    ORDER BY t4.publish_date DESC, t4.create_time DESC, t4.file_id DESC
    LIMIT 24
    """
    missing_file_id_sample_sql = f"""
    SELECT file_type, file_id, platform_source_id, spu, title, full_path, publish_date
    FROM {db}.{table}
    WHERE file_id IS NULL OR TRIM(CAST(file_id AS STRING)) = ''
    ORDER BY publish_date DESC, create_time DESC
    LIMIT 24
    """

    checks: list[CheckResult] = [
        CheckResult(
            check_name="桥接键唯一性",
            severity="high" if row_count != distinct_key else "low",
            status="fail" if row_count != distinct_key else "pass",
            detail="现网只能用 (file_id, platform_source_id, spu) 近似桥接键做唯一性校验，因为 sku 尚未落表。",
            row_count=max(row_count - distinct_key, 0),
            row_rate_pct=round(max(row_count - distinct_key, 0) * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="core dim: sku 已落表",
            severity="critical" if "sku" not in columns else "low",
            status="fail" if "sku" not in columns else "pass",
            detail="four_table.pdf 将 sku 定义为桥接核心维度键；当前实表缺失时，配色级归因无法落地。",
        ),
        CheckResult(
            check_name="core dim: weight_factor 已落表",
            severity="critical" if "weight_factor" not in columns else "low",
            status="fail" if "weight_factor" not in columns else "pass",
            detail="four_table.pdf 将 weight_factor 定义为归因权重；当前实表缺失时，多商品内容无法防止双重计量。",
        ),
        CheckResult(
            check_name="platform_source_id 非空覆盖率",
            severity="high" if platform_source_null > 0 else "low",
            status="fail" if platform_source_null > 0 else "pass",
            detail="若 platform_source_id 为空，文件资产无法稳定挂到平台发布实例。",
            row_count=platform_source_null,
            row_rate_pct=round(platform_source_null * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="file_id 非空覆盖率",
            severity="high" if file_id_null > 0 else "low",
            status="fail" if file_id_null > 0 else "pass",
            detail="桥接表若 file_id 为空，将无法回溯到本地原始素材。",
            row_count=file_id_null,
            row_rate_pct=round(file_id_null * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="spu 非空覆盖率",
            severity="medium" if spu_null > 0 else "low",
            status="fail" if spu_null > 0 else "pass",
            detail="spu 为空会削弱款级聚合能力。",
            row_count=spu_null,
            row_rate_pct=round(spu_null * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="回链表1文件资源层",
            severity="high" if unmatched_t1 > 0 else "low",
            status="fail" if unmatched_t1 > 0 else "pass",
            detail="表4按规划应由表1、表2、表3桥接而来；若 file_id 找不到表1，对应资产实体不完整。",
            row_count=unmatched_t1,
            row_rate_pct=round(unmatched_t1 * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="回链表2语义层",
            severity="high" if unmatched_t2 > 0 else "low",
            status="fail" if unmatched_t2 > 0 else "pass",
            detail="表4若无法回链表2，则 file_id / spu 桥接后的商品语义不完整。",
            row_count=unmatched_t2,
            row_rate_pct=round(unmatched_t2 * 100.0 / row_count, 4) if row_count else 0.0,
        ),
        CheckResult(
            check_name="回链表3平台表达层",
            severity="medium" if unmatched_t3 > 0 else "low",
            status="fail" if unmatched_t3 > 0 else "pass",
            detail="platform_source_id 若回链不到表3，则发布实例维度未闭环。",
            row_count=unmatched_t3,
            row_rate_pct=round(unmatched_t3 * 100.0 / row_count, 4) if row_count else 0.0,
        ),
    ]

    return TableReport(
        table_id=planned.table_id,
        plan_name=planned.plan_name,
        target_table=f"{db}.{table}",
        logic_desc=planned.logic_desc or "规划上属于 1/2/3 的关联桥接表。",
        code_excerpt=planned.code_excerpt,
        source_tables=planned.source_tables,
        downstream_notes=[
            "下游承接图片/视频聚合层与后续内容消费层；current-version.image-update.sql 也直接消费该表。",
            "若表4桥接键不稳定，会直接传导到图片/视频标签结果表。",
        ],
        schema_columns=sorted(columns),
        row_count=row_count,
        distinct_key_count=distinct_key,
        key_definition="当前实表近似键：(file_id, platform_source_id, spu)；规划键：file_id + platform_source_id + sku",
        freshness=collect_freshness(runner, db, table, [col for col in ["create_time", "publish_date"] if col in columns]),
        source_snapshots=snapshot_sources(
            runner,
            planned.source_tables,
            ["file_id", "platform_source_id", "spu", "sku", "weight_factor"],
        ),
        field_metrics=field_metrics,
        checks=checks,
        samples=[
            SampleSet(
                title="platform_source_id 为空样本",
                description="抽样展示桥接表中平台素材实例缺失的记录。",
                columns=["file_type", "file_id", "platform_source_id", "spu", "title", "publish_date", "full_path"],
                rows=runner.exec_sql(missing_platform_sql, db),
            ),
            SampleSet(
                title="无法回链表2的桥接样本",
                description="抽样展示桥接表中 file_id + spu 在表2缺失的记录。",
                columns=["file_type", "file_id", "platform_source_id", "spu", "title", "full_path"],
                rows=runner.exec_sql(unmatched_t2_sample_sql, db),
            ),
            SampleSet(
                title="file_id 为空样本",
                description="抽样展示桥接表中 file_id 为空的记录。",
                columns=["file_type", "file_id", "platform_source_id", "spu", "title", "full_path", "publish_date"],
                rows=runner.exec_sql(missing_file_id_sample_sql, db),
            ),
        ],
    )


def build_summary_findings(table_reports: list[TableReport]) -> list[str]:
    findings: list[str] = []
    lookup = {item.table_id: item for item in table_reports}
    t1 = lookup.get("01")
    t2 = lookup.get("02")
    t4 = lookup.get("04")
    if t1:
        for check in t1.checks:
            if check.check_name == "file_id 唯一性" and check.row_count:
                findings.append(
                    f"表1 file_id 非唯一，重复扩张 {check.row_count} 行，说明资源层还不是稳定的一条文件一行。"
                )
            if check.check_name == "图片链路混入视频后缀" and check.row_count:
                findings.append(
                    f"表1 图片子集中仍有 {check.row_count} 行视频后缀路径，会直接污染表4和图片消费层。"
                )
    if t2:
        for check in t2.checks:
            if check.check_name == "file_id 上游回链表1" and check.row_count:
                findings.append(f"表2 有 {check.row_count} 行无法回链表1，说明语义层与资源层存在断点。")
    if t4:
        for check in t4.checks:
            if check.check_name in {"core dim: sku 已落表", "core dim: weight_factor 已落表"} and check.status == "fail":
                findings.append(check.detail)
            if check.check_name == "platform_source_id 非空覆盖率" and check.row_count:
                findings.append(
                    f"表4 platform_source_id 为空 {check.row_count} 行，占比 {check.row_rate_pct:.2f}%，桥接表尚未稳定承接平台发布实例。"
                )
            if check.check_name == "file_id 非空覆盖率" and check.row_count:
                findings.append(
                    f"表4 file_id 为空 {check.row_count} 行，占比 {check.row_rate_pct:.2f}%，桥接层无法稳定回溯本地资产。"
                )
    return findings


def build_pdf(payload: Payload, pdf_path: Path) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_cn",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=18,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17375E"),
    )
    heading_style = ParagraphStyle(
        "heading_cn",
        parent=styles["Heading2"],
        fontName="STSong-Light",
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#1F4E78"),
        spaceBefore=8,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "body_cn",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=9.3,
        leading=14,
        spaceAfter=4,
    )
    small_style = ParagraphStyle(
        "small_cn",
        parent=body_style,
        fontSize=8.2,
        leading=12,
    )
    story: list[Any] = []
    story.append(Paragraph("表1 / 表2 / 表4 数据质量检测报告", title_style))
    story.append(
        Paragraph(
            escape(
                f"生成时间：{payload.generated_at}；规划稿：{Path(payload.workbook_path).name}；桥接设计稿：{Path(payload.four_table_pdf_path).name}"
            ),
            body_style,
        )
    )
    story.append(Paragraph("一、核心结论", heading_style))
    for item in payload.summary_findings:
        story.append(Paragraph(f"• {escape(item)}", body_style))

    story.append(Paragraph("二、表4设计期望（来自 four_table.pdf）", heading_style))
    story.append(Paragraph(escape("核心桥接维度：" + ", ".join(payload.four_table_expected_dims)), body_style))
    story.append(Paragraph(escape(payload.four_table_pdf_excerpt), small_style))

    for report in payload.table_reports:
        story.append(PageBreak())
        story.append(Paragraph(f"{report.table_id} {report.plan_name}", heading_style))
        story.append(Paragraph(escape(f"目标表：{report.target_table}"), body_style))
        story.append(Paragraph(escape(f"粒度/键定义：{report.key_definition}"), body_style))
        story.append(Paragraph(escape(f"上游来源：{'; '.join(report.source_tables)}"), small_style))
        for note in report.downstream_notes:
            story.append(Paragraph(f"• {escape(note)}", small_style))

        stats_rows = [
            ["指标", "值"],
            ["总行数", str(report.row_count)],
            ["distinct key", str(report.distinct_key_count) if report.distinct_key_count is not None else "NA"],
        ]
        stats_table = LongTable(stats_rows, repeatRows=1, colWidths=[50 * mm, 40 * mm])
        stats_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9FBAD0")),
                ]
            )
        )
        story.append(stats_table)
        story.append(Spacer(1, 2 * mm))

        if report.freshness:
            freshness_rows = [["字段", "最早", "最新"]] + report.freshness
            freshness_table = LongTable(freshness_rows, repeatRows=1, colWidths=[36 * mm, 50 * mm, 50 * mm])
            freshness_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2F0D9")),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A8C68F")),
                    ]
                )
            )
            story.append(Paragraph("新鲜度", body_style))
            story.append(freshness_table)

        if report.source_snapshots:
            source_rows = [["上游表", "存在", "记录数hint", "列数", "关键列"]]
            for item in report.source_snapshots:
                source_rows.append(
                    [
                        item.full_name,
                        "Y" if item.exists else "N",
                        "" if item.record_count_hint is None else str(item.record_count_hint),
                        "" if item.column_count is None else str(item.column_count),
                        ", ".join(item.key_columns_present),
                    ]
                )
            source_table = LongTable(source_rows, repeatRows=1, colWidths=[58 * mm, 12 * mm, 24 * mm, 16 * mm, 56 * mm])
            source_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7.6),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EDEDED")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9C9C9")),
                    ]
                )
            )
            story.append(Paragraph("上游血缘快照", body_style))
            story.append(source_table)

        story.append(Paragraph("质量检查", body_style))
        check_rows = [["检查项", "级别", "状态", "行数", "占比", "说明"]]
        for check in report.checks:
            pct = "" if check.row_rate_pct is None else f"{check.row_rate_pct:.2f}%"
            check_rows.append(
                [
                    check.check_name,
                    check.severity,
                    check.status,
                    "" if check.row_count is None else str(check.row_count),
                    pct,
                    check.detail,
                ]
            )
        check_table = LongTable(check_rows, repeatRows=1, colWidths=[42 * mm, 18 * mm, 16 * mm, 16 * mm, 16 * mm, 80 * mm])
        check_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.8),
                    ("LEADING", (0, 0), (-1, -1), 9.5),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FCE4D6")),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E6B8A2")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(check_table)

        story.append(Paragraph("字段质量", body_style))
        metric_rows = [["字段", "schema", "填充", "缺失", "缺失率", "备注"]]
        for metric in report.field_metrics:
            metric_rows.append(
                [
                    metric.field_name,
                    "Y" if metric.present_in_schema else "N",
                    "" if metric.fill_count is None else str(metric.fill_count),
                    "" if metric.miss_count is None else str(metric.miss_count),
                    "" if metric.miss_rate_pct is None else f"{metric.miss_rate_pct:.2f}%",
                    metric.note,
                ]
            )
        metric_table = LongTable(metric_rows, repeatRows=1, colWidths=[32 * mm, 12 * mm, 18 * mm, 18 * mm, 18 * mm, 78 * mm])
        metric_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9C9C9")),
                ]
            )
        )
        story.append(metric_table)

        for sample in report.samples:
            if not sample.rows:
                continue
            story.append(Paragraph(sample.title, body_style))
            story.append(Paragraph(escape(sample.description), small_style))
            sample_rows = [sample.columns] + sample.rows[:5]
            sample_table = LongTable(
                sample_rows,
                repeatRows=1,
                colWidths=[min(max(18, 160 / max(len(sample.columns), 1)), 45) * mm for _ in sample.columns],
            )
            sample_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7.2),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F7F7F7")),
                        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#D0D0D0")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(sample_table)

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    doc.build(story)


def build_sample_frames(table_reports: list[TableReport]) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for report in table_reports:
        check_frame = pd.DataFrame([asdict(item) for item in report.checks])
        metric_frame = pd.DataFrame([asdict(item) for item in report.field_metrics])
        source_frame = pd.DataFrame([asdict(item) for item in report.source_snapshots])
        frames[f"{report.table_id}_checks"] = check_frame
        frames[f"{report.table_id}_metrics"] = metric_frame
        frames[f"{report.table_id}_sources"] = source_frame
        for idx, sample in enumerate(report.samples, start=1):
            frames[f"{report.table_id}_sample_{idx}"] = pd.DataFrame(sample.rows, columns=sample.columns)
    return frames


def autofit_excel(writer, sheet_to_frame: dict[str, pd.DataFrame]) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="top", wrap_text=True)
    for sheet_name, frame in sheet_to_frame.items():
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        for idx, column in enumerate(frame.columns, start=1):
            max_len = len(str(column))
            if not frame.empty:
                sample_values = frame[column].head(200).tolist()
                if sample_values:
                    max_len = max(max_len, max(len(str(v)) for v in sample_values))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 60)


def write_outputs(payload: Payload, output_dir: Path, date_tag: str) -> tuple[Path, Path, Path]:
    pdf_path = output_dir / DEFAULT_PDF_NAME.format(date_tag=date_tag)
    json_path = output_dir / DEFAULT_JSON_NAME.format(date_tag=date_tag)
    xlsx_path = output_dir / DEFAULT_XLSX_NAME.format(date_tag=date_tag)

    json_path.write_text(
        json.dumps(
            {
                **asdict(payload),
                "table_reports": [asdict(item) for item in payload.table_reports],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    frames = build_sample_frames(payload.table_reports)
    frames["summary"] = pd.DataFrame({"summary_finding": payload.summary_findings})
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        autofit_excel(writer, {k[:31]: v for k, v in frames.items()})

    build_pdf(payload, pdf_path)
    return pdf_path, json_path, xlsx_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成表1/表2/表4数据质量检测报告")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="规划 Excel 路径")
    parser.add_argument("--four-table-pdf", default=str(DEFAULT_FOUR_TABLE_PDF), help="表4设计 PDF 路径")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    date_tag = now_tag()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    planned_tables = build_planned_tables(Path(args.workbook))
    pdf_excerpt, expected_dims = extract_four_table_pdf(Path(args.four_table_pdf))

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), args.max_wait)
    reports = [
        build_table1_report(runner, planned_tables["01"]),
        build_table2_report(runner, planned_tables["02"]),
        build_table4_report(runner, planned_tables["04"], expected_dims),
    ]
    payload = Payload(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        workbook_path=str(Path(args.workbook).resolve()),
        four_table_pdf_path=str(Path(args.four_table_pdf).resolve()),
        four_table_expected_dims=expected_dims,
        four_table_pdf_excerpt=pdf_excerpt,
        summary_findings=build_summary_findings(reports),
        table_reports=reports,
    )
    pdf_path, json_path, xlsx_path = write_outputs(payload, output_dir, date_tag)
    print(f"[OK] PDF:  {pdf_path}")
    print(f"[OK] JSON: {json_path}")
    print(f"[OK] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
