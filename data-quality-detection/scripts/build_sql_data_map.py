#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

try:
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlglot import exp, lineage, parse, parse_one
    from sqlglot.optimizer import qualify
except ImportError as exc:
    raise SystemExit(f"缺少依赖: {exc}") from exc


DIALECT = "spark"
WORKSPACE_ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = WORKSPACE_ROOT / "sql-data-map-output"
KEYWORD_SPLIT_RE = re.compile(
    r"(?im)^(?=(insert\s+overwrite\s+table|insert\s+into|create\s+(?:or\s+replace\s+)?temp\s+view|create\s+view|create\s+table))"
)
VAR_RE = re.compile(r"\$\{[^}]+\}")
RLIKE_TYPO_RE = re.compile(r"\bRLIike\b", re.IGNORECASE)
DDL_START_RE = re.compile(
    r"(?is)create\s+table\s+(?:if\s+not\s+exists\s+)?`?(?P<db>[^`. \n]+)`?\.`?(?P<table>[^`( \n]+)`?\s*\(",
)
DDL_COLUMN_RE = re.compile(
    r"^\s*`(?P<name>[^`]+)`\s+(?P<type>.*?)(?:\s+COMMENT\s+'(?P<comment>(?:''|[^'])*)')?\s*,?\s*$",
    re.IGNORECASE,
)
TABLE_COMMENT_RE = re.compile(r"(?i)\)\s*COMMENT\s+'((?:''|[^'])*)'")
FQN_RE = re.compile(r"^[a-zA-Z_][\w$]*\.[a-zA-Z_][\w$]*$")
TABLE_NAME_ALIASES = {
    "data_dwd.dwd_t_file_resource_id_test_001": "data_dwd.dwd_t_file_resource_id",
}


@dataclass
class ColumnMeta:
    name: str
    data_type: str | None = None
    comment: str | None = None
    comment_source: str = "unknown"


@dataclass
class TableMeta:
    table: str
    comment: str | None = None
    comment_source: str = "unknown"
    columns: dict[str, ColumnMeta] = field(default_factory=dict)
    table_type: str | None = None
    owner: str | None = None
    update_time: str | None = None


@dataclass
class StatementInfo:
    statement_id: str
    file_path: str
    statement_index: int
    statement_type: str
    target_dataset: str | None
    target_kind: str
    target_columns: list[str]
    query_sql: str | None
    source_tables: list[str]
    parse_status: str
    note: str = ""


@dataclass
class ColumnEdge:
    statement_id: str
    file_path: str
    target_dataset: str
    target_column: str
    target_column_position: int
    source_dataset: str | None
    source_column: str | None
    source_is_local_produced: bool
    expression_sql: str
    lineage_status: str
    target_kind: str
    source_hop_dataset: str | None = None
    source_hop_column: str | None = None


@dataclass
class RemoteAttempt:
    object_name: str
    object_type: str
    status: str
    detail: str = ""


def normalize_identifier(value: str) -> str:
    return value.strip().strip("`").lower()


def normalize_column_name(value: str) -> str:
    return value.strip().strip("`").lower()


def normalize_table_name(value: str) -> str:
    cleaned = value.strip().strip("`")
    parts = [part.strip("`") for part in cleaned.split(".") if part.strip("`")]
    normalized = ".".join(part.lower() for part in parts)
    return TABLE_NAME_ALIASES.get(normalized, normalized)


def is_physical_table_name(value: str | None) -> bool:
    return bool(value and "." in value and not value.startswith("virtual."))


def infer_table_role(table_name: str) -> str:
    short_name = table_name.split(".")[-1]
    if short_name.startswith("ods_"):
        return "ODS 原始层"
    if short_name.startswith("dwd_"):
        return "DWD 明细层"
    if short_name.startswith("dws_"):
        return "DWS 汇总层"
    if short_name.startswith("dim_") or ".tb" in table_name:
        return "DIM 维度层"
    if short_name.startswith("dm_"):
        return "DM 应用层"
    if short_name.startswith("ads_"):
        return "ADS 应用层"
    if short_name.startswith("tmp_"):
        return "TMP 临时层"
    return "未识别层级"


def infer_table_comment(table_name: str) -> str:
    return f"{infer_table_role(table_name)}表，名称推断自 `{table_name}`。"


def infer_column_comment(column_name: str) -> str:
    name = normalize_column_name(column_name)
    exact = {
        "file_id": "文件唯一标识。",
        "platform_source_id": "平台侧素材实例 ID。",
        "platform": "平台标识。",
        "brand": "品牌标识。",
        "spu": "商品 SPU。",
        "sku": "商品 SKU。",
        "sku_id": "商品 SKU ID。",
        "product_id": "商品 ID。",
        "product_name": "商品名称。",
        "title": "标题。",
        "author": "作者或达人。",
        "good_ids": "挂车商品 ID 集合。",
        "full_path": "完整路径。",
        "concatenated_path": "标准化拼接路径。",
        "file_name": "文件名。",
        "file_type": "文件类型。",
        "picture_type": "图片类型。",
        "picture_wear": "图片穿搭属性。",
        "picture_size": "图片文件大小。",
        "video_duration_type": "视频时长分类。",
        "video_wide_range": "视频画幅分类。",
        "scene": "场景标签。",
        "style": "风格标签。",
        "gender": "性别标签。",
        "big_cate": "品牌大类。",
        "mid_cate": "中类规整。",
        "sub_track": "细分赛道。",
        "publish_date": "发布日期。",
        "create_time": "创建时间。",
        "update_time": "更新时间。",
        "width": "宽度像素。",
        "height": "高度像素。",
        "resolution": "分辨率。",
        "room_id": "直播间或场次 ID。",
        "child_order_no": "子订单号。",
        "order_amount": "订单金额。",
        "pay_amount": "支付金额。",
        "item_num": "商品件数。",
        "shop_name": "店铺名称。",
        "erp_shop_id": "ERP 店铺 ID。",
    }
    if name in exact:
        return exact[name]
    if name.endswith("_id"):
        return f"{column_name} 标识字段。"
    if name.endswith("_time"):
        return f"{column_name} 时间字段。"
    if name.endswith("_date"):
        return f"{column_name} 日期字段。"
    if name.endswith("_cnt") or name.endswith("_count"):
        return f"{column_name} 计数字段。"
    if name.endswith("_amount") or name.endswith("_amt"):
        return f"{column_name} 金额字段。"
    if name.endswith("_rate"):
        return f"{column_name} 比率字段。"
    if "path" in name:
        return f"{column_name} 路径字段。"
    if "name" in name:
        return f"{column_name} 名称字段。"
    return f"{column_name} 字段，暂无显式注释，按字段名推断。"


def load_sql_files(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*.sql") if path.is_file())


def preprocess_sql(sql_text: str) -> str:
    sql_text = sql_text.replace("\ufeff", "")
    sql_text = VAR_RE.sub("'__VAR__'", sql_text)
    sql_text = RLIKE_TYPO_RE.sub("RLIKE", sql_text)
    sql_text = KEYWORD_SPLIT_RE.sub(";\n", sql_text).lstrip(";\n")
    return sql_text


def soft_normalize_sql(sql_text: str) -> str:
    sql_text = sql_text.replace("\ufeff", "")
    sql_text = VAR_RE.sub("'__VAR__'", sql_text)
    sql_text = RLIKE_TYPO_RE.sub("RLIKE", sql_text)
    return sql_text


def parse_local_ddl_metadata(sql_files: list[Path]) -> dict[str, TableMeta]:
    metadata: dict[str, TableMeta] = {}
    for file_path in sql_files:
        lines = file_path.read_text(encoding="utf-8").splitlines()
        current_table: str | None = None
        current_meta: TableMeta | None = None
        for line in lines:
            if current_table is None:
                match = DDL_START_RE.search(line)
                if match:
                    current_table = normalize_table_name(f"{match.group('db')}.{match.group('table')}")
                    current_meta = metadata.setdefault(current_table, TableMeta(table=current_table))
                    current_meta.comment_source = "local_ddl"
                continue

            column_match = DDL_COLUMN_RE.match(line)
            if column_match:
                name = column_match.group("name")
                data_type = (column_match.group("type") or "").strip()
                comment = column_match.group("comment")
                current_meta.columns[normalize_column_name(name)] = ColumnMeta(
                    name=name,
                    data_type=data_type or None,
                    comment=(comment or "").replace("''", "'") or None,
                    comment_source="local_ddl" if comment else "unknown",
                )
                continue

            table_comment_match = TABLE_COMMENT_RE.search(line)
            if table_comment_match and current_meta:
                current_meta.comment = table_comment_match.group(1).replace("''", "'")
                current_meta.comment_source = "local_ddl"
                current_table = None
                current_meta = None
                continue

            if line.strip().startswith(")") and current_meta:
                current_table = None
                current_meta = None
    return metadata


def parse_statements(sql_text: str) -> list[Any]:
    candidates: list[str] = []
    for candidate in [sql_text, soft_normalize_sql(sql_text), preprocess_sql(sql_text)]:
        if candidate not in candidates:
            candidates.append(candidate)

    for candidate in candidates:
        for dialect in [DIALECT, "mysql", "hive"]:
            try:
                statements = parse(candidate, read=dialect)
                if statements:
                    return statements
            except Exception:
                continue

    chunks = [chunk.strip() for chunk in preprocess_sql(sql_text).split(";") if chunk.strip()]
    statements: list[Any] = []
    for chunk in chunks:
        for dialect in [DIALECT, "mysql", "hive"]:
            try:
                statements.extend(parse(chunk, read=dialect))
                break
            except Exception:
                continue
    return statements


def build_schema_map(metadata: dict[str, TableMeta], produced_schema: dict[str, list[str]]) -> dict[str, Any]:
    schema: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    for table_name, table_meta in metadata.items():
        if "." not in table_name:
            continue
        if not table_meta.columns:
            continue
        db_name, short_name = table_name.split(".", 1)
        schema.setdefault(db_name, {})[short_name] = {
            col.name: (col.data_type or "string") for col in table_meta.columns.values()
        }
    for table_name, columns in produced_schema.items():
        if "." not in table_name:
            continue
        db_name, short_name = table_name.split(".", 1)
        table_schema = schema.setdefault(db_name, {}).setdefault(short_name, {})
        for column in columns:
            table_schema.setdefault(column, "string")
    return schema


def table_name_from_expr(table_expr: exp.Expression | None) -> str | None:
    if table_expr is None:
        return None
    if isinstance(table_expr, exp.Schema):
        table_expr = table_expr.this
    if not isinstance(table_expr, exp.Table):
        return None
    parts = []
    if table_expr.catalog:
        parts.append(table_expr.catalog.name if hasattr(table_expr.catalog, "name") else str(table_expr.catalog))
    if table_expr.db:
        parts.append(table_expr.db.name if hasattr(table_expr.db, "name") else str(table_expr.db))
    if table_expr.name:
        parts.append(table_expr.name)
    if not parts:
        return None
    return normalize_table_name(".".join(parts))


def cte_names_from_query(query: exp.Expression | None) -> set[str]:
    if query is None:
        return set()
    names: set[str] = set()
    with_expr = query.args.get("with") or query.args.get("with_")
    if with_expr:
        for cte in with_expr.expressions or []:
            alias = cte.alias
            if alias:
                names.add(normalize_identifier(alias))
    return names


def attach_with_clause(owner: exp.Expression, query: exp.Expression | None) -> exp.Expression | None:
    if query is None:
        return None
    with_expr = owner.args.get("with") or owner.args.get("with_")
    if not with_expr:
        return query
    query = query.copy()
    query.set("with_", with_expr.copy())
    return query


def extract_statement_target(stmt: exp.Expression, file_path: str, statement_index: int) -> tuple[str | None, str, exp.Expression | None]:
    if isinstance(stmt, exp.Insert):
        return table_name_from_expr(stmt.this), "physical_table", attach_with_clause(stmt, stmt.expression)
    if isinstance(stmt, exp.Create):
        kind = "temp_view" if any(isinstance(prop, exp.TemporaryProperty) for prop in (stmt.args.get("properties") or []).expressions or []) else "view"
        if str(stmt.args.get("kind") or "").upper() == "TABLE":
            kind = "physical_table"
        if stmt.expression is None:
            return table_name_from_expr(stmt.this), kind, None
        return table_name_from_expr(stmt.this) or normalize_table_name(stmt.this.sql()), kind, attach_with_clause(stmt, stmt.expression)
    if isinstance(stmt, exp.Select) or isinstance(stmt, exp.Union):
        target = f"virtual.{normalize_table_name(Path(file_path).stem)}.stmt_{statement_index}"
        return target, "virtual_query", stmt
    return None, "unsupported", None


def qualify_query(query: exp.Expression, schema: dict[str, Any]) -> exp.Expression:
    try:
        return qualify.qualify(
            query.copy(),
            dialect=DIALECT,
            schema=schema,
            validate_qualify_columns=False,
            identify=False,
        )
    except Exception:
        return query.copy()


def get_query_output_columns(query: exp.Expression, target_columns_hint: list[str] | None = None) -> list[str]:
    named = [name for name in query.named_selects if name]
    if target_columns_hint and len(target_columns_hint) == len(query.selects):
        return target_columns_hint
    if named:
        return named
    columns: list[str] = []
    for idx, select_expr in enumerate(query.selects, start=1):
        alias = select_expr.alias_or_name
        columns.append(alias or f"expr_{idx}")
    return columns


def guess_source_tables(query: exp.Expression) -> list[str]:
    cte_names = cte_names_from_query(query)
    names: set[str] = set()
    for table_expr in query.find_all(exp.Table):
        table_name = table_name_from_expr(table_expr)
        if not table_name:
            alias_name = normalize_identifier(table_expr.name)
            if alias_name and alias_name not in cte_names:
                names.add(alias_name)
            continue
        short_name = normalize_identifier(table_expr.name)
        if short_name in cte_names:
            continue
        names.add(table_name)
    return sorted(names)


def alias_query_outputs(query: exp.Expression, target_columns: list[str]) -> exp.Expression:
    query = query.copy()
    if isinstance(query, exp.Select) and len(query.selects) == len(target_columns):
        aliased: list[exp.Expression] = []
        for select_expr, target_column in zip(query.selects, target_columns):
            aliased.append(select_expr.as_(target_column, copy=True))
        query.set("expressions", aliased)
    return query


def extract_leaf_sources(node: lineage.Node) -> list[tuple[str, str]]:
    leaf_sources: set[tuple[str, str]] = set()
    for item in node.walk():
        if isinstance(item.expression, exp.Table):
            table_name = table_name_from_expr(item.expression)
            if not table_name:
                continue
            column_name = item.name.rsplit(".", 1)[-1]
            leaf_sources.add((table_name, normalize_column_name(column_name)))
    return sorted(leaf_sources)


def ensure_table_meta(table_metadata: dict[str, TableMeta], table_name: str) -> TableMeta:
    table_name = normalize_table_name(table_name)
    return table_metadata.setdefault(table_name, TableMeta(table=table_name))


def merge_table_metadata(target: TableMeta, incoming: TableMeta) -> None:
    if incoming.comment and (not target.comment or target.comment_source == "unknown"):
        target.comment = incoming.comment
        target.comment_source = incoming.comment_source
    if incoming.table_type and not target.table_type:
        target.table_type = incoming.table_type
    if incoming.owner and not target.owner:
        target.owner = incoming.owner
    if incoming.update_time and not target.update_time:
        target.update_time = incoming.update_time
    for key, value in incoming.columns.items():
        if key not in target.columns:
            target.columns[key] = value
            continue
        target_col = target.columns[key]
        if value.data_type and not target_col.data_type:
            target_col.data_type = value.data_type
        if value.comment and (not target_col.comment or target_col.comment_source == "unknown"):
            target_col.comment = value.comment
            target_col.comment_source = value.comment_source


def discover_credentials(root: Path) -> tuple[str | None, str | None]:
    secret_id = os.environ.get("DLC_USER") or os.environ.get("TENCENTCLOUD_SECRET_ID")
    secret_key = os.environ.get("DLC_PASSWORD") or os.environ.get("TENCENTCLOUD_SECRET_KEY")
    if secret_id and secret_key:
        return secret_id, secret_key
    for candidate in sorted(root.rglob("skill.json")):
        try:
            data = json.loads(candidate.read_text(encoding="utf-8"))
        except Exception:
            continue
        skills = data.get("skills") or {}
        for skill in skills.values():
            config = skill.get("config") or {}
            candidate_id = config.get("DLC_USER") or config.get("TENCENTCLOUD_SECRET_ID")
            candidate_key = config.get("DLC_PASSWORD") or config.get("TENCENTCLOUD_SECRET_KEY")
            if candidate_id and candidate_key:
                return str(candidate_id), str(candidate_key)
    return None, None


class DlcMetadataClient:
    def __init__(self, secret_id: str, secret_key: str, region: str = "ap-shanghai"):
        from tencentcloud.common import credential
        from tencentcloud.common.profile.client_profile import ClientProfile
        from tencentcloud.common.profile.http_profile import HttpProfile
        from tencentcloud.dlc.v20210125 import dlc_client

        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, region, client_profile)

    def describe_table(self, table_name: str) -> TableMeta:
        from tencentcloud.dlc.v20210125 import models

        db_name, short_name = table_name.split(".", 1)
        req = models.DescribeTableRequest()
        req.DatabaseName = db_name
        req.TableName = short_name
        response = self.client.DescribeTable(req)
        payload = json.loads(response.to_json_string())
        table_info = payload.get("Table", payload)
        table_meta = TableMeta(
            table=table_name,
            comment=table_info.get("Comment") or table_info.get("Description") or None,
            comment_source="dlc_describe",
            table_type=table_info.get("TableType") or None,
            owner=table_info.get("Owner") or None,
            update_time=table_info.get("UpdateTime") or None,
        )
        for raw_col in table_info.get("Columns") or []:
            name = raw_col.get("Name")
            if not name:
                continue
            table_meta.columns[normalize_column_name(name)] = ColumnMeta(
                name=name,
                data_type=raw_col.get("Type") or None,
                comment=raw_col.get("Comment") or raw_col.get("Description") or None,
                comment_source="dlc_describe",
            )
        return table_meta


def collect_remote_metadata(
    table_names: list[str],
    table_metadata: dict[str, TableMeta],
    root: Path,
    enable_remote: bool,
) -> list[RemoteAttempt]:
    attempts: list[RemoteAttempt] = []
    if not enable_remote:
        attempts.append(RemoteAttempt("all", "dlc_metadata", "skipped", "remote disabled"))
        return attempts

    secret_id, secret_key = discover_credentials(root)
    if not secret_id or not secret_key:
        attempts.append(RemoteAttempt("all", "dlc_metadata", "skipped", "credentials not found"))
        return attempts

    try:
        client = DlcMetadataClient(secret_id, secret_key)
    except Exception as exc:
        attempts.append(RemoteAttempt("all", "dlc_metadata", "failed", f"client init failed: {exc}"))
        return attempts

    for table_name in sorted({name for name in table_names if is_physical_table_name(name)}):
        try:
            remote_meta = client.describe_table(table_name)
            merge_table_metadata(ensure_table_meta(table_metadata, table_name), remote_meta)
            attempts.append(RemoteAttempt(table_name, "table", "success"))
        except Exception as exc:
            attempts.append(RemoteAttempt(table_name, "table", "failed", str(exc)))
    return attempts


def first_pass_statement_scan(
    sql_files: list[Path],
    schema: dict[str, Any],
) -> tuple[list[StatementInfo], dict[str, list[str]], set[str], list[dict[str, str]]]:
    statements: list[StatementInfo] = []
    produced_schema: dict[str, list[str]] = {}
    referenced_tables: set[str] = set()
    parse_issues: list[dict[str, str]] = []

    for file_path in sql_files:
        raw_sql = file_path.read_text(encoding="utf-8")
        parsed = parse_statements(raw_sql)
        if not parsed:
            parse_issues.append({"file_path": str(file_path), "issue": "parse_failed", "detail": "no statements parsed"})
            continue

        for idx, stmt in enumerate(parsed, start=1):
            target_dataset, target_kind, query = extract_statement_target(stmt, str(file_path), idx)
            statement_id = f"{file_path.relative_to(WORKSPACE_ROOT)}::stmt_{idx}"
            source_tables: list[str] = []
            target_columns: list[str] = []
            parse_status = "ok"
            note = ""

            if query is not None:
                qualified = qualify_query(query, schema)
                source_tables = guess_source_tables(qualified)
                target_columns = get_query_output_columns(qualified)
                for source_table in source_tables:
                    referenced_tables.add(source_table)
                if target_dataset and target_kind in {"physical_table", "view", "temp_view", "virtual_query"}:
                    produced_schema[target_dataset] = target_columns
            else:
                if target_dataset:
                    produced_schema.setdefault(target_dataset, [])

            if target_dataset:
                referenced_tables.add(target_dataset)

            statements.append(
                StatementInfo(
                    statement_id=statement_id,
                    file_path=str(file_path),
                    statement_index=idx,
                    statement_type=type(stmt).__name__.lower(),
                    target_dataset=target_dataset,
                    target_kind=target_kind,
                    target_columns=target_columns,
                    query_sql=query.sql(pretty=False, dialect=DIALECT) if query is not None else None,
                    source_tables=source_tables,
                    parse_status=parse_status,
                    note=note,
                )
            )

    return statements, produced_schema, referenced_tables, parse_issues


def second_pass_lineage_scan(
    statements: list[StatementInfo],
    schema: dict[str, Any],
    produced_datasets: set[str],
) -> tuple[list[ColumnEdge], list[dict[str, str]]]:
    edges: list[ColumnEdge] = []
    issues: list[dict[str, str]] = []

    for stmt_info in statements:
        if not stmt_info.query_sql or not stmt_info.target_dataset:
            continue

        try:
            query = parse_one(stmt_info.query_sql, read=DIALECT)
        except Exception as exc:
            issues.append({"statement_id": stmt_info.statement_id, "issue": "query_reparse_failed", "detail": str(exc)})
            continue

        target_columns = stmt_info.target_columns[:]
        if not target_columns:
            qualified = qualify_query(query, schema)
            target_columns = get_query_output_columns(qualified)

        aliased_query = alias_query_outputs(query, target_columns)
        for position, target_column in enumerate(target_columns, start=1):
            try:
                node = lineage.lineage(
                    target_column,
                    aliased_query,
                    schema=schema,
                    dialect=DIALECT,
                    validate_qualify_columns=False,
                    identify=False,
                )
                sources = extract_leaf_sources(node)
                expression_sql = node.expression.sql(dialect=DIALECT)
                if not sources:
                    edges.append(
                        ColumnEdge(
                            statement_id=stmt_info.statement_id,
                            file_path=stmt_info.file_path,
                            target_dataset=stmt_info.target_dataset,
                            target_column=target_column,
                            target_column_position=position,
                            source_dataset=None,
                            source_column=None,
                            source_is_local_produced=False,
                            expression_sql=expression_sql,
                            lineage_status="constant_or_unresolved",
                            target_kind=stmt_info.target_kind,
                        )
                    )
                    continue

                for source_dataset, source_column in sources:
                    edges.append(
                        ColumnEdge(
                            statement_id=stmt_info.statement_id,
                            file_path=stmt_info.file_path,
                            target_dataset=stmt_info.target_dataset,
                            target_column=target_column,
                            target_column_position=position,
                            source_dataset=source_dataset,
                            source_column=source_column,
                            source_is_local_produced=source_dataset in produced_datasets,
                            expression_sql=expression_sql,
                            lineage_status="ok",
                            target_kind=stmt_info.target_kind,
                            source_hop_dataset=source_dataset,
                            source_hop_column=source_column,
                        )
                    )
            except Exception as exc:
                issues.append(
                    {
                        "statement_id": stmt_info.statement_id,
                        "issue": "column_lineage_failed",
                        "detail": f"{stmt_info.target_dataset}.{target_column}: {exc}",
                    }
                )
                edges.append(
                    ColumnEdge(
                        statement_id=stmt_info.statement_id,
                        file_path=stmt_info.file_path,
                        target_dataset=stmt_info.target_dataset,
                        target_column=target_column,
                        target_column_position=position,
                        source_dataset=None,
                        source_column=None,
                        source_is_local_produced=False,
                        expression_sql=target_column,
                        lineage_status="failed",
                        target_kind=stmt_info.target_kind,
                    )
                )
    return edges, issues


def build_base_table_sets(statements: list[StatementInfo]) -> tuple[set[str], set[str], set[str]]:
    produced_tables = {
        normalize_table_name(stmt.target_dataset)
        for stmt in statements
        if stmt.target_dataset and stmt.target_kind in {"physical_table", "view", "temp_view"}
    }
    source_tables = {
        normalize_table_name(table_name)
        for stmt in statements
        for table_name in stmt.source_tables
        if is_physical_table_name(table_name) or table_name in produced_tables
    }
    base_tables = {table_name for table_name in source_tables if table_name not in produced_tables and is_physical_table_name(table_name)}
    return base_tables, produced_tables, source_tables


def build_table_edges(statements: list[StatementInfo], produced_tables: set[str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for stmt in statements:
        if not stmt.target_dataset:
            continue
        for source_table in stmt.source_tables:
            normalized_source = normalize_table_name(source_table)
            if not is_physical_table_name(normalized_source) and normalized_source not in produced_tables:
                continue
            rows.append(
                {
                    "statement_id": stmt.statement_id,
                    "file_path": stmt.file_path,
                    "source_table": normalized_source,
                    "target_table": stmt.target_dataset,
                    "target_kind": stmt.target_kind,
                    "statement_type": stmt.statement_type,
                }
            )
    return pd.DataFrame(rows).drop_duplicates()


def build_transitive_table_paths(
    table_edges: pd.DataFrame,
    base_tables: set[str],
) -> pd.DataFrame:
    adjacency: dict[str, set[str]] = defaultdict(set)
    for row in table_edges.to_dict("records"):
        adjacency[row["source_table"]].add(row["target_table"])

    rows: list[dict[str, Any]] = []
    for base_table in sorted(base_tables):
        queue: deque[tuple[str, list[str]]] = deque([(base_table, [base_table])])
        visited: set[str] = {base_table}
        while queue:
            current, path = queue.popleft()
            next_nodes = sorted(adjacency.get(current, set()))
            for next_node in next_nodes:
                new_path = path + [next_node]
                rows.append(
                    {
                        "base_table": base_table,
                        "downstream_table": next_node,
                        "hop_count": len(new_path) - 1,
                        "path": " -> ".join(new_path),
                    }
                )
                if next_node not in visited:
                    visited.add(next_node)
                    queue.append((next_node, new_path))
    return pd.DataFrame(rows)


def build_immediate_column_df(edges: list[ColumnEdge]) -> pd.DataFrame:
    rows = []
    for edge in edges:
        rows.append(
            {
                "statement_id": edge.statement_id,
                "file_path": edge.file_path,
                "target_table": edge.target_dataset,
                "target_column": edge.target_column,
                "target_column_position": edge.target_column_position,
                "source_table": edge.source_dataset,
                "source_column": edge.source_column,
                "source_is_local_produced": edge.source_is_local_produced,
                "expression_sql": edge.expression_sql,
                "lineage_status": edge.lineage_status,
                "target_kind": edge.target_kind,
            }
        )
    return pd.DataFrame(rows)


def build_recursive_column_df(
    edges: list[ColumnEdge],
    statements: list[StatementInfo],
    base_tables: set[str],
) -> pd.DataFrame:
    produced_by_dataset: dict[str, list[str]] = defaultdict(list)
    for stmt in statements:
        if stmt.target_dataset and stmt.query_sql:
            produced_by_dataset[stmt.target_dataset].append(stmt.statement_id)

    immediate_by_stmt_col: dict[tuple[str, str], list[ColumnEdge]] = defaultdict(list)
    target_kind_by_stmt: dict[str, str] = {}
    for edge in edges:
        immediate_by_stmt_col[(edge.statement_id, normalize_column_name(edge.target_column))].append(edge)
        target_kind_by_stmt[edge.statement_id] = edge.target_kind

    statement_by_id = {stmt.statement_id: stmt for stmt in statements}

    def resolve(dataset: str, column: str, visited: set[tuple[str, str]]) -> list[dict[str, Any]]:
        key = (dataset, normalize_column_name(column))
        if key in visited:
            return []
        if dataset in base_tables:
            return [
                {
                    "base_table": dataset,
                    "base_column": normalize_column_name(column),
                    "path": f"{dataset}.{normalize_column_name(column)}",
                    "producer_chain": "",
                    "resolution_status": "base",
                }
            ]

        statement_ids = produced_by_dataset.get(dataset, [])
        if not statement_ids:
            return []

        resolved_rows: list[dict[str, Any]] = []
        next_visited = set(visited)
        next_visited.add(key)
        for statement_id in statement_ids:
            immediate_edges = immediate_by_stmt_col.get((statement_id, normalize_column_name(column)), [])
            if not immediate_edges:
                resolved_rows.append(
                    {
                        "base_table": None,
                        "base_column": None,
                        "path": f"{dataset}.{normalize_column_name(column)}",
                        "producer_chain": statement_id,
                        "resolution_status": "unresolved_local_column",
                    }
                )
                continue
            for immediate_edge in immediate_edges:
                if not immediate_edge.source_dataset or not immediate_edge.source_column:
                    resolved_rows.append(
                        {
                            "base_table": None,
                            "base_column": None,
                            "path": f"{dataset}.{normalize_column_name(column)}",
                            "producer_chain": statement_id,
                            "resolution_status": immediate_edge.lineage_status,
                        }
                    )
                    continue
                upstream_rows = resolve(immediate_edge.source_dataset, immediate_edge.source_column, next_visited)
                for upstream in upstream_rows:
                    upstream["path"] = f"{upstream['path']} -> {dataset}.{normalize_column_name(column)}"
                    upstream["producer_chain"] = " -> ".join(part for part in [upstream.get("producer_chain", ""), statement_id] if part)
                    resolved_rows.append(upstream)
        return resolved_rows

    rows: list[dict[str, Any]] = []
    for edge in edges:
        if edge.lineage_status not in {"ok"} or not edge.source_dataset or not edge.source_column:
            continue
        resolved = resolve(edge.source_dataset, edge.source_column, set())
        for item in resolved:
            if not item["base_table"] or item["base_table"] not in base_tables:
                continue
            rows.append(
                {
                    "statement_id": edge.statement_id,
                    "file_path": edge.file_path,
                    "target_table": edge.target_dataset,
                    "target_column": edge.target_column,
                    "target_kind": edge.target_kind,
                    "base_table": item["base_table"],
                    "base_column": item["base_column"],
                    "path": item["path"],
                    "producer_chain": item["producer_chain"],
                    "expression_sql": edge.expression_sql,
                    "resolution_status": item["resolution_status"],
                }
            )
    return pd.DataFrame(rows).drop_duplicates()


def enrich_metadata_with_inference(
    table_metadata: dict[str, TableMeta],
    all_tables: set[str],
    produced_schema: dict[str, list[str]],
) -> None:
    for table_name in sorted(all_tables):
        meta = ensure_table_meta(table_metadata, table_name)
        if not meta.comment:
            meta.comment = infer_table_comment(table_name)
            meta.comment_source = "inferred"
        for column_name in produced_schema.get(table_name, []):
            key = normalize_column_name(column_name)
            if key not in meta.columns:
                meta.columns[key] = ColumnMeta(
                    name=column_name,
                    data_type=None,
                    comment=infer_column_comment(column_name),
                    comment_source="inferred",
                )
            elif not meta.columns[key].comment:
                meta.columns[key].comment = infer_column_comment(column_name)
                meta.columns[key].comment_source = "inferred"
    for meta in table_metadata.values():
        for column_key, column_meta in meta.columns.items():
            if not column_meta.comment:
                column_meta.comment = infer_column_comment(column_meta.name)
                column_meta.comment_source = "inferred"


def build_base_table_summary(
    base_tables: set[str],
    table_paths: pd.DataFrame,
    table_edges: pd.DataFrame,
    table_metadata: dict[str, TableMeta],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for table_name in sorted(base_tables):
        meta = table_metadata.get(table_name) or TableMeta(table=table_name)
        downstream_df = table_paths[table_paths["base_table"] == table_name] if not table_paths.empty else pd.DataFrame()
        direct_df = table_edges[table_edges["source_table"] == table_name] if not table_edges.empty else pd.DataFrame()
        rows.append(
            {
                "base_table": table_name,
                "table_role": infer_table_role(table_name),
                "table_comment": meta.comment,
                "table_comment_source": meta.comment_source,
                "column_count": len(meta.columns),
                "direct_downstream_count": int(direct_df["target_table"].nunique()) if not direct_df.empty else 0,
                "all_downstream_count": int(downstream_df["downstream_table"].nunique()) if not downstream_df.empty else 0,
                "direct_downstream_tables": ", ".join(sorted(direct_df["target_table"].dropna().unique())) if not direct_df.empty else "",
            }
        )
    return pd.DataFrame(rows)


def build_dictionary_df(
    table_metadata: dict[str, TableMeta],
    base_tables: set[str],
    produced_tables: set[str],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    all_tables = sorted(table_metadata)
    for table_name in all_tables:
        meta = table_metadata[table_name]
        if table_name in base_tables:
            lineage_role = "base_table"
        elif table_name in produced_tables:
            lineage_role = "produced_local"
        elif table_name.startswith("virtual."):
            lineage_role = "virtual"
        else:
            lineage_role = "external_reference"
        if not meta.columns:
            rows.append(
                {
                    "table_name": table_name,
                    "lineage_role": lineage_role,
                    "table_comment": meta.comment,
                    "table_comment_source": meta.comment_source,
                    "column_name": "",
                    "data_type": "",
                    "column_comment": "",
                    "column_comment_source": "",
                }
            )
            continue
        for column_meta in sorted(meta.columns.values(), key=lambda item: item.name):
            rows.append(
                {
                    "table_name": table_name,
                    "lineage_role": lineage_role,
                    "table_comment": meta.comment,
                    "table_comment_source": meta.comment_source,
                    "column_name": column_meta.name,
                    "data_type": column_meta.data_type or "",
                    "column_comment": column_meta.comment or "",
                    "column_comment_source": column_meta.comment_source,
                }
            )
    return pd.DataFrame(rows)


def build_assessment_df(
    sql_files: list[Path],
    statements: list[StatementInfo],
    base_tables: set[str],
    produced_tables: set[str],
    table_edges: pd.DataFrame,
    immediate_df: pd.DataFrame,
    recursive_df: pd.DataFrame,
    parse_issues: list[dict[str, str]],
    lineage_issues: list[dict[str, str]],
    remote_attempts: list[RemoteAttempt],
) -> pd.DataFrame:
    failed_remote = [item for item in remote_attempts if item.status == "failed"]
    rows = [
        {"category": "coverage", "item": "sql_file_count", "value": len(sql_files), "detail": ""},
        {"category": "coverage", "item": "statement_count", "value": len(statements), "detail": ""},
        {"category": "coverage", "item": "base_table_count", "value": len(base_tables), "detail": ""},
        {"category": "coverage", "item": "produced_table_count", "value": len(produced_tables), "detail": ""},
        {"category": "coverage", "item": "table_edge_count", "value": int(len(table_edges)), "detail": ""},
        {"category": "coverage", "item": "immediate_column_lineage_rows", "value": int(len(immediate_df)), "detail": ""},
        {"category": "coverage", "item": "base_column_lineage_rows", "value": int(len(recursive_df)), "detail": ""},
        {"category": "quality", "item": "parse_issue_count", "value": len(parse_issues), "detail": "; ".join(item["file_path"] for item in parse_issues[:10])},
        {"category": "quality", "item": "lineage_issue_count", "value": len(lineage_issues), "detail": "; ".join(item["statement_id"] for item in lineage_issues[:10])},
        {"category": "remote", "item": "remote_attempt_count", "value": len(remote_attempts), "detail": ""},
        {"category": "remote", "item": "remote_failure_count", "value": len(failed_remote), "detail": "; ".join(item.object_name for item in failed_remote[:10])},
    ]
    return pd.DataFrame(rows)


def build_issue_df(
    parse_issues: list[dict[str, str]],
    lineage_issues: list[dict[str, str]],
    remote_attempts: list[RemoteAttempt],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for item in parse_issues:
        rows.append(
            {
                "issue_type": "parse_issue",
                "statement_id": item.get("statement_id", ""),
                "file_path": item.get("file_path", ""),
                "object_name": "",
                "detail": item.get("detail", ""),
            }
        )
    for item in lineage_issues:
        rows.append(
            {
                "issue_type": "lineage_issue",
                "statement_id": item.get("statement_id", ""),
                "file_path": "",
                "object_name": "",
                "detail": item.get("detail", ""),
            }
        )
    for item in remote_attempts:
        if item.status == "failed":
            rows.append(
                {
                    "issue_type": "remote_failure",
                    "statement_id": "",
                    "file_path": "",
                    "object_name": item.object_name,
                    "detail": item.detail,
                }
            )
    return pd.DataFrame(rows)


def build_table_index_df(
    dictionary_df: pd.DataFrame,
    table_edges_df: pd.DataFrame,
    table_paths_df: pd.DataFrame,
    base_summary_df: pd.DataFrame,
) -> pd.DataFrame:
    if dictionary_df.empty:
        return pd.DataFrame(
            columns=[
                "table_name",
                "table_role",
                "lineage_role",
                "table_comment",
                "column_count",
                "direct_upstream_count",
                "direct_downstream_count",
                "all_downstream_count",
                "direct_upstream_tables",
                "direct_downstream_tables",
            ]
        )

    base_summary_map = {}
    if not base_summary_df.empty:
        base_summary_map = {
            row["base_table"]: row for row in base_summary_df.to_dict("records")
        }

    rows: list[dict[str, Any]] = []
    for table_name, group in dictionary_df.groupby("table_name", dropna=False):
        valid_columns = group[group["column_name"].fillna("") != ""]
        upstream_df = table_edges_df[table_edges_df["target_table"] == table_name] if not table_edges_df.empty else pd.DataFrame()
        downstream_df = table_edges_df[table_edges_df["source_table"] == table_name] if not table_edges_df.empty else pd.DataFrame()
        path_df = table_paths_df[table_paths_df["base_table"] == table_name] if not table_paths_df.empty else pd.DataFrame()
        base_row = base_summary_map.get(table_name, {})
        first = group.iloc[0]
        rows.append(
            {
                "table_name": table_name,
                "table_role": infer_table_role(table_name),
                "lineage_role": first.get("lineage_role", ""),
                "table_comment": first.get("table_comment", ""),
                "column_count": int(valid_columns["column_name"].nunique()) if not valid_columns.empty else 0,
                "direct_upstream_count": int(upstream_df["source_table"].nunique()) if not upstream_df.empty else 0,
                "direct_downstream_count": int(downstream_df["target_table"].nunique()) if not downstream_df.empty else 0,
                "all_downstream_count": int(base_row.get("all_downstream_count", 0) or 0),
                "direct_upstream_tables": ", ".join(sorted(upstream_df["source_table"].dropna().astype(str).unique())) if not upstream_df.empty else "",
                "direct_downstream_tables": ", ".join(sorted(downstream_df["target_table"].dropna().astype(str).unique())) if not downstream_df.empty else "",
                "all_downstream_tables": ", ".join(sorted(path_df["downstream_table"].dropna().astype(str).unique())) if not path_df.empty else "",
                "search_hint": f"可按 table_name 精确筛选: {table_name}",
            }
        )
    return pd.DataFrame(rows).sort_values(["table_name"], kind="stable").reset_index(drop=True)


def build_field_index_df(
    dictionary_df: pd.DataFrame,
    immediate_df: pd.DataFrame,
    recursive_df: pd.DataFrame,
) -> pd.DataFrame:
    if dictionary_df.empty:
        return pd.DataFrame(
            columns=[
                "field_name",
                "table_name",
                "table_role",
                "lineage_role",
                "data_type",
                "column_comment",
                "direct_upstream_count",
                "direct_upstream_tables",
                "direct_downstream_count",
                "direct_downstream_tables",
                "base_source_count",
                "base_source_tables",
            ]
        )

    rows: list[dict[str, Any]] = []
    detail_df = dictionary_df[dictionary_df["column_name"].fillna("") != ""].copy()
    for record in detail_df.to_dict("records"):
        table_name = str(record["table_name"])
        column_name = str(record["column_name"])
        upstream_df = immediate_df[
            (immediate_df["target_table"] == table_name)
            & (immediate_df["target_column"] == column_name)
            & (immediate_df["source_table"].notna())
        ] if not immediate_df.empty else pd.DataFrame()
        downstream_df = immediate_df[
            (immediate_df["source_table"] == table_name)
            & (immediate_df["source_column"] == column_name)
            & (immediate_df["target_table"].notna())
        ] if not immediate_df.empty else pd.DataFrame()
        base_df = recursive_df[
            (recursive_df["target_table"] == table_name)
            & (recursive_df["target_column"] == column_name)
            & (recursive_df["base_table"].notna())
        ] if not recursive_df.empty else pd.DataFrame()
        rows.append(
            {
                "field_name": column_name,
                "table_name": table_name,
                "table_role": infer_table_role(table_name),
                "lineage_role": record.get("lineage_role", ""),
                "data_type": record.get("data_type", ""),
                "column_comment": record.get("column_comment", ""),
                "column_comment_source": record.get("column_comment_source", ""),
                "direct_upstream_count": int(upstream_df["source_table"].nunique()) if not upstream_df.empty else 0,
                "direct_upstream_tables": ", ".join(sorted(upstream_df["source_table"].dropna().astype(str).unique())) if not upstream_df.empty else "",
                "direct_downstream_count": int(downstream_df["target_table"].nunique()) if not downstream_df.empty else 0,
                "direct_downstream_tables": ", ".join(sorted(downstream_df["target_table"].dropna().astype(str).unique())) if not downstream_df.empty else "",
                "base_source_count": int(base_df["base_table"].nunique()) if not base_df.empty else 0,
                "base_source_tables": ", ".join(sorted(base_df["base_table"].dropna().astype(str).unique())) if not base_df.empty else "",
                "search_hint": f"可按 field_name 精确筛选: {column_name}",
            }
        )
    return pd.DataFrame(rows).sort_values(["field_name", "table_name"], kind="stable").reset_index(drop=True)


def build_statement_df(statements: list[StatementInfo]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for stmt in statements:
        rows.append(
            {
                "statement_id": stmt.statement_id,
                "file_path": stmt.file_path,
                "statement_index": stmt.statement_index,
                "statement_type": stmt.statement_type,
                "target_dataset": stmt.target_dataset,
                "target_kind": stmt.target_kind,
                "target_columns": ", ".join(stmt.target_columns),
                "source_tables": ", ".join(stmt.source_tables),
                "parse_status": stmt.parse_status,
                "note": stmt.note,
                "query_sql": stmt.query_sql or "",
            }
        )
    return pd.DataFrame(rows)


def build_remote_attempt_df(remote_attempts: list[RemoteAttempt]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "object_name": item.object_name,
                "object_type": item.object_type,
                "status": item.status,
                "detail": item.detail,
            }
            for item in remote_attempts
        ]
    )


def build_readme_df(output_file: Path, root: Path, enable_remote: bool) -> pd.DataFrame:
    rows = [
        {"section": "生成时间", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"section": "扫描目录", "value": str(root)},
        {"section": "产物文件", "value": str(output_file)},
        {"section": "SQL 方言", "value": DIALECT},
        {"section": "远端补采", "value": "启用" if enable_remote else "关闭"},
        {"section": "说明", "value": "字段级血缘以 SQL 静态解析为主；表/字段备注优先取本地 DDL，再补 DLC DescribeTable，最后用规则推断。"},
        {"section": "说明", "value": "若同一目标表存在多个本地生产版本，Excel 会保留多条变体血缘，不会擅自覆盖。"},
        {"section": "说明", "value": "基表定义为：在当前扫描范围内被引用但未被本地 SQL 产出的物理表。"},
    ]
    return pd.DataFrame(rows)


def autosize_workbook(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in worksheet.columns:
            max_len = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 80)


def write_outputs(
    output_dir: Path,
    readme_df: pd.DataFrame,
    base_summary_df: pd.DataFrame,
    table_edges_df: pd.DataFrame,
    table_paths_df: pd.DataFrame,
    immediate_df: pd.DataFrame,
    recursive_df: pd.DataFrame,
    table_index_df: pd.DataFrame,
    field_index_df: pd.DataFrame,
    dictionary_df: pd.DataFrame,
    statements_df: pd.DataFrame,
    assessment_df: pd.DataFrame,
    issue_df: pd.DataFrame,
    remote_attempt_df: pd.DataFrame,
    raw_payload: dict[str, Any],
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "sql_data_map.xlsx"
    json_path = output_dir / "sql_data_map_raw.json"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False)
        base_summary_df.to_excel(writer, sheet_name="基表总览", index=False)
        table_edges_df.to_excel(writer, sheet_name="表级血缘_直接", index=False)
        table_paths_df.to_excel(writer, sheet_name="表级血缘_路径", index=False)
        immediate_df.to_excel(writer, sheet_name="字段级血缘_直接", index=False)
        recursive_df.to_excel(writer, sheet_name="字段级血缘_基表", index=False)
        table_index_df.to_excel(writer, sheet_name="索引_按表名", index=False)
        field_index_df.to_excel(writer, sheet_name="索引_按字段名", index=False)
        dictionary_df.to_excel(writer, sheet_name="表字段字典", index=False)
        statements_df.to_excel(writer, sheet_name="加工节点", index=False)
        assessment_df.to_excel(writer, sheet_name="评估", index=False)
        issue_df.to_excel(writer, sheet_name="问题明细", index=False)
        remote_attempt_df.to_excel(writer, sheet_name="远端补采", index=False)
        autosize_workbook(writer)

    json_path.write_text(json.dumps(raw_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return xlsx_path, json_path


def main(argv: list[str]) -> int:
    import argparse

    parser = argparse.ArgumentParser(description="扫描 SQL 目录并生成表级/字段级数据地图 Excel。")
    parser.add_argument("--root", default=str(WORKSPACE_ROOT), help="待扫描根目录，默认当前工作区。")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录。")
    parser.add_argument("--disable-remote", action="store_true", help="关闭 DLC 远端元数据补采。")
    args = parser.parse_args(argv)

    root = Path(args.root).resolve()
    output_dir = Path(args.output_dir).resolve()

    sql_files = load_sql_files(root)
    if not sql_files:
        raise SystemExit(f"未在 {root} 找到 SQL 文件。")

    local_metadata = parse_local_ddl_metadata(sql_files)
    initial_schema = build_schema_map(local_metadata, {})
    statements, produced_schema, referenced_tables, parse_issues = first_pass_statement_scan(sql_files, initial_schema)
    remote_attempts = collect_remote_metadata(
        sorted(referenced_tables | set(produced_schema)),
        local_metadata,
        root,
        enable_remote=not args.disable_remote,
    )
    enrich_metadata_with_inference(local_metadata, referenced_tables | set(produced_schema), produced_schema)
    final_schema = build_schema_map(local_metadata, produced_schema)
    base_tables, produced_tables, _ = build_base_table_sets(statements)
    immediate_edges, lineage_issues = second_pass_lineage_scan(statements, final_schema, produced_tables)

    table_edges_df = build_table_edges(statements, produced_tables)
    table_paths_df = build_transitive_table_paths(table_edges_df, base_tables)
    immediate_df = build_immediate_column_df(immediate_edges)
    recursive_df = build_recursive_column_df(immediate_edges, statements, base_tables)
    base_summary_df = build_base_table_summary(base_tables, table_paths_df, table_edges_df, local_metadata)
    dictionary_df = build_dictionary_df(local_metadata, base_tables, produced_tables)
    table_index_df = build_table_index_df(dictionary_df, table_edges_df, table_paths_df, base_summary_df)
    field_index_df = build_field_index_df(dictionary_df, immediate_df, recursive_df)
    statements_df = build_statement_df(statements)
    assessment_df = build_assessment_df(
        sql_files,
        statements,
        base_tables,
        produced_tables,
        table_edges_df,
        immediate_df,
        recursive_df,
        parse_issues,
        lineage_issues,
        remote_attempts,
    )
    issue_df = build_issue_df(parse_issues, lineage_issues, remote_attempts)
    remote_attempt_df = build_remote_attempt_df(remote_attempts)
    readme_df = build_readme_df(output_dir / "sql_data_map.xlsx", root, enable_remote=not args.disable_remote)

    raw_payload = {
        "generated_at": datetime.now().isoformat(),
        "root": str(root),
        "sql_file_count": len(sql_files),
        "base_tables": sorted(base_tables),
        "produced_tables": sorted(produced_tables),
        "parse_issues": parse_issues,
        "lineage_issues": lineage_issues,
        "remote_attempts": [item.__dict__ for item in remote_attempts],
        "statements": [item.__dict__ for item in statements],
        "immediate_column_lineage": immediate_df.to_dict("records"),
        "base_column_lineage": recursive_df.to_dict("records"),
        "table_edges": table_edges_df.to_dict("records"),
        "table_paths": table_paths_df.to_dict("records"),
    }

    xlsx_path, json_path = write_outputs(
        output_dir=output_dir,
        readme_df=readme_df,
        base_summary_df=base_summary_df,
        table_edges_df=table_edges_df,
        table_paths_df=table_paths_df,
        immediate_df=immediate_df,
        recursive_df=recursive_df,
        table_index_df=table_index_df,
        field_index_df=field_index_df,
        dictionary_df=dictionary_df,
        statements_df=statements_df,
        assessment_df=assessment_df,
        issue_df=issue_df,
        remote_attempt_df=remote_attempt_df,
        raw_payload=raw_payload,
    )

    print(json.dumps({"xlsx": str(xlsx_path), "json": str(json_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
