#!/usr/bin/env python3
"""
面向 current-version.image-update.sql 图片链路的源表数据质量检测脚本。

设计原则：
1. 直接连接腾讯云 DLC，基于实时数据做检测，不依赖静态快照。
2. 不只检查源表本身，还模拟 current-version.image-update.sql 的图片投影逻辑，
   用于区分“源表混入视频数据”和“目标字段语义映射错位”两类问题。
3. 输出 PDF / JSON / Excel，兼顾业务阅读、抽样复核和技术追溯。

重点回答的业务问题：
- 为什么业务会在图片目标表的 shared_folder_path 字段中看到 .mp4？
- 这是源表把视频混进图片集，还是 SQL 将完整文件路径误写进 folder_path 语义字段？
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape


# Proxy bypass: 与仓库内其他 DLC 脚本保持一致
os.environ["NO_PROXY"] = "*"
os.environ["no_proxy"] = "*"
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"]:
    os.environ.pop(_key, None)

import requests

_orig_request = requests.Session.request


def _no_proxy(self, method, url, **kwargs):
    kwargs["proxies"] = {"http": "", "https": ""}
    return _orig_request(self, method, url, **kwargs)


requests.Session.request = _no_proxy

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit("缺少 pandas，请先安装：python -m pip install pandas openpyxl") from exc

try:
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：python -m pip install openpyxl") from exc

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle
except ImportError as exc:
    raise SystemExit("缺少 reportlab，请先安装：python -m pip install reportlab") from exc

try:
    from tencentcloud.common import credential
    from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
    from tencentcloud.common.profile.client_profile import ClientProfile
    from tencentcloud.common.profile.http_profile import HttpProfile
    from tencentcloud.dlc.v20210125 import dlc_client, models
except ImportError as exc:
    raise SystemExit(
        "缺少 tencentcloud-sdk-python，请先安装：python -m pip install tencentcloud-sdk-python"
    ) from exc


REGION = "ap-shanghai"
SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = SCRIPT_DIR.parents[1]
DEFAULT_SQL_PATH = WORKSPACE_DIR / "current-version.image-update.sql"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR
DEFAULT_SOURCE_DB = "data_dws"
DEFAULT_SOURCE_TABLE = "dws_platform_file_resource_label_id"
DEFAULT_TARGET_TABLE = "data_dim.dim_picture_material_data_enriched"
DEFAULT_PDF_NAME = "current-version-image-source-dq-report-{date_tag}.pdf"
DEFAULT_JSON_NAME = "current-version-image-source-dq-report-{date_tag}.json"
DEFAULT_XLSX_NAME = "current-version-image-source-dq-report-{date_tag}.xlsx"

IMAGE_EXTENSIONS = ("jpg", "jpeg", "png", "webp", "bmp", "gif", "tif", "tiff", "heic", "avif")
VIDEO_EXTENSIONS = ("mp4", "mov", "avi", "mkv", "flv", "wmv", "mpeg", "mpg", "m4v", "webm", "ts", "3gp")


@dataclass
class IssueSpec:
    code: str
    title: str
    severity: str
    where_sql: str
    business_meaning: str
    sample_basis: str
    diagnosis_hint: str
    order_sql: str


@dataclass
class IssueResult:
    code: str
    title: str
    severity: str
    row_count: int
    row_rate_pct: float
    business_meaning: str
    sample_basis: str
    diagnosis_hint: str
    diagnosis: str
    samples: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ReportPayload:
    generated_at: str
    sql_path: str
    source_table: str
    target_table: str
    source_total_rows: int
    schema_columns: list[str]
    projection_notes: list[str]
    freshness: list[list[str]]
    health_metrics: list[dict[str, Any]]
    extension_distribution: list[dict[str, Any]]
    issue_results: list[IssueResult]
    key_findings: list[str]
    recommendations: list[str]


def now_tag() -> str:
    return datetime.now().strftime("%Y%m%d")


def sql_quote(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def backtick(name: str) -> str:
    return f"`{name}`"


def to_int(value: Any) -> int:
    try:
        return int(str(value))
    except Exception:
        return 0


def to_float(value: Any) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def value_at(rows: list[list[Any]] | None, row: int = 0, col: int = 0, default: Any = None) -> Any:
    try:
        return rows[row][col]
    except Exception:
        return default


def choose_column(columns: set[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def format_pct(value: float) -> str:
    return f"{value:.2f}%"


def sql_string_expr(col: str | None) -> str:
    if not col:
        return "CAST(NULL AS STRING)"
    return f"CAST({backtick(col)} AS STRING)"


def sql_non_empty_expr(alias: str) -> str:
    return f"`{alias}` IS NOT NULL AND TRIM(CAST(`{alias}` AS STRING)) <> ''"


def ext_list_sql(exts: tuple[str, ...]) -> str:
    return ", ".join(sql_quote(ext) for ext in exts)


class DlcRunner:
    def __init__(self, secret_id: str | None, secret_key: str | None, max_wait: int = 300):
        if not secret_id or not secret_key:
            raise SystemExit("缺少 DLC_USER / DLC_PASSWORD 环境变量。")
        cred = credential.Credential(secret_id, secret_key)
        http_profile = HttpProfile()
        http_profile.endpoint = "dlc.tencentcloudapi.com"
        client_profile = ClientProfile()
        client_profile.httpProfile = http_profile
        self.client = dlc_client.DlcClient(cred, REGION, client_profile)
        self.max_wait = max_wait

    def exec_sql(self, sql: str, db: str) -> list[list[Any]]:
        sql_b64 = base64.b64encode(sql.encode("utf-8")).decode("utf-8")
        task = models.Task()
        task.SparkSQLTask = {"SQL": sql_b64}
        req = models.CreateTaskRequest()
        req.DatabaseName = db
        req.DataEngineName = "SparkSQL"
        req.Task = task
        preview = " ".join(sql.strip().split())[:140]
        print(f"[SQL] {db}: {preview}...", flush=True)
        resp = self.client.CreateTask(req)
        task_id = json.loads(resp.to_json_string()).get("TaskId")
        if not task_id:
            raise RuntimeError("CreateTask 未返回 TaskId")

        elapsed = 0
        while elapsed < self.max_wait:
            time.sleep(3)
            elapsed += 3
            req2 = models.DescribeTaskResultRequest()
            req2.TaskId = str(task_id)
            result = json.loads(self.client.DescribeTaskResult(req2).to_json_string())
            task_info = result.get("TaskInfo", result)
            state = task_info.get("State", "")
            if state == 2:
                result_set = task_info.get("ResultSet", "[]")
                if isinstance(result_set, str):
                    try:
                        return json.loads(base64.b64decode(result_set).decode("utf-8"))
                    except Exception:
                        return json.loads(result_set)
                return result_set or []
            if state == 3:
                raise RuntimeError(task_info.get("OutputMessage", "SQL failed"))
        raise TimeoutError(f"SQL 执行超时，等待了 {self.max_wait}s")

    def describe_table(self, db: str, table: str) -> dict[str, Any]:
        req = models.DescribeTableRequest()
        req.DatabaseName = db
        req.TableName = table
        try:
            resp = self.client.DescribeTable(req)
            data = json.loads(resp.to_json_string())
            return data.get("Table", data)
        except TencentCloudSDKException as exc:
            raise RuntimeError(f"DescribeTable 失败: {db}.{table}: {exc}") from exc


def build_projection_spec(columns: set[str]) -> dict[str, str]:
    file_type_col = choose_column(columns, ["file_type"])
    if not file_type_col:
        raise RuntimeError("源表缺少 file_type，无法按 current-version.image-update.sql 的图片口径过滤。")

    spec = {
        "platform_source_id": choose_column(columns, ["platform_source_id"]),
        "file_id": choose_column(columns, ["file_id"]),
        "title": choose_column(columns, ["title", "image_name"]),
        "platform": choose_column(columns, ["platform"]),
        "file_type": file_type_col,
        "full_path": choose_column(columns, ["full_path", "image_full_path"]),
        "folder_path": choose_column(columns, ["folder_path"]),
        "file_name": choose_column(columns, ["file_name", "image_name"]),
        "data_update_time": choose_column(columns, ["data_update_time"]),
        "image_modify_time": choose_column(columns, ["image_modify_time"]),
        "create_time": choose_column(columns, ["create_time"]),
        "brand": choose_column(columns, ["brand"]),
        "spu": choose_column(columns, ["spu"]),
        "width": choose_column(columns, ["width"]),
        "height": choose_column(columns, ["height"]),
        "picture_size": choose_column(columns, ["picture_size", "file_size", "length"]),
        "picture_wear": choose_column(columns, ["picture_wear"]),
        "picture_type": choose_column(columns, ["picture_type"]),
        "scene": choose_column(columns, ["scene"]),
        "style": choose_column(columns, ["style"]),
        "big_cate": choose_column(columns, ["big_cate"]),
        "mid_cate": choose_column(columns, ["mid_cate"]),
        "product_name": choose_column(columns, ["product_name"]),
        "sub_track": choose_column(columns, ["sub_track"]),
        "gender": choose_column(columns, ["gender"]),
    }
    if not any(spec.get(name) for name in ["full_path", "folder_path", "file_name"]):
        raise RuntimeError("源表缺少 full_path / folder_path / file_name，无法评估图片路径逻辑。")
    return spec


def build_projection_cte(source_db: str, source_table: str, spec: dict[str, str]) -> str:
    alias_selects = []
    for alias, column in spec.items():
        alias_selects.append(f"{sql_string_expr(column)} AS `{alias}`")
    full_path_std = """
        CASE
            WHEN folder_path IS NOT NULL
             AND TRIM(CAST(folder_path AS STRING)) <> ''
             AND file_name IS NOT NULL
             AND TRIM(CAST(file_name AS STRING)) <> ''
            THEN concat(folder_path, '\\\\', file_name)
            ELSE full_path
        END
    """
    file_name_std = f"""
        COALESCE(
            file_name,
            regexp_extract(
                {full_path_std},
                '([^\\\\\\\\/]+)$',
                1
            )
        )
    """
    return f"""
    WITH picture_src AS (
        SELECT
            {", ".join(alias_selects)}
        FROM {source_db}.{source_table}
        WHERE {backtick(spec['file_type'])} = '图片'
    ),
    projected AS (
        SELECT
            *,
            {full_path_std} AS full_path_std,
            {file_name_std} AS file_name_std,
            lower(regexp_extract({full_path_std}, '\\\\.([A-Za-z0-9]{{2,8}})$', 1)) AS path_ext_std,
            lower(regexp_extract({file_name_std}, '\\\\.([A-Za-z0-9]{{2,8}})$', 1)) AS file_name_ext_std,
            COALESCE(data_update_time, image_modify_time) AS image_update_time,
            CASE
                WHEN folder_path IS NOT NULL
                 AND TRIM(CAST(folder_path AS STRING)) <> ''
                 AND file_name IS NOT NULL
                 AND TRIM(CAST(file_name AS STRING)) <> ''
                THEN 1 ELSE 0
            END AS use_concat_path_flag,
            CASE
                WHEN (folder_path IS NULL OR TRIM(CAST(folder_path AS STRING)) = '')
                  OR (file_name IS NULL OR TRIM(CAST(file_name AS STRING)) = '')
                THEN 1 ELSE 0
            END AS fallback_to_full_path_flag,
            CASE
                WHEN full_path IS NOT NULL
                 AND TRIM(CAST(full_path AS STRING)) <> ''
                 AND file_name IS NOT NULL
                 AND TRIM(CAST(file_name AS STRING)) <> ''
                 AND lower(regexp_extract(full_path, '([^\\\\\\\\/]+)$', 1)) <> lower(file_name)
                THEN 1 ELSE 0
            END AS basename_conflict_flag,
            CASE
                WHEN {full_path_std} RLIKE '(?i)\\\\.[a-z0-9]{{2,8}}$' THEN 1 ELSE 0
            END AS target_folder_semantic_suffix_flag,
            CASE
                WHEN lower(regexp_extract({full_path_std}, '\\\\.([A-Za-z0-9]{{2,8}})$', 1)) IN ({ext_list_sql(IMAGE_EXTENSIONS)})
                THEN 'image'
                WHEN lower(regexp_extract({full_path_std}, '\\\\.([A-Za-z0-9]{{2,8}})$', 1)) IN ({ext_list_sql(VIDEO_EXTENSIONS)})
                THEN 'video'
                WHEN lower(regexp_extract({full_path_std}, '\\\\.([A-Za-z0-9]{{2,8}})$', 1)) = ''
                THEN 'none'
                ELSE 'other'
            END AS path_suffix_family
        FROM picture_src
    )
    """


def fetch_schema_columns(runner: DlcRunner, source_db: str, source_table: str) -> list[str]:
    desc = runner.describe_table(source_db, source_table)
    return [col["Name"] for col in desc.get("Columns", [])]


def fetch_health_metrics(runner: DlcRunner, source_db: str, projection_cte: str) -> tuple[int, list[dict[str, Any]]]:
    sql = f"""
    {projection_cte}
    SELECT
        COUNT(*) AS total_rows,
        SUM(CASE WHEN {sql_non_empty_expr('full_path_std')} THEN 1 ELSE 0 END) AS filled_full_path_std_rows,
        SUM(CASE WHEN use_concat_path_flag = 1 THEN 1 ELSE 0 END) AS use_concat_path_rows,
        SUM(CASE WHEN fallback_to_full_path_flag = 1 THEN 1 ELSE 0 END) AS fallback_to_full_path_rows,
        SUM(CASE WHEN basename_conflict_flag = 1 THEN 1 ELSE 0 END) AS basename_conflict_rows,
        SUM(CASE WHEN {sql_non_empty_expr('image_update_time')} THEN 1 ELSE 0 END) AS image_update_time_rows,
        SUM(CASE WHEN target_folder_semantic_suffix_flag = 1 THEN 1 ELSE 0 END) AS target_folder_semantic_suffix_rows,
        COUNT(DISTINCT file_id) AS distinct_file_id_rows,
        COUNT(DISTINCT platform_source_id) AS distinct_platform_source_id_rows
    FROM projected
    """
    res = runner.exec_sql(sql, source_db)
    total_rows = to_int(value_at(res, 0, 0, 0))
    metric_defs = [
        ("full_path_std 非空", 1, "图片投影后是否能得到标准化完整路径"),
        ("使用 folder_path + file_name 拼接", 2, "结构化路径字段可用，能直接反映目录与文件名"),
        ("回退到 full_path", 3, "folder_path/file_name 不全，只能依赖原始 full_path"),
        ("full_path 与 file_name 基名冲突", 4, "源表同一条记录存在路径/文件名不一致"),
        ("image_update_time 非空", 5, "最终图片更新时间字段可稳定回填"),
        ("目标 folder 字段将带后缀", 6, "当前 SQL 会把完整路径写入 image_share_folder_path，语义上存在错位"),
    ]
    metrics: list[dict[str, Any]] = []
    for label, idx, meaning in metric_defs:
        value = to_int(value_at(res, 0, idx, 0))
        rate = round(value * 100.0 / total_rows, 2) if total_rows else 0.0
        metrics.append({"metric": label, "row_count": value, "row_rate_pct": rate, "meaning": meaning})

    distinct_file_id = to_int(value_at(res, 0, 7, 0))
    distinct_platform_source_id = to_int(value_at(res, 0, 8, 0))
    metrics.append(
        {
            "metric": "file_id 去重后行数",
            "row_count": distinct_file_id,
            "row_rate_pct": round(distinct_file_id * 100.0 / total_rows, 2) if total_rows else 0.0,
            "meaning": "用于判断 file_id 是否存在重复污染",
        }
    )
    metrics.append(
        {
            "metric": "platform_source_id 去重后行数",
            "row_count": distinct_platform_source_id,
            "row_rate_pct": round(distinct_platform_source_id * 100.0 / total_rows, 2) if total_rows else 0.0,
            "meaning": "用于判断平台素材主键是否存在重复污染",
        }
    )
    return total_rows, metrics


def fetch_extension_distribution(runner: DlcRunner, source_db: str, projection_cte: str) -> list[dict[str, Any]]:
    sql = f"""
    {projection_cte}
    SELECT
        COALESCE(path_ext_std, '') AS path_ext_std,
        path_suffix_family,
        COUNT(*) AS row_count
    FROM projected
    GROUP BY COALESCE(path_ext_std, ''), path_suffix_family
    ORDER BY row_count DESC, path_ext_std
    LIMIT 30
    """
    rows = runner.exec_sql(sql, source_db)
    return [
        {
            "path_ext_std": row[0],
            "path_suffix_family": row[1],
            "row_count": to_int(row[2]),
        }
        for row in rows
    ]


def fetch_freshness(runner: DlcRunner, source_db: str, projection_cte: str) -> list[list[str]]:
    freshness: list[list[str]] = []
    for col in ["create_time", "data_update_time", "image_modify_time", "image_update_time"]:
        sql = f"""
        {projection_cte}
        SELECT
            MIN(CAST({backtick(col)} AS STRING)),
            MAX(CAST({backtick(col)} AS STRING))
        FROM projected
        WHERE {sql_non_empty_expr(col)}
        """
        rows = runner.exec_sql(sql, source_db)
        freshness.append([col, str(value_at(rows, 0, 0, "NA")), str(value_at(rows, 0, 1, "NA"))])
    return freshness


def build_issue_specs() -> list[IssueSpec]:
    video_exts = ext_list_sql(VIDEO_EXTENSIONS)
    image_and_video_exts = ext_list_sql(IMAGE_EXTENSIONS + VIDEO_EXTENSIONS)
    return [
        IssueSpec(
            code="VIDEO_SUFFIX_IN_IMAGE_SET",
            title="图片集混入视频后缀",
            severity="high",
            where_sql=f"path_ext_std IN ({video_exts})",
            business_meaning="这类记录会直接把 .mp4/.mov 等后缀带入目标图片表，是业务抱怨最强的一类硬证据。",
            sample_basis="按 image_update_time 倒序抽样，优先看最近进入链路的数据。",
            diagnosis_hint="若命中，说明源表 file_type='图片' 子集内已经混入视频资源或视频路径。",
            order_sql="image_update_time DESC, file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="TARGET_FOLDER_FIELD_SUFFIX_RISK",
            title="目标 shared_folder_path 字段语义错位",
            severity="high",
            where_sql="target_folder_semantic_suffix_flag = 1",
            business_meaning="当前 SQL 会把 full_path_std 填到 image_share_folder_path；只要路径带文件名，目标字段就会出现后缀，业务会误以为 folder_path 不纯。",
            sample_basis="按 image_update_time 倒序抽样，观察目标字段将出现怎样的后缀表现。",
            diagnosis_hint="这类问题不等同于源表污染，更像目标字段语义与 SQL 赋值口径不一致。",
            order_sql="image_update_time DESC, file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="UNKNOWN_SUFFIX_IN_IMAGE_SET",
            title="图片集出现非图片非视频后缀",
            severity="medium",
            where_sql=f"path_ext_std <> '' AND path_ext_std NOT IN ({image_and_video_exts})",
            business_meaning="这类记录不一定是视频，但已经偏离图片链路的主流文件格式，需要人工判断是否为压缩包、中间件文件或脏路径。",
            sample_basis="按 image_update_time 倒序抽样，优先验证近期新增异常格式。",
            diagnosis_hint="常见于脏数据、目录名误写、临时文件或外部系统落盘异常。",
            order_sql="image_update_time DESC, file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="FALLBACK_TO_FULL_PATH",
            title="结构化路径字段缺失，只能回退 full_path",
            severity="medium",
            where_sql="fallback_to_full_path_flag = 1",
            business_meaning="folder_path/file_name 不完整时，SQL 无法严格区分目录和文件名，会削弱路径治理和 folder 级分析。",
            sample_basis="按 image_update_time 倒序抽样，定位最近缺结构化路径的来源记录。",
            diagnosis_hint="这类问题会放大 shared_folder_path 的字段语义歧义。",
            order_sql="image_update_time DESC, file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="BASENAME_CONFLICT",
            title="full_path 与 file_name 基名不一致",
            severity="medium",
            where_sql="basename_conflict_flag = 1",
            business_meaning="同一条记录的 full_path 与 file_name 指向不同文件名时，路径标准化结果会不稳定，业务难以信任目录和文件映射。",
            sample_basis="按 image_update_time 倒序抽样，查看路径与文件名冲突的真实样本。",
            diagnosis_hint="若命中，优先排查上游同步是否发生覆盖、截断或二次加工。",
            order_sql="image_update_time DESC, file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="MISSING_IMAGE_UPDATE_TIME",
            title="图片更新时间缺失",
            severity="medium",
            where_sql="image_update_time IS NULL OR TRIM(CAST(image_update_time AS STRING)) = ''",
            business_meaning="最终图片表的 image_modify_time 为空，会影响业务追踪近期变更和重跑判断。",
            sample_basis="按 file_id 倒序抽样，快速复核更新时间缺失的来源分布。",
            diagnosis_hint="若命中较高，说明 data_update_time / image_modify_time 两个来源字段都不稳定。",
            order_sql="file_id DESC, platform_source_id DESC",
        ),
        IssueSpec(
            code="MISSING_STANDARD_PATH",
            title="标准化完整路径缺失",
            severity="high",
            where_sql="full_path_std IS NULL OR TRIM(CAST(full_path_std AS STRING)) = ''",
            business_meaning="连标准化后的路径都缺失，意味着目标表核心定位字段会直接空掉，业务无法回溯素材文件。",
            sample_basis="按 file_id 倒序抽样，快速锁定完全缺路径的数据来源。",
            diagnosis_hint="这类问题比字段语义错位更严重，属于关键主链路信息缺失。",
            order_sql="file_id DESC, platform_source_id DESC",
        ),
    ]


def collect_schema_gaps(spec: dict[str, str]) -> list[str]:
    gaps: list[str] = []
    if not spec.get("folder_path"):
        gaps.append("folder_path")
    if not spec.get("data_update_time"):
        gaps.append("data_update_time")
    if not spec.get("image_modify_time"):
        gaps.append("image_modify_time")
    return gaps


def diagnose_issue(spec: IssueSpec, row_count: int, total_rows: int, schema_gaps: set[str]) -> str:
    rate = round(row_count * 100.0 / total_rows, 2) if total_rows else 0.0
    if row_count == 0:
        return "本次实时检测未命中该问题。"
    if spec.code == "FALLBACK_TO_FULL_PATH" and "folder_path" in schema_gaps:
        return (
            f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。实时源表 schema 本身就没有 folder_path，"
            "所以 current-version.image-update.sql 在当前实库上只能 100% 回退 full_path。"
        )
    if spec.code == "MISSING_IMAGE_UPDATE_TIME" and {"data_update_time", "image_modify_time"} <= schema_gaps:
        return (
            f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。实时源表 schema 同时缺少 data_update_time 和 image_modify_time，"
            "因此投影后的 image_update_time 必然全空。"
        )
    if spec.code == "VIDEO_SUFFIX_IN_IMAGE_SET":
        return f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。这可以直接佐证“图片链路混入视频路径”的业务直觉。"
    if spec.code == "TARGET_FOLDER_FIELD_SUFFIX_RISK":
        return (
            f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。这说明即便源表本身是图片，"
            "当前 SQL 也会把完整文件路径写进 folder 语义字段，业务看到后缀并不意外。"
        )
    if spec.code == "FALLBACK_TO_FULL_PATH":
        return f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。folder_path/file_name 的结构化治理还不稳定。"
    if spec.code == "MISSING_IMAGE_UPDATE_TIME":
        return f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。最终 image_modify_time 的可用性存在缺口。"
    return f"命中 {row_count} 行，占图片子集 {format_pct(rate)}。{spec.diagnosis_hint}"


def fetch_issue_results(
    runner: DlcRunner,
    source_db: str,
    projection_cte: str,
    issue_specs: list[IssueSpec],
    total_rows: int,
    sample_limit: int,
    schema_gaps: set[str],
) -> list[IssueResult]:
    issue_results: list[IssueResult] = []
    for spec in issue_specs:
        count_sql = f"""
        {projection_cte}
        SELECT COUNT(*)
        FROM projected
        WHERE {spec.where_sql}
        """
        row_count = to_int(value_at(runner.exec_sql(count_sql, source_db), 0, 0, 0))
        row_rate_pct = round(row_count * 100.0 / total_rows, 2) if total_rows else 0.0

        sample_sql = f"""
        {projection_cte}
        SELECT
            platform_source_id,
            file_id,
            platform,
            title,
            file_type,
            full_path,
            folder_path,
            file_name,
            full_path_std,
            file_name_std,
            path_ext_std,
            image_update_time,
            image_modify_time,
            data_update_time,
            brand,
            spu
        FROM projected
        WHERE {spec.where_sql}
        ORDER BY {spec.order_sql}
        LIMIT {sample_limit}
        """
        sample_rows = runner.exec_sql(sample_sql, source_db)
        samples = [
            {
                "platform_source_id": row[0],
                "file_id": row[1],
                "platform": row[2],
                "title": row[3],
                "file_type": row[4],
                "full_path": row[5],
                "folder_path": row[6],
                "file_name": row[7],
                "full_path_std": row[8],
                "file_name_std": row[9],
                "path_ext_std": row[10],
                "image_update_time": row[11],
                "image_modify_time": row[12],
                "data_update_time": row[13],
                "brand": row[14],
                "spu": row[15],
            }
            for row in sample_rows
        ]
        issue_results.append(
            IssueResult(
                code=spec.code,
                title=spec.title,
                severity=spec.severity,
                row_count=row_count,
                row_rate_pct=row_rate_pct,
                business_meaning=spec.business_meaning,
                sample_basis=spec.sample_basis,
                diagnosis_hint=spec.diagnosis_hint,
                diagnosis=diagnose_issue(spec, row_count, total_rows, schema_gaps),
                samples=samples,
            )
        )
    issue_results.sort(key=lambda item: ({"high": 0, "medium": 1, "low": 2}.get(item.severity, 9), -item.row_count, item.code))
    return issue_results


def build_key_findings(
    total_rows: int,
    metrics: list[dict[str, Any]],
    issues: list[IssueResult],
    schema_gaps: set[str],
) -> list[str]:
    findings: list[str] = []
    issue_map = {item.code: item for item in issues}
    metric_map = {item["metric"]: item for item in metrics}

    if schema_gaps:
        findings.append(
            f"实时源表 schema 缺少 {', '.join(sorted(schema_gaps))}。这意味着 current-version.image-update.sql 对这些字段的依赖，在当前实库上并未真正具备。"
        )

    video_issue = issue_map.get("VIDEO_SUFFIX_IN_IMAGE_SET")
    if video_issue and video_issue.row_count > 0:
        findings.append(
            f"图片子集里直接命中 {video_issue.row_count} 行视频后缀路径，占比 {format_pct(video_issue.row_rate_pct)}，可以直接支撑“图片表混入视频格式”的判断。"
        )
    else:
        findings.append("本次实时检测未命中视频后缀路径，业务看到 .mp4 更可能来自其他链路、历史批次或当前样本外数据。")

    semantic_metric = metric_map.get("目标 folder 字段将带后缀")
    if semantic_metric and semantic_metric["row_count"] > 0:
        findings.append(
            f"有 {semantic_metric['row_count']} 行在当前 SQL 下会把完整文件路径写入 image_share_folder_path，占比 {format_pct(semantic_metric['row_rate_pct'])}。"
            " 这说明 shared_folder_path 出现后缀不只是数据脏，更有字段语义映射问题。"
        )

    fallback_metric = metric_map.get("回退到 full_path")
    if fallback_metric and fallback_metric["row_rate_pct"] >= 20:
        findings.append(
            f"有 {fallback_metric['row_count']} 行缺少 folder_path 或 file_name，只能回退 full_path，占比 {format_pct(fallback_metric['row_rate_pct'])}。"
            " 这会削弱基于目录的业务分析和核查。"
        )

    update_metric = metric_map.get("image_update_time 非空")
    if update_metric and update_metric["row_rate_pct"] < 95:
        miss_pct = round(100.0 - update_metric["row_rate_pct"], 2)
        findings.append(
            f"最终图片更新时间覆盖率只有 {format_pct(update_metric['row_rate_pct'])}，仍有 {format_pct(miss_pct)} 的记录无法稳定用于增量追踪。"
        )

    full_path_metric = metric_map.get("full_path_std 非空")
    if full_path_metric and full_path_metric["row_rate_pct"] < 99:
        findings.append(
            f"标准化完整路径覆盖率仅 {format_pct(full_path_metric['row_rate_pct'])}，说明仍有记录无法回溯到物理文件。"
        )

    if total_rows == 0:
        findings.append("当前 file_type='图片' 子集为空，本次报告无法对图片链路做有效判断。")
    return findings


def build_recommendations(issues: list[IssueResult], schema_gaps: set[str]) -> list[str]:
    issue_map = {item.code: item for item in issues}
    recs = [
        "若业务字段语义上要求目录路径，建议把目标表 image_share_folder_path 改为接 folder_path，而把完整路径单独保留在 full_path。",
        "在 current-version.image-update.sql 上游增加图片后缀白名单校验，至少拦截 .mp4/.mov 等视频后缀进入图片链路。",
        "对 folder_path、file_name、full_path 建立一致性校验，避免路径与文件名互相矛盾。",
        "把本脚本的异常样本抽样结果交给业务复核，先确认是否为真实图片素材，再决定是清洗还是改口径。",
    ]
    if schema_gaps:
        recs.append(f"先补齐实时源表 schema：{', '.join(sorted(schema_gaps))}。否则 SQL 中对应逻辑只能停留在设计稿，无法在实库稳定生效。")
    if issue_map.get("VIDEO_SUFFIX_IN_IMAGE_SET") and issue_map["VIDEO_SUFFIX_IN_IMAGE_SET"].row_count == 0:
        recs.append("若业务侧仍持续观察到 .mp4，请优先核查目标表历史分区、其他批次脚本或非当前 SQL 写入链路。")
    return recs


def build_issue_summary_frame(issue_results: list[IssueResult]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "issue_code": item.code,
                "issue_title": item.title,
                "severity": item.severity,
                "row_count": item.row_count,
                "row_rate_pct": item.row_rate_pct,
                "business_meaning": item.business_meaning,
                "sample_basis": item.sample_basis,
                "diagnosis": item.diagnosis,
            }
            for item in issue_results
        ]
    )


def build_issue_samples_frame(issue_results: list[IssueResult]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for item in issue_results:
        for sample in item.samples:
            records.append({"issue_code": item.code, "issue_title": item.title, **sample})
    return pd.DataFrame(records)


def autofit_excel(writer, sheet_to_frame: dict[str, pd.DataFrame]) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="top", wrap_text=True)
    for sheet_name, frame in sheet_to_frame.items():
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
        for idx, column in enumerate(frame.columns, start=1):
            max_len = len(str(column))
            if not frame.empty:
                sample_series = frame[column].astype(str).head(200)
                max_len = max(max_len, sample_series.map(len).max())
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 2, 60)


def build_pdf(payload: ReportPayload, pdf_path: Path) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_cn",
        parent=styles["Title"],
        fontName="STSong-Light",
        fontSize=18,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17375E"),
        spaceAfter=10,
    )
    heading_style = ParagraphStyle(
        "heading_cn",
        parent=styles["Heading2"],
        fontName="STSong-Light",
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#1F4E78"),
        spaceBefore=8,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "body_cn",
        parent=styles["BodyText"],
        fontName="STSong-Light",
        fontSize=9.5,
        leading=14,
        spaceAfter=4,
    )
    small_style = ParagraphStyle(
        "small_cn",
        parent=body_style,
        fontSize=8.5,
        leading=12,
    )

    story: list[Any] = []
    story.append(Paragraph("current-version.image-update.sql 图片源表数据质量报告", title_style))
    story.append(
        Paragraph(
            escape(
                f"生成时间：{payload.generated_at}；源表：{payload.source_table}；目标表：{payload.target_table}；图片子集总量：{payload.source_total_rows}"
            ),
            body_style,
        )
    )
    story.append(Paragraph(escape(f"SQL 文件：{payload.sql_path}"), small_style))
    story.append(Spacer(1, 4 * mm))

    story.append(Paragraph("一、检测立场", heading_style))
    for note in payload.projection_notes:
        story.append(Paragraph(f"• {escape(note)}", body_style))

    story.append(Paragraph("二、关键结论", heading_style))
    for finding in payload.key_findings:
        story.append(Paragraph(f"• {escape(finding)}", body_style))

    metric_rows = [["指标", "行数", "占比", "业务含义"]]
    for metric in payload.health_metrics:
        metric_rows.append(
            [
                metric["metric"],
                str(metric["row_count"]),
                format_pct(float(metric["row_rate_pct"])),
                metric["meaning"],
            ]
        )
    metric_table = LongTable(metric_rows, repeatRows=1, colWidths=[48 * mm, 22 * mm, 20 * mm, 88 * mm])
    metric_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("LEADING", (0, 0), (-1, -1), 11),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9FBAD0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(Paragraph("三、核心体检指标", heading_style))
    story.append(metric_table)
    story.append(Spacer(1, 4 * mm))

    ext_rows = [["后缀", "分类", "行数"]]
    for row in payload.extension_distribution[:15]:
        ext_rows.append([row["path_ext_std"] or "(空)", row["path_suffix_family"], str(row["row_count"])])
    ext_table = LongTable(ext_rows, repeatRows=1, colWidths=[40 * mm, 35 * mm, 25 * mm])
    ext_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2F0D9")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A8C68F")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(Paragraph("四、路径后缀分布", heading_style))
    story.append(ext_table)
    story.append(Spacer(1, 4 * mm))

    issue_rows = [["问题", "严重级别", "行数", "占比", "诊断"]]
    for item in payload.issue_results:
        issue_rows.append(
            [item.title, item.severity, str(item.row_count), format_pct(item.row_rate_pct), item.diagnosis]
        )
    issue_table = LongTable(issue_rows, repeatRows=1, colWidths=[34 * mm, 18 * mm, 18 * mm, 18 * mm, 92 * mm])
    issue_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FCE4D6")),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E6B8A2")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(Paragraph("五、异常问题汇总", heading_style))
    story.append(issue_table)
    story.append(PageBreak())

    story.append(Paragraph("六、抽样样本", heading_style))
    for item in payload.issue_results:
        if not item.samples:
            continue
        story.append(Paragraph(f"{item.title}（{item.code}）", body_style))
        story.append(Paragraph(f"抽样依据：{escape(item.sample_basis)}", small_style))
        sample_rows = [["file_id", "platform", "path_ext", "file_name_std", "full_path_std", "image_update_time"]]
        for sample in item.samples[:5]:
            sample_rows.append(
                [
                    str(sample.get("file_id", "")),
                    str(sample.get("platform", "")),
                    str(sample.get("path_ext_std", "")),
                    str(sample.get("file_name_std", ""))[:32],
                    str(sample.get("full_path_std", ""))[:58],
                    str(sample.get("image_update_time", "")),
                ]
            )
        sample_table = LongTable(sample_rows, repeatRows=1, colWidths=[22 * mm, 20 * mm, 16 * mm, 34 * mm, 72 * mm, 28 * mm])
        sample_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("LEADING", (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9C9C9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(sample_table)
        story.append(Spacer(1, 3 * mm))

    story.append(Paragraph("七、建议动作", heading_style))
    for rec in payload.recommendations:
        story.append(Paragraph(f"• {escape(rec)}", body_style))

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    doc.build(story)


def write_outputs(payload: ReportPayload, output_dir: Path, date_tag: str) -> tuple[Path, Path, Path]:
    pdf_path = output_dir / DEFAULT_PDF_NAME.format(date_tag=date_tag)
    json_path = output_dir / DEFAULT_JSON_NAME.format(date_tag=date_tag)
    xlsx_path = output_dir / DEFAULT_XLSX_NAME.format(date_tag=date_tag)

    json_path.write_text(
        json.dumps(
            {
                **asdict(payload),
                "issue_results": [asdict(item) for item in payload.issue_results],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    issue_summary = build_issue_summary_frame(payload.issue_results)
    issue_samples = build_issue_samples_frame(payload.issue_results)
    health_metrics = pd.DataFrame(payload.health_metrics)
    extension_distribution = pd.DataFrame(payload.extension_distribution)
    freshness = pd.DataFrame(payload.freshness, columns=["field_name", "min_value", "max_value"])
    projection_notes = pd.DataFrame({"projection_note": payload.projection_notes})
    recommendations = pd.DataFrame({"recommendation": payload.recommendations})

    sheets = {
        "issue_summary": issue_summary,
        "issue_samples": issue_samples,
        "health_metrics": health_metrics,
        "extension_distribution": extension_distribution,
        "freshness": freshness,
        "projection_notes": projection_notes,
        "recommendations": recommendations,
    }
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
        autofit_excel(writer, sheets)

    build_pdf(payload, pdf_path)
    return pdf_path, json_path, xlsx_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="检测 current-version.image-update.sql 的图片源表数据质量")
    parser.add_argument("--sql-path", default=str(DEFAULT_SQL_PATH), help="current-version.image-update.sql 路径")
    parser.add_argument("--source-db", default=DEFAULT_SOURCE_DB, help="源库名，默认 data_dws")
    parser.add_argument("--source-table", default=DEFAULT_SOURCE_TABLE, help="源表名，默认 dws_platform_file_resource_label_id")
    parser.add_argument("--target-table", default=DEFAULT_TARGET_TABLE, help="目标表名，仅用于报告展示")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--sample-limit", type=int, default=20, help="每类问题抽样条数，默认 20")
    parser.add_argument("--max-wait", type=int, default=300, help="单条 SQL 最长等待秒数")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    date_tag = now_tag()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    runner = DlcRunner(os.environ.get("DLC_USER"), os.environ.get("DLC_PASSWORD"), max_wait=args.max_wait)
    schema_columns = fetch_schema_columns(runner, args.source_db, args.source_table)
    projection_spec = build_projection_spec(set(schema_columns))
    schema_gaps = set(collect_schema_gaps(projection_spec))
    projection_cte = build_projection_cte(args.source_db, args.source_table, projection_spec)

    total_rows, health_metrics = fetch_health_metrics(runner, args.source_db, projection_cte)
    extension_distribution = fetch_extension_distribution(runner, args.source_db, projection_cte)
    freshness = fetch_freshness(runner, args.source_db, projection_cte)
    issue_results = fetch_issue_results(
        runner,
        args.source_db,
        projection_cte,
        build_issue_specs(),
        total_rows,
        args.sample_limit,
        schema_gaps,
    )

    projection_notes = [
        "本报告不直接扫目标表，而是先还原 current-version.image-update.sql 的图片投影逻辑，再对投影结果做质检。",
        "若 shared_folder_path 中出现后缀，可能来自两类原因：一是图片子集真的混入了视频或异常文件；二是 SQL 把完整 full_path 写进了 folder 语义字段。",
        "抽样均按近期 image_update_time 优先，目的是先验证当前正在影响业务使用的数据，而不是只看历史脏样本。",
        "本报告默认图片子集定义为源表 file_type = '图片'，与 current-version.image-update.sql 保持一致。",
    ]
    if schema_gaps:
        projection_notes.append(
            f"实时源表 schema 缺少 {', '.join(sorted(schema_gaps))}；相关口径在当前实库环境下会表现为全量回退或全量为空。"
        )
    key_findings = build_key_findings(total_rows, health_metrics, issue_results, schema_gaps)
    recommendations = build_recommendations(issue_results, schema_gaps)
    payload = ReportPayload(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        sql_path=str(Path(args.sql_path).resolve()),
        source_table=f"{args.source_db}.{args.source_table}",
        target_table=args.target_table,
        source_total_rows=total_rows,
        schema_columns=schema_columns,
        projection_notes=projection_notes,
        freshness=freshness,
        health_metrics=health_metrics,
        extension_distribution=extension_distribution,
        issue_results=issue_results,
        key_findings=key_findings,
        recommendations=recommendations,
    )
    pdf_path, json_path, xlsx_path = write_outputs(payload, output_dir, date_tag)

    print(f"[OK] PDF:  {pdf_path}")
    print(f"[OK] JSON: {json_path}")
    print(f"[OK] XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
